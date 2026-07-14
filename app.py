import streamlit as st
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
import fiona
import io
import datetime
import re

# Enable KML support in GeoPandas
fiona.drvsupport.supported_drivers['KML'] = 'rw'
fiona.drvsupport.supported_drivers['LIBKML'] = 'rw'

# --- 1. CONFIGURATION & UI SETUP ---
st.set_page_config(page_title="Territory Audit Engine", layout="wide")

st.title("Congregation Territory Analysis Engine")
st.markdown("Upload your territories KML map to generate a complete, filtered address database & analysis.")

st.sidebar.header("Step 1: Configuration")
congregation_name = st.sidebar.text_input("Congregation Name (No Spaces)", "ExampleCongregation")
selected_county = st.sidebar.selectbox("Select County Data", ["Milwaukee"]) 
goal_range = st.sidebar.selectbox("Goal # of Addresses Per Territory", 
                                  ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"])

st.header("Step 2: Upload Territory Map")
uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

# Parse Goal Range
MIN_GOAL, MAX_GOAL = [int(x) for x in goal_range.split("-")]

# --- 2. DATA LOADING & CACHING ---
@st.cache_data
def load_county_data(county_name):
    if county_name == "Milwaukee":
        file_path = "zip://data/Milwaukee_Datapoints07072026.zip"
        try:
            gdf = gpd.read_file(file_path)
            return gdf
        except Exception as e:
            st.error(f"Error loading county shapefile. Check that the zip is in the /data/ folder. Error: {e}")
            return None
    return None

# --- NATURAL SORTING HELPER ---
def natural_keys(text):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

# --- ADDRESS BUILDER ---
def build_addresses(row):
    house = str(row['HouseNo']).replace('.0', '').strip() if pd.notna(row['HouseNo']) and str(row['HouseNo']).lower() != "nan" else ""
    house_sx = str(row['HouseSx']).strip() if pd.notna(row['HouseSx']) and str(row['HouseSx']).lower() != "nan" else ""
    direction = str(row['Dir']).strip() if pd.notna(row['Dir']) and str(row['Dir']).lower() != "nan" else ""
    street = str(row['Street']).strip() if pd.notna(row['Street']) and str(row['Street']).lower() != "nan" else ""
    st_type = str(row['StType']).strip() if pd.notna(row['StType']) and str(row['StType']).lower() != "nan" else ""
    muni = str(row['Muni']).strip() if pd.notna(row['Muni']) and str(row['Muni']).lower() != "nan" else ""
    zip_c = str(row['Zip_Code']).strip() if pd.notna(row['Zip_Code']) and str(row['Zip_Code']).lower() != "nan" else ""
    
    unit_val = str(row['Unit']).strip() if pd.notna(row['Unit']) and str(row['Unit']).lower() != "nan" else ""
    unit_str = f" Apt {unit_val}" if unit_val else ""

    full_house_num = f"{house}{house_sx}"
    street_parts = [direction, street, st_type]
    full_street = " ".join([p for p in street_parts if p])

    base_addr_line = f"{full_house_num} {full_street}".strip()
    base_addr = f"{base_addr_line}, {muni}, WI {zip_c}".replace(" ,", ",").strip(" ,")

    mailable_addr_line = f"{base_addr_line}{unit_str}".strip()
    mailable_addr = f"{mailable_addr_line}, {muni}, WI {zip_c}".replace(" ,", ",").strip(" ,")

    return pd.Series([base_addr, mailable_addr])


# --- 3. EXCEL GENERATION ENGINE ---
def generate_excel_report(joined_gdf, kml_gdf, min_goal, max_goal, cong_name):
    output = io.BytesIO()
    
    joined_gdf['Zip_Code'] = joined_gdf['Zip_Code'].astype(str).str[:5]
    joined_gdf[['Base_Address', 'Mailable_Address']] = joined_gdf.apply(build_addresses, axis=1)
    
    invalid_statuses = [
        'Undeveloped', 'Parking Lot', 'ROW', 'Park or Recreational Facility',
        'Undeveloped Outlot', 'Sliver or Remnant', 'Non Addressable Assoc with Adj Parcel'
    ]
    excluded_gdf = joined_gdf[joined_gdf['Addr_Statu'].isin(invalid_statuses)].copy()
    valid_gdf = joined_gdf[~joined_gdf['Addr_Statu'].isin(invalid_statuses)].copy()

    unique_territories = valid_gdf['Territory_Name'].unique().tolist()
    unique_territories.sort(key=natural_keys)
    valid_gdf['Territory_Name'] = pd.Categorical(valid_gdf['Territory_Name'], categories=unique_territories, ordered=True)
    
    if not excluded_gdf.empty:
        excluded_unique = excluded_gdf['Territory_Name'].unique().tolist()
        excluded_unique.sort(key=natural_keys)
        excluded_gdf['Territory_Name'] = pd.Categorical(excluded_gdf['Territory_Name'], categories=excluded_unique, ordered=True)

    counts_df = valid_gdf.groupby('Territory_Name', observed=True).size().reset_index(name='Total_Addresses')
    counts_df = counts_df[counts_df['Total_Addresses'] > 0]
    
    def get_category(count):
        if count < min_goal: return "Undersized"
        elif min_goal <= count <= max_goal: return "Ideal"
        else: return "Oversized"
        
    counts_df['Category'] = counts_df['Total_Addresses'].apply(get_category)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- TAB 1: DASHBOARD ---
        total_territories = len(counts_df)
        total_addresses = counts_df['Total_Addresses'].sum()
        largest_terr = counts_df.loc[counts_df['Total_Addresses'].idxmax()] if total_territories > 0 else None
        smallest_terr = counts_df.loc[counts_df['Total_Addresses'].idxmin()] if total_territories > 0 else None
        ideal_pct = (len(counts_df[counts_df['Category'] == 'Ideal']) / total_territories) * 100 if total_territories > 0 else 0
        
        largest_name = largest_terr['Territory_Name'] if largest_terr is not None else ""
        largest_count = largest_terr['Total_Addresses'] if largest_terr is not None else 0
        smallest_name = smallest_terr['Territory_Name'] if smallest_terr is not None else ""
        smallest_count = smallest_terr['Total_Addresses'] if smallest_terr is not None else 0

        dashboard_top = [
            [f"Territory Analysis: {cong_name}"],
            [f"Generated {datetime.datetime.now().strftime('%B %Y')} by Territory Analysis Engine."],
            [""],
            [f"Total Territories: {total_territories}"],
            [f"Total Valid Addresses: {total_addresses}"],
            [f"Excluded Addresses (See Tab 6): {len(excluded_gdf)}"],
            [f"The largest territory has {largest_count} addresses in it ({largest_name})."],
            [f"The smallest territory has {smallest_count} addresses in it ({smallest_name})."],
            [""],
            [f"Goal Range: {min_goal}-{max_goal}"],
            [""] 
        ]
        pd.DataFrame(dashboard_top).to_excel(writer, sheet_name="Dashboard", index=False, header=False)
        
        ranges = ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"]
        distribution = []
        for r in ranges:
            rmin, rmax = [int(x) for x in r.split("-")]
            count = len(counts_df[(counts_df['Total_Addresses'] >= rmin) & (counts_df['Total_Addresses'] <= rmax)])
            cat = "Ideal" if rmin == min_goal else ("Undersized" if rmax <= min_goal else "Oversized")
            distribution.append([cat, r, count])
            
        pd.DataFrame(distribution, columns=["Category", "Range", "Count"]).to_excel(writer, sheet_name="Dashboard", startrow=11, index=False)

        ws1 = writer.sheets['Dashboard']
        ws1.column_dimensions['A'].width = 15
        
        ws1['A1'].font = Font(size=20, bold=True)
        ws1['A2'].hyperlink = "http://www.territoryanalysis.com/"
        ws1['A2'].font = Font(color="0563C1", underline="single")
        
        bold_inline = InlineFont(b=True)
        ws1['A11'].value = CellRichText([
            "About ",
            TextBlock(bold_inline, f"{ideal_pct:.1f}%"),
            " of territories fall within this range."
        ])

        header_fill = PatternFill(start_color="C7CDDB", end_color="C7CDDB", fill_type="solid")
        for col in range(1, 4):
            ws1.cell(row=12, column=col).fill = header_fill

        for r in range(13, 19):
            if ws1.cell(row=r, column=1).value == "Ideal":
                for col in range(1, 4):
                    ws1.cell(row=r, column=col).font = Font(bold=True)

        ws1['A20'].value = CellRichText(["As a part of this analysis, every ", TextBlock(bold_inline, "address point"), " within your territory was collected & identified."])
        ws1['A21'].value = "These addresses, with a little reformatting, can be added to NWS or other programs (Please see http://www.territoryanalysis.com/ to see if your system is supported.)"
        ws1['A21'].hyperlink = "http://www.territoryanalysis.com/"
        ws1['A21'].font = Font(color="0563C1", underline="single")
        ws1['A22'].value = "It's suggested to export this file into a program you can easily edit, like excel or google sheets."
        ws1['A23'].value = "That will allow you to expand cells to read easier, create custom filters to see specific data, and customize the sheet to make it more legible."
        ws1['A24'].value = ""
        ws1['A25'].value = CellRichText(["The ", TextBlock(bold_inline, "DASHBOARD"), " tab displays basic statistics about the territory that was analyzed"])
        ws1['A26'].value = CellRichText(["The ", TextBlock(bold_inline, "COUNTS"), " tab organizes territories by size. This is done by 'counting' workable addresses, not geographical size."])
        ws1['A27'].value = CellRichText(["The ", TextBlock(bold_inline, "ADDRESS LIST"), " tab displays every workable address in your territory."])
        ws1['A28'].value = CellRichText(["The ", TextBlock(bold_inline, "APARTMENTS"), " tab displays every multifamily above 5 units in your territory. Large units can be explanations for inflated door-to-door territories."])
        ws1['A29'].value = CellRichText(["The ", TextBlock(bold_inline, "BORDER REWRITES"), " tab displays borders within your territory that may benefit from being redrawn. The intent is to shrink oversized territories adjacent to undersized territories. These are just suggestions."])
        ws1['A30'].value = CellRichText(["The ", TextBlock(bold_inline, "EXCLUDED AUDIT"), " tab displays addresses that are NOT counted towards your territory. These are usually addresses of highways, vacant lots, parks, etc. This is included for confidence."])

        # --- TAB 2: COUNT PER TERRITORY ---
        counts_df_sorted = counts_df.sort_values(by='Territory_Name').rename(columns={
            'Territory_Name': 'Territory Name', 
            'Total_Addresses': '# of Addresses'
        })
        counts_df_sorted.to_excel(writer, sheet_name="Counts", index=False)
        ws2 = writer.sheets['Counts']
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 15
        
        for row in range(2, len(counts_df_sorted) + 2):
            ws2[f'B{row}'].alignment = Alignment(horizontal='center')
            cell = ws2[f'C{row}']
            if cell.value == 'Ideal':
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif cell.value == 'Undersized':
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif cell.value == 'Oversized':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # --- TAB 3: ADDRESS LIST ---
        valid_gdf['HouseNum_Sort'] = pd.to_numeric(valid_gdf['HouseNo'], errors='coerce').fillna(0)
        address_list_df = valid_gdf.sort_values(by=['Territory_Name', 'Street', 'HouseNum_Sort', 'Unit'])
        
        export_df = address_list_df[['Territory_Name', 'Mailable_Address', 'HouseNo', 'Street', 'Unit', 'Zip_Code']].rename(columns={
            'Territory_Name': 'Territory Name', 
            'Mailable_Address': 'Mailable Address'
        })
        export_df.to_excel(writer, sheet_name="Address List", index=False)
        
        ws3 = writer.sheets['Address List']
        ws3.column_dimensions['C'].hidden = True
        ws3.column_dimensions['D'].hidden = True
        ws3.column_dimensions['E'].hidden = True
        ws3.column_dimensions['F'].hidden = True
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 55

        # --- TAB 4: APARTMENTS / POTENTIAL LETTER WRITING ---
        apt_groups = valid_gdf.groupby(['Territory_Name', 'Base_Address'], observed=True).size().reset_index(name='Total Units')
        apt_groups = apt_groups[apt_groups['Total Units'] >= 5]
        
        if not counts_df.empty:
            cat_mapping = counts_df.set_index('Territory_Name')['Category'].to_dict()
            apt_groups['Status'] = apt_groups['Territory_Name'].map(cat_mapping)
        else:
            apt_groups['Status'] = "Unknown"
        
        def format_terr_name(row):
            return f"{row['Territory_Name']} [{row['Status']}]"
            
        if not apt_groups.empty:
            apt_groups['Territory Name'] = apt_groups.apply(format_terr_name, axis=1)
            apt_groups.rename(columns={'Base_Address': 'Base Address'}, inplace=True)
            apt_export = apt_groups[['Territory Name', 'Base Address', 'Total Units']]
        else:
            apt_export = pd.DataFrame(columns=['Territory Name', 'Base Address', 'Total Units'])
            
        apt_export.to_excel(writer, sheet_name="Apartments", index=False)
        ws4 = writer.sheets['Apartments']
        ws4.column_dimensions['A'].width = 30
        ws4.column_dimensions['B'].width = 40
        ws4.column_dimensions['C'].width = 15
        
        for row in range(2, len(apt_export) + 2):
            ws4[f'C{row}'].alignment = Alignment(horizontal='center')

        # --- TAB 5: BORDER REWRITES ---
        oversized = counts_df[counts_df['Category'] == 'Oversized']['Territory_Name'].tolist() if not counts_df.empty else []
        undersized = counts_df[counts_df['Category'] == 'Undersized']['Territory_Name'].tolist() if not counts_df.empty else []
        
        terr_geoms = kml_gdf.drop_duplicates('Territory_Name').set_index('Territory_Name')
        suggestions = []
        
        for over_name in oversized:
            if over_name in terr_geoms.index:
                over_geom = terr_geoms.loc[over_name, 'geometry_terr']
                over_count = counts_df[counts_df['Territory_Name'] == over_name]['Total_Addresses'].values[0]
                
                for under_name in undersized:
                    if under_name in terr_geoms.index:
                        under_geom = terr_geoms.loc[under_name, 'geometry_terr']
                        if over_geom.touches(under_geom) or over_geom.intersects(under_geom):
                            under_count = counts_df[counts_df['Territory_Name'] == under_name]['Total_Addresses'].values[0]
                            suggestions.append([over_name, over_count, under_name, under_count, ""])
                        
        pd.DataFrame(suggestions, columns=["Too Large", "Count", "Too Small", "Count ", "Recommendation"]).to_excel(writer, sheet_name="Border Rewrites", index=False)
        ws5 = writer.sheets['Border Rewrites']
        ws5.column_dimensions['A'].width = 15
        ws5.column_dimensions['C'].width = 15
        ws5.column_dimensions['E'].width = 85

        for r in range(2, len(suggestions) + 2):
            diff = abs(suggestions[r-2][1] - suggestions[r-2][3])
            ws5.cell(row=r, column=5).value = CellRichText([
                "That is a ",
                TextBlock(bold_inline, f"{diff} address difference"),
                f". Shrink {suggestions[r-2][0]} & Expand {suggestions[r-2][2]}."
            ])

        # --- TAB 6: EXCLUDED AUDIT ---
        if not excluded_gdf.empty:
            excluded_gdf['HouseNum_Sort'] = pd.to_numeric(excluded_gdf['HouseNo'], errors='coerce').fillna(0)
            excluded_list_df = excluded_gdf.sort_values(by=['Territory_Name', 'Street', 'HouseNum_Sort', 'Unit'])
            
            export_ex_df = excluded_list_df[['Territory_Name', 'Mailable_Address', 'Addr_Statu', 'HouseNo', 'Street', 'Unit', 'Zip_Code']].rename(columns={
                'Territory_Name': 'Territory Name', 
                'Mailable_Address': 'Mailable Address'
            })
            export_ex_df.to_excel(writer, sheet_name="Excluded Audit", index=False)
            
            ws6 = writer.sheets['Excluded Audit']
            ws6.column_dimensions['D'].hidden = True
            ws6.column_dimensions['E'].hidden = True
            ws6.column_dimensions['F'].hidden = True
            ws6.column_dimensions['G'].hidden = True
            ws6.column_dimensions['A'].width = 15
            ws6.column_dimensions['B'].width = 55
            ws6.column_dimensions['C'].width = 28
        else:
            pd.DataFrame(columns=["Notice"]).to_excel(writer, sheet_name="Excluded Audit", index=False)
            writer.sheets['Excluded Audit'].cell(row=2, column=1, value="No addresses were excluded in this map area.")

        # --- EXCEL UX POLISH (FREEZE PANES & BOLD HEADERS) ---
        tabs_to_format = ["Counts", "Address List", "Apartments", "Border Rewrites", "Excluded Audit"]
        for tab_name in tabs_to_format:
            ws = writer.sheets[tab_name]
            ws.freeze_panes = 'A2'
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal=cell.alignment.horizontal if cell.alignment else 'left')

        # Tab colorization
        writer.sheets['Dashboard'].sheet_properties.tabColor = "1E90FF"
        writer.sheets['Counts'].sheet_properties.tabColor = "32CD32"
        writer.sheets['Address List'].sheet_properties.tabColor = "32CD32"
        writer.sheets['Apartments'].sheet_properties.tabColor = "FF8C00"
        writer.sheets['Border Rewrites'].sheet_properties.tabColor = "FF0000"
        writer.sheets['Excluded Audit'].sheet_properties.tabColor = "808080" 

    output.seek(0)
    return output

# --- 4. EXECUTION FLOW ---
if 'last_uploaded_kml' not in st.session_state:
    st.session_state['last_uploaded_kml'] = None

if uploaded_kml != st.session_state['last_uploaded_kml']:
    if 'excel_data' in st.session_state:
        del st.session_state['excel_data']
    st.session_state['last_uploaded_kml'] = uploaded_kml

if uploaded_kml:
    if st.button("Generate Territory Analysis"):
        with st.spinner(f"Loading Master {selected_county} County Data..."):
            parcel_gdf = load_county_data(selected_county)
            
        if parcel_gdf is not None:
            with st.spinner("Parsing KML Territories & Executing Spatial Join..."):
                try:
                    kml_gdf = gpd.read_file(uploaded_kml, driver="KML")
                    kml_gdf['geometry'] = kml_gdf['geometry'].make_valid()
                    
                    fallback_names = "Territory_" + kml_gdf.index.to_series().astype(str)
                    if 'Name' in kml_gdf.columns:
                        kml_gdf['Territory_Name'] = kml_gdf['Name'].fillna(fallback_names)
                    elif 'Description' in kml_gdf.columns:
                        kml_gdf['Territory_Name'] = kml_gdf['Description'].fillna(fallback_names)
                    else:
                        kml_gdf['Territory_Name'] = fallback_names
                    
                    if parcel_gdf.crs != kml_gdf.crs:
                        parcel_gdf = parcel_gdf.to_crs(kml_gdf.crs)
                        
                    bounding_box = kml_gdf.unary_union.envelope
                    parcel_gdf = gpd.clip(parcel_gdf, bounding_box)
                    
                    kml_gdf = kml_gdf.rename(columns={'geometry': 'geometry_terr'})
                    kml_gdf = kml_gdf.set_geometry('geometry_terr')
                    
                    joined_gdf = gpd.sjoin(parcel_gdf, kml_gdf, how="inner", predicate="within")
                    joined_gdf = joined_gdf.dropna(subset=['Territory_Name'])
                    
                    with st.spinner("Generating Excel Report..."):
                        excel_file = generate_excel_report(joined_gdf, kml_gdf, MIN_GOAL, MAX_GOAL, congregation_name.replace(" ", ""))
                        filename = f"{congregation_name.replace(' ', '')}_{datetime.datetime.now().strftime('%B%Y')}_TerritoryAnalysis.xlsx"
                        
                        st.session_state['excel_data'] = excel_file.getvalue()
                        st.session_state['excel_filename'] = filename
                        
                        st.success("Analysis Complete!")
                        
                except Exception as e:
                    st.error(f"An error occurred during processing: {e}")

    if 'excel_data' in st.session_state:
        st.info("Analysis results ready for download.")
        st.download_button(
            label="⬇️ Download Excel Analysis",
            data=st.session_state['excel_data'],
            file_name=st.session_state['excel_filename'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
