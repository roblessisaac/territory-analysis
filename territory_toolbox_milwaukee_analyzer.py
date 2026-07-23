import gc
import io
import re
import hashlib
import datetime as dt
from dataclasses import dataclass, asdict
from typing import Dict, Iterable, List, Optional, Tuple

import streamlit as st
import pandas as pd
import numpy as np
import geopandas as gpd
import fiona

from shapely.geometry import Point
from shapely.ops import unary_union

from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule

# Enable KML support in GeoPandas/Fiona
fiona.drvsupport.supported_drivers["KML"] = "rw"
fiona.drvsupport.supported_drivers["LIBKML"] = "rw"


# ============================================================
# 1. CONFIGURATION
# ============================================================

APP_VERSION = "TerritoryToolbox Milwaukee Analyzer v0.4.0"
APP_NAME = "TerritoryToolbox"
DEFAULT_MIN_GOAL = 80
DEFAULT_MAX_GOAL = 100
DEFAULT_APARTMENT_THRESHOLD = 5
DEFAULT_BORDER_DISTANCE_METERS = 50

MILWAUKEE_PROFILE = {
    "county_name": "Milwaukee",
    "state": "WI",
    "county_data_path": "zip://data/Milwaukee_Datapoints07072026.zip",
    "county_data_label": "Milwaukee_Datapoints07072026.zip",
    "geometry_contract": "Milwaukee source layer is expected to contain address points. Polygon inputs are converted to representative points.",
    "spatial_predicate": "intersects",  # catches boundary points and allows conflict QA
    "required_columns": [
        "HouseNo",
        "Street",
        "Addr_Statu",
    ],
    "optional_columns": [
        "HouseSx",
        "Dir",
        "StType",
        "Unit",
        "Muni",
        "Zip_Code",
    ],
    "column_map": {
        "HouseNo": "HouseNo",
        "HouseSx": "HouseSx",
        "Street_PreDir": "Dir",
        "Street_Name": "Street",
        "Street_Type": "StType",
        "Unit": "Unit",
        "Muni": "Muni",
        "Zip_Code": "Zip_Code",
        "Address_Status": "Addr_Statu",
    },
    "excluded_statuses": [
        "Undeveloped",
        "Parking Lot",
        "ROW",
        "Park or Recreational Facility",
        "Undeveloped Outlot",
        "Sliver or Remnant",
        "Non Addressable Assoc with Adj Parcel",
    ],
}

STATUS_COLOR = {
    "Ideal": "C6EFCE",
    "Undersized": "FFEB9C",
    "Oversized": "FFC7CE",
    "Empty": "D9EAD3",
    "Needs Review": "FCE4D6",
}

HEADER_FILL = "1F4E3D"
HEADER_FONT = "FFFFFF"
LIGHT_GREEN = "E2F0D9"
LIGHT_BLUE = "DDEBF7"
LIGHT_ORANGE = "FCE4D6"
LIGHT_GRAY = "F2F2F2"
WARNING_FILL = "FFF2CC"
ERROR_FILL = "F4CCCC"


# ============================================================
# 2. VALIDATION TYPES
# ============================================================

@dataclass
class ValidationIssue:
    severity: str  # ERROR / WARNING / INFO
    code: str
    message: str
    count: int = 0
    details: str = ""

    def to_dict(self) -> Dict[str, object]:
        return asdict(self)


def add_issue(
    issues: List[ValidationIssue],
    severity: str,
    code: str,
    message: str,
    count: int = 0,
    details: str = "",
) -> None:
    issues.append(ValidationIssue(severity, code, message, int(count or 0), details or ""))


def has_fatal_errors(issues: List[ValidationIssue]) -> bool:
    return any(issue.severity.upper() == "ERROR" for issue in issues)


# ============================================================
# 3. GENERAL HELPERS
# ============================================================

_NULLISH = {"", "nan", "none", "null", "nat", "<na>"}


def clean_scalar(value) -> str:
    """Safely normalize one scalar to a stripped string without turning blanks into 'nan'."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in _NULLISH:
        return ""
    return text


def clean_text_series(series: pd.Series) -> pd.Series:
    """Vectorized string cleanup for source fields."""
    return (
        series.fillna("")
        .astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.strip()
        .replace({"nan": "", "None": "", "NULL": "", "<NA>": ""})
    )


def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_scalar(c).replace("\ufeff", "").strip() for c in df.columns]
    return df


def natural_keys(text) -> List[object]:
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r"(\d+)", str(text))]


def natural_sort_df(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if df.empty or col not in df.columns:
        return df
    return df.assign(_sort_key=df[col].map(natural_keys)).sort_values("_sort_key").drop(columns="_sort_key")


def make_safe_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", " ", str(name)).strip()
    return safe[:31] if safe else "Sheet"


def sanitize_table_name(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not re.match(r"^[A-Za-z_]", safe):
        safe = f"T_{safe}"
    return safe[:200]


def md5_short(text: str, length: int = 12) -> str:
    return hashlib.md5(text.encode("utf-8", errors="ignore")).hexdigest()[:length]


def remove_trailing_dot_zero(value: str) -> str:
    """Remove only a terminal .0 artifact. Do not corrupt values like 101.05."""
    return re.sub(r"\.0$", "", clean_scalar(value))


def first_existing_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return None


def safe_unique_join(values: Iterable[object], max_items: int = 8) -> str:
    seen = []
    for value in values:
        val = clean_scalar(value)
        if val and val not in seen:
            seen.append(val)
        if len(seen) >= max_items:
            break
    return ", ".join(seen)


def normalize_status_value(value: str) -> str:
    text = re.sub(r"\s+", " ", clean_scalar(value)).strip()
    return text.casefold()


def display_status_value(value: str) -> str:
    text = re.sub(r"\s+", " ", clean_scalar(value)).strip()
    return text


def normalize_zip_value(value: str) -> str:
    """
    Preserve ZIP as text. Supports 53208, 53208-1234, and numeric artifacts like 53208.0.
    For Milwaukee output, this returns the leading 5-digit ZIP when present.
    """
    text = remove_trailing_dot_zero(value)
    if not text:
        return ""
    match = re.search(r"\d{5}", text)
    return match.group(0) if match else text


def extract_house_sort(value: str) -> int:
    text = clean_scalar(value)
    match = re.match(r"^(\d+)", text)
    return int(match.group(1)) if match else 0


def extract_territory_number(name: str) -> str:
    """Extract trailing territory number for downstream tools, without relying on it as the primary key."""
    text = clean_scalar(name)
    match = re.search(r"(\d+)\s*$", text)
    return match.group(1) if match else ""


def extract_territory_category(name: str) -> str:
    """Conservative display-category extraction from territory name. Example: Hi-Mount 99 -> Hi-Mount."""
    text = clean_scalar(name)
    if not text:
        return ""
    return re.sub(r"\s*\d+\s*$", "", text).strip()


# ============================================================
# 4. DATA LOADING
# ============================================================

@st.cache_data(show_spinner=False)
def load_county_data(county_name: str) -> Optional[gpd.GeoDataFrame]:
    """Load Milwaukee source address data. Kept cacheable for Streamlit Cloud."""
    if county_name != "Milwaukee":
        return None

    path = MILWAUKEE_PROFILE["county_data_path"]

    try:
        try:
            # pyogrio is faster when available; Fiona fallback remains compatible.
            return gpd.read_file(path, engine="pyogrio")
        except Exception:
            return gpd.read_file(path)
    except Exception as exc:
        st.error(
            "Could not load Milwaukee county data. Confirm that "
            "data/Milwaukee_Datapoints07072026.zip exists in your repo.\n\n"
            f"Error: {exc}"
        )
        return None


def load_kml(uploaded_kml) -> gpd.GeoDataFrame:
    """Read uploaded KML and normalize CRS/name fields."""
    kml_gdf = gpd.read_file(uploaded_kml, driver="KML")
    kml_gdf = clean_column_names(kml_gdf)

    if kml_gdf.empty:
        return kml_gdf

    # KML is WGS84 by definition. GeoPandas sometimes does not attach CRS.
    if kml_gdf.crs is None:
        kml_gdf = kml_gdf.set_crs("EPSG:4326", allow_override=True)

    # Repair invalid geometries where possible.
    kml_gdf["_geometry_was_invalid"] = ~kml_gdf.geometry.is_valid
    try:
        kml_gdf["geometry"] = kml_gdf.geometry.make_valid()
    except Exception:
        # Conservative fallback for environments without make_valid support.
        kml_gdf["geometry"] = kml_gdf.geometry.buffer(0)

    fallback_names = "Territory_" + kml_gdf.index.to_series().astype(str)
    used_fallback = pd.Series(False, index=kml_gdf.index)

    if "Name" in kml_gdf.columns:
        names = clean_text_series(kml_gdf["Name"])
        used_fallback = names.eq("")
        kml_gdf["Territory_Name"] = names.where(~used_fallback, fallback_names)
    elif "Description" in kml_gdf.columns:
        names = clean_text_series(kml_gdf["Description"])
        used_fallback = names.eq("")
        kml_gdf["Territory_Name"] = names.where(~used_fallback, fallback_names)
    else:
        used_fallback = pd.Series(True, index=kml_gdf.index)
        kml_gdf["Territory_Name"] = fallback_names

    kml_gdf["Territory_Name"] = clean_text_series(kml_gdf["Territory_Name"])
    kml_gdf["Territory_Number"] = kml_gdf["Territory_Name"].map(extract_territory_number)
    kml_gdf["Territory_CategoryCode"] = kml_gdf["Territory_Name"].map(extract_territory_category)
    kml_gdf["Territory_Key"] = kml_gdf["Territory_Name"].map(lambda x: md5_short(x.upper()))
    kml_gdf["_used_fallback_name"] = used_fallback.astype(bool)

    # Dissolve multiple KML features with the same territory name into one territory.
    # This supports multi-part territory shapes while making duplicate-name behavior explicit.
    dup_counts = kml_gdf.groupby("Territory_Name").size().rename("_feature_count")

    non_geom_cols = [
        "Territory_Name",
        "Territory_Number",
        "Territory_CategoryCode",
        "Territory_Key",
        "_used_fallback_name",
        "_geometry_was_invalid",
    ]
    keep = kml_gdf[non_geom_cols + ["geometry"]].copy()
    dissolved = keep.dissolve(by="Territory_Name", as_index=False, aggfunc={
        "Territory_Number": "first",
        "Territory_CategoryCode": "first",
        "Territory_Key": "first",
        "_used_fallback_name": "max",
        "_geometry_was_invalid": "max",
    })
    dissolved = dissolved.merge(dup_counts.reset_index(), on="Territory_Name", how="left")
    dissolved = natural_sort_df(dissolved, "Territory_Name")
    return dissolved


# ============================================================
# 5. VALIDATION
# ============================================================


def validate_county_schema(parcel_gdf: gpd.GeoDataFrame) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = []

    if parcel_gdf is None or parcel_gdf.empty:
        add_issue(issues, "ERROR", "COUNTY_EMPTY", "County address data could not be loaded or is empty.")
        return issues

    missing_required = [c for c in MILWAUKEE_PROFILE["required_columns"] if c not in parcel_gdf.columns]
    if missing_required:
        add_issue(
            issues,
            "ERROR",
            "MISSING_REQUIRED_COLUMNS",
            "Milwaukee source data is missing required column(s).",
            len(missing_required),
            ", ".join(missing_required),
        )

    missing_optional = [c for c in MILWAUKEE_PROFILE["optional_columns"] if c not in parcel_gdf.columns]
    if missing_optional:
        add_issue(
            issues,
            "WARNING",
            "MISSING_OPTIONAL_COLUMNS",
            "Milwaukee source data is missing optional column(s). Blank fallback values will be used.",
            len(missing_optional),
            ", ".join(missing_optional),
        )

    if parcel_gdf.crs is None:
        add_issue(issues, "ERROR", "COUNTY_MISSING_CRS", "County source data has no CRS. Spatial assignment cannot be trusted.")

    if "geometry" not in parcel_gdf.columns:
        add_issue(issues, "ERROR", "COUNTY_MISSING_GEOMETRY", "County source data has no geometry column.")
    else:
        null_geom = int(parcel_gdf.geometry.isna().sum())
        if null_geom:
            add_issue(issues, "WARNING", "COUNTY_NULL_GEOMETRIES", "County source data contains null geometries.", null_geom)

    return issues


def validate_kml(kml_gdf: gpd.GeoDataFrame) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = []

    if kml_gdf is None or kml_gdf.empty:
        add_issue(issues, "ERROR", "KML_EMPTY", "Uploaded KML has no territory features.")
        return issues

    if kml_gdf.crs is None:
        add_issue(issues, "ERROR", "KML_MISSING_CRS", "Uploaded KML has no CRS. KML should normally be EPSG:4326.")

    null_names = int(kml_gdf["Territory_Name"].fillna("").astype(str).str.strip().eq("").sum())
    if null_names:
        add_issue(issues, "ERROR", "KML_BLANK_NAMES", "Some KML territory names are blank.", null_names)

    fallback_names = int(kml_gdf.get("_used_fallback_name", pd.Series(dtype=bool)).sum())
    if fallback_names:
        add_issue(issues, "WARNING", "KML_FALLBACK_NAMES", "Some KML features had no usable Name/Description and received fallback names.", fallback_names)

    invalid_before = int(kml_gdf.get("_geometry_was_invalid", pd.Series(dtype=bool)).sum())
    if invalid_before:
        add_issue(issues, "WARNING", "KML_GEOMETRY_REPAIRED", "Some KML geometries were invalid and were repaired before analysis.", invalid_before)

    multipart = int((kml_gdf.get("_feature_count", pd.Series(1, index=kml_gdf.index)) > 1).sum())
    if multipart:
        add_issue(
            issues,
            "INFO",
            "KML_DUPLICATE_NAMES_DISSOLVED",
            "Some repeated KML names were dissolved into multi-part territories.",
            multipart,
            "Repeated names are treated as the same territory.",
        )

    empty_geom = int(kml_gdf.geometry.is_empty.sum())
    if empty_geom:
        add_issue(issues, "ERROR", "KML_EMPTY_GEOMETRIES", "Some KML territories have empty geometry.", empty_geom)

    return issues


def validate_goal_range(min_goal: int, max_goal: int) -> List[ValidationIssue]:
    issues: List[ValidationIssue] = []
    if min_goal <= 0 or max_goal <= 0:
        add_issue(issues, "ERROR", "INVALID_GOAL_RANGE", "Goal range values must be positive.")
    if min_goal > max_goal:
        add_issue(issues, "ERROR", "INVALID_GOAL_RANGE", "Minimum goal cannot be greater than maximum goal.")
    return issues


# ============================================================
# 6. GEOMETRY PREPARATION AND SPATIAL ASSIGNMENT
# ============================================================


def ensure_profile_columns(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Ensure optional Milwaukee columns exist so later vectorized code is stable."""
    gdf = clean_column_names(gdf)
    for col in MILWAUKEE_PROFILE["optional_columns"]:
        if col not in gdf.columns:
            gdf[col] = ""
    return gdf


def make_source_record_id(gdf: gpd.GeoDataFrame) -> pd.Series:
    """Create stable source IDs even when the county file lacks an explicit parcel/address ID."""
    candidate_cols = [
        "OBJECTID",
        "ObjectID",
        "FID",
        "PARCELID",
        "ParcelID",
        "TAXKEY",
        "Taxkey",
        "Tax_Key",
        "ADDR_ID",
        "AddressID",
    ]
    existing = first_existing_column(gdf, candidate_cols)
    if existing:
        return clean_text_series(gdf[existing]).where(clean_text_series(gdf[existing]).ne(""), "SRC_" + gdf.index.astype(str))
    return "SRC_" + gdf.index.astype(str)


def get_geometry_mode(gdf: gpd.GeoDataFrame) -> str:
    geom_types = set(gdf.geometry.geom_type.dropna().unique().tolist())
    if geom_types.issubset({"Point", "MultiPoint"}):
        return "point"
    return "representative_point"


def prepare_address_points(parcel_gdf: gpd.GeoDataFrame, issues: List[ValidationIssue]) -> gpd.GeoDataFrame:
    """Prepare source records as address points for point-in-territory assignment."""
    gdf = ensure_profile_columns(parcel_gdf.copy())
    gdf["Source_Record_ID"] = make_source_record_id(gdf)

    geometry_mode = get_geometry_mode(gdf)
    gdf["Geometry_Mode"] = geometry_mode

    if geometry_mode != "point":
        add_issue(
            issues,
            "WARNING",
            "COUNTY_POLYGON_TO_POINT",
            "County layer is not purely point geometry. Representative points were used for spatial assignment.",
            len(gdf),
        )
        try:
            gdf["geometry"] = gdf.geometry.make_valid().representative_point()
        except Exception:
            gdf["geometry"] = gdf.geometry.buffer(0).representative_point()

    return gdf


def subset_county_to_kml_extent(
    address_points: gpd.GeoDataFrame,
    kml_gdf: gpd.GeoDataFrame,
    issues: List[ValidationIssue],
) -> Tuple[gpd.GeoDataFrame, gpd.GeoDataFrame]:
    """Reproject KML to county CRS and subset source points to the KML bounding box."""
    if address_points.crs is None:
        add_issue(issues, "ERROR", "COUNTY_MISSING_CRS", "County source data has no CRS.")
        return address_points.iloc[0:0].copy(), kml_gdf

    kml_for_join = kml_gdf.to_crs(address_points.crs) if kml_gdf.crs != address_points.crs else kml_gdf.copy()

    try:
        minx, miny, maxx, maxy = kml_for_join.total_bounds
        subset = address_points.cx[minx:maxx, miny:maxy].copy()
    except Exception:
        add_issue(issues, "WARNING", "BBOX_FILTER_FAILED", "Bounding-box filtering failed; falling back to full county dataset.")
        subset = address_points.copy()

    if subset.empty:
        add_issue(
            issues,
            "ERROR",
            "NO_CANDIDATE_ADDRESSES_IN_EXTENT",
            "No county address points fall inside the uploaded KML bounding box. Check CRS, county selection, or KML boundaries.",
        )
    else:
        add_issue(
            issues,
            "INFO",
            "CANDIDATE_ADDRESSES_IN_EXTENT",
            "Candidate source records inside KML bounding box.",
            len(subset),
        )

    return subset, kml_for_join


def assign_addresses_to_territories(
    address_points: gpd.GeoDataFrame,
    kml_for_join: gpd.GeoDataFrame,
    issues: List[ValidationIssue],
) -> Tuple[gpd.GeoDataFrame, pd.DataFrame]:
    """
    Spatially assign address points to KML territories.
    Uses left join for QA, then chooses one assignment per source record while flagging conflicts.
    """
    predicate = MILWAUKEE_PROFILE["spatial_predicate"]

    territories = kml_for_join[["Territory_Name", "Territory_Number", "Territory_CategoryCode", "Territory_Key", "geometry"]].copy()

    joined = gpd.sjoin(
        address_points,
        territories,
        how="left",
        predicate=predicate,
    )

    joined = joined.drop(columns=[c for c in ["index_right"] if c in joined.columns])

    # Build assignment QA before deduping conflict rows.
    assigned_names = (
        joined.dropna(subset=["Territory_Name"])
        .groupby("Source_Record_ID")["Territory_Name"]
        .agg(lambda s: sorted(set(clean_scalar(v) for v in s if clean_scalar(v))))
        .rename("Matched_Territories_List")
    )
    assignment_counts = assigned_names.map(len).rename("Matched_Territory_Count")
    assignment_summary = pd.concat([assigned_names, assignment_counts], axis=1).reset_index()
    assignment_summary["Matched_Territories"] = assignment_summary["Matched_Territories_List"].map(lambda x: "; ".join(x))
    assignment_summary = assignment_summary.drop(columns="Matched_Territories_List")

    joined = joined.merge(assignment_summary, on="Source_Record_ID", how="left")
    joined["Matched_Territory_Count"] = joined["Matched_Territory_Count"].fillna(0).astype(int)
    joined["Matched_Territories"] = clean_text_series(joined.get("Matched_Territories", pd.Series("", index=joined.index)))

    unassigned_count = int(joined["Territory_Name"].isna().sum())
    multi_count = int((joined["Matched_Territory_Count"] > 1).sum())

    if unassigned_count:
        add_issue(
            issues,
            "WARNING",
            "UNASSIGNED_ADDRESS_POINTS",
            "Some candidate address points in the KML extent were not assigned to any territory.",
            unassigned_count,
            "These are included in the Spatial QA tab, not the Address List.",
        )

    unique_multi_sources = int(assignment_summary[assignment_summary["Matched_Territory_Count"] > 1]["Source_Record_ID"].nunique())
    if unique_multi_sources:
        add_issue(
            issues,
            "WARNING",
            "MULTI_ASSIGNED_ADDRESS_POINTS",
            "Some address points matched multiple territories, likely due to overlaps or boundary points.",
            unique_multi_sources,
            "The output keeps one deterministic assignment and flags the row for review.",
        )

    # Deterministic single assignment: sort territory naturally and keep first per source record.
    joined["_territory_sort"] = joined["Territory_Name"].fillna("ZZZ_UNASSIGNED").map(natural_keys)
    joined = joined.sort_values(["Source_Record_ID", "_territory_sort"], kind="mergesort")
    assigned_single = joined.drop_duplicates(subset=["Source_Record_ID"], keep="first").drop(columns="_territory_sort")

    spatial_qa = joined[[
        "Source_Record_ID",
        "Territory_Name",
        "Matched_Territory_Count",
        "Matched_Territories",
    ]].copy()
    spatial_qa["QA_Status"] = np.select(
        [
            spatial_qa["Territory_Name"].isna(),
            spatial_qa["Matched_Territory_Count"] > 1,
        ],
        [
            "Unassigned",
            "Matched Multiple Territories",
        ],
        default="Assigned",
    )

    return assigned_single, spatial_qa


# ============================================================
# 7. ADDRESS NORMALIZATION
# ============================================================


def normalize_milwaukee_addresses(gdf: gpd.GeoDataFrame, issues: List[ValidationIssue]) -> gpd.GeoDataFrame:
    """Normalize Milwaukee source columns into source-of-truth output fields."""
    df = gdf.copy()
    profile = MILWAUKEE_PROFILE
    cmap = profile["column_map"]

    df["Raw_Address_Status"] = clean_text_series(df[cmap["Address_Status"]]) if cmap["Address_Status"] in df.columns else ""
    df["Normalized_Address_Status"] = df["Raw_Address_Status"].map(normalize_status_value)

    excluded_lookup = {normalize_status_value(s): s for s in profile["excluded_statuses"]}
    df["Is_Excluded"] = df["Normalized_Address_Status"].isin(set(excluded_lookup.keys()))
    df["Exclusion_Reason"] = df["Normalized_Address_Status"].map(excluded_lookup).fillna("")
    df["Exclusion_Rule"] = np.where(df["Is_Excluded"], "Milwaukee invalid address status", "")
    df["Review_Recommended"] = False

    # Core address fields
    house_col = cmap["HouseNo"]
    housesx_col = cmap["HouseSx"]
    predir_col = cmap["Street_PreDir"]
    street_col = cmap["Street_Name"]
    sttype_col = cmap["Street_Type"]
    unit_col = cmap["Unit"]
    muni_col = cmap["Muni"]
    zip_col = cmap["Zip_Code"]

    df["HouseNo_Raw"] = clean_text_series(df[house_col]) if house_col in df.columns else ""
    df["HouseNo_Display"] = df["HouseNo_Raw"].map(remove_trailing_dot_zero)
    df["HouseSx"] = clean_text_series(df[housesx_col]) if housesx_col in df.columns else ""
    df["Full_HouseNo"] = (df["HouseNo_Display"] + df["HouseSx"]).str.strip()
    df["HouseNo_Sort"] = df["HouseNo_Display"].map(extract_house_sort)

    df["Street_PreDir"] = clean_text_series(df[predir_col]) if predir_col in df.columns else ""
    df["Street_Name"] = clean_text_series(df[street_col]) if street_col in df.columns else ""
    df["Street_Type"] = clean_text_series(df[sttype_col]) if sttype_col in df.columns else ""
    df["Street_PostDir"] = ""  # Milwaukee source currently does not provide a separate post-direction field.

    street_parts = ["Street_PreDir", "Street_Name", "Street_Type", "Street_PostDir"]
    df["Street_Full"] = (
        df[street_parts]
        .fillna("")
        .agg(lambda row: " ".join([clean_scalar(x) for x in row if clean_scalar(x)]), axis=1)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )

    df["Unit"] = clean_text_series(df[unit_col]) if unit_col in df.columns else ""
    df["Unit_Normalized"] = df["Unit"].str.upper().str.replace(r"\s+", " ", regex=True).str.strip()
    df["Muni"] = clean_text_series(df[muni_col]) if muni_col in df.columns else ""
    df["State"] = profile["state"]
    df["Zip_Code"] = clean_text_series(df[zip_col]).map(normalize_zip_value) if zip_col in df.columns else ""

    df["Base_Address_Line"] = (df["Full_HouseNo"] + " " + df["Street_Full"]).str.replace(r"\s+", " ", regex=True).str.strip()
    df["Unit_Address_Part"] = np.where(df["Unit_Normalized"].ne(""), " Apt " + df["Unit_Normalized"], "")
    df["Mailable_Address_Line"] = (df["Base_Address_Line"] + df["Unit_Address_Part"]).str.replace(r"\s+", " ", regex=True).str.strip()

    # Avoid malformed double commas when municipality is missing.
    location_with_muni = (df["Muni"] + ", " + df["State"] + " " + df["Zip_Code"]).str.strip()
    location_no_muni = (df["State"] + " " + df["Zip_Code"]).str.strip()
    df["Location_Display"] = np.where(df["Muni"].ne(""), location_with_muni, location_no_muni)
    df["Location_Display"] = df["Location_Display"].str.replace(r"\s+", " ", regex=True).str.strip(" ,")

    df["Base_Address"] = np.where(
        df["Location_Display"].ne(""),
        df["Base_Address_Line"] + ", " + df["Location_Display"],
        df["Base_Address_Line"],
    )
    df["Mailable_Address"] = np.where(
        df["Location_Display"].ne(""),
        df["Mailable_Address_Line"] + ", " + df["Location_Display"],
        df["Mailable_Address_Line"],
    )
    df["Base_Address"] = clean_text_series(df["Base_Address"])
    df["Mailable_Address"] = clean_text_series(df["Mailable_Address"])

    # Territory fields for downstream tools.
    df["Territory_Name"] = clean_text_series(df.get("Territory_Name", pd.Series("", index=df.index)))
    df["Territory_DisplayName"] = df["Territory_Name"]
    df["Territory_Number"] = clean_text_series(df.get("Territory_Number", pd.Series("", index=df.index)))
    df["Territory_CategoryCode"] = clean_text_series(df.get("Territory_CategoryCode", pd.Series("", index=df.index)))
    df["Territory_Key"] = clean_text_series(df.get("Territory_Key", pd.Series("", index=df.index)))

    # Coordinates: always output WGS84 lat/lon for downstream systems.
    try:
        ll = df.to_crs("EPSG:4326") if df.crs and str(df.crs).upper() != "EPSG:4326" else df
        df["Latitude"] = ll.geometry.y
        df["Longitude"] = ll.geometry.x
    except Exception:
        add_issue(issues, "WARNING", "LAT_LON_FAILED", "Could not derive WGS84 latitude/longitude for some records.")
        df["Latitude"] = np.nan
        df["Longitude"] = np.nan

    # Stable keys
    df["Address_Key"] = (
        df["Muni"].str.upper() + "|" +
        df["Zip_Code"].str.upper() + "|" +
        df["Full_HouseNo"].str.upper() + "|" +
        df["Street_Full"].str.upper() + "|" +
        df["Unit_Normalized"].str.upper()
    ).str.replace(r"\s+", " ", regex=True)

    df["Building_Key"] = (
        df["Muni"].str.upper() + "|" +
        df["Zip_Code"].str.upper() + "|" +
        df["Full_HouseNo"].str.upper() + "|" +
        df["Street_Full"].str.upper()
    ).str.replace(r"\s+", " ", regex=True)

    df["Territory_Building_Key"] = df["Territory_Key"].astype(str) + "|" + df["Building_Key"].astype(str)
    df["Unit_Key"] = df["Building_Key"].astype(str) + "|" + df["Unit_Normalized"].astype(str)

    # Duplicate detection among assigned source records.
    assigned_mask = df["Territory_Name"].ne("")
    duplicate_address_key = df.loc[assigned_mask, "Address_Key"].duplicated(keep=False)
    duplicate_keys = set(df.loc[assigned_mask].loc[duplicate_address_key, "Address_Key"])
    df["Duplicate_Address_Key"] = df["Address_Key"].isin(duplicate_keys) & df["Address_Key"].ne("||||")

    if duplicate_keys:
        add_issue(
            issues,
            "WARNING",
            "DUPLICATE_ADDRESS_KEYS",
            "Duplicate normalized address keys were found in assigned source records.",
            len(duplicate_keys),
            "Review duplicate rows in Address List / Excluded Audit.",
        )

    # Data quality flags
    flags = []
    for _, row in df.iterrows():
        row_flags = []
        if not clean_scalar(row.get("Territory_Name", "")):
            row_flags.append("Unassigned to territory")
        if int(row.get("Matched_Territory_Count", 0) or 0) > 1:
            row_flags.append("Assigned to multiple territories")
        if not clean_scalar(row.get("HouseNo_Display", "")):
            row_flags.append("Missing house number")
        if not clean_scalar(row.get("Street_Name", "")):
            row_flags.append("Missing street name")
        if not clean_scalar(row.get("Street_Full", "")):
            row_flags.append("Missing full street")
        if not clean_scalar(row.get("Muni", "")):
            row_flags.append("Missing municipality")
        if not clean_scalar(row.get("Zip_Code", "")):
            row_flags.append("Missing ZIP")
        if not clean_scalar(row.get("Raw_Address_Status", "")):
            row_flags.append("Missing address status")
        if bool(row.get("Duplicate_Address_Key", False)):
            row_flags.append("Duplicate normalized address")
        if "nan" in clean_scalar(row.get("Mailable_Address", "")).lower() or ",," in clean_scalar(row.get("Mailable_Address", "")):
            row_flags.append("Malformed mailable address")
        flags.append("; ".join(row_flags))

    df["Data_Quality_Flags"] = flags
    df["Review_Recommended"] = df["Data_Quality_Flags"].ne("") | df["Review_Recommended"].astype(bool)

    df["Assignment_Confidence"] = np.select(
        [
            df["Territory_Name"].eq(""),
            df["Matched_Territory_Count"].gt(1),
            df["Data_Quality_Flags"].ne(""),
        ],
        ["Low", "Low", "Medium"],
        default="High",
    )

    return df


# ============================================================
# 8. ANALYSIS TABLES
# ============================================================


def classify_count(count: int, min_goal: int, max_goal: int) -> str:
    if count == 0:
        return "Empty"
    if count < min_goal:
        return "Undersized"
    if count <= max_goal:
        return "Ideal"
    return "Oversized"


def suggested_action(row: pd.Series, min_goal: int, max_goal: int) -> str:
    status = clean_scalar(row.get("Status", ""))
    count = int(row.get("# of Addresses", 0) or 0)
    apt_units = int(row.get("Apartment Units", 0) or 0)

    if status == "Empty":
        return "Review KML boundary / no valid addresses found"
    if status == "Undersized":
        return "Candidate to expand from adjacent oversized territory"
    if status == "Ideal":
        return "No immediate balancing action needed"
    if status == "Oversized" and apt_units >= max(10, (max_goal - min_goal + 1)):
        return "Review apartment-heavy territory before redrawing"
    if status == "Oversized":
        return "Candidate to shrink, split, or redraw"
    return "Review"


def detect_apartments(valid_gdf: gpd.GeoDataFrame, apartment_threshold: int) -> pd.DataFrame:
    if valid_gdf.empty:
        return pd.DataFrame(columns=[
            "Territory Key", "Territory Name", "Building Key", "Base Address", "Total Units",
            "Unit Examples", "Muni", "Zip_Code", "Territory Status", "Territory Count",
            "Count After Removing Building", "Suggested Handling", "Confidence"
        ])

    group_cols = ["Territory_Key", "Territory_Name", "Building_Key", "Base_Address", "Muni", "Zip_Code"]
    groups = (
        valid_gdf.groupby(group_cols, dropna=False, observed=True)
        .agg(
            **{
                "Total Units": ("Address_Key", "size"),
                "Unit Examples": ("Unit_Normalized", lambda s: safe_unique_join(s, 10)),
                "Rows With Flags": ("Data_Quality_Flags", lambda s: int((s.fillna("").astype(str).str.strip() != "").sum())),
            }
        )
        .reset_index()
    )

    groups = groups[groups["Total Units"] >= apartment_threshold].copy()
    if groups.empty:
        return pd.DataFrame(columns=[
            "Territory Key", "Territory Name", "Building Key", "Base Address", "Total Units",
            "Unit Examples", "Muni", "Zip_Code", "Rows With Flags", "Territory Status", "Territory Count",
            "Count After Removing Building", "Suggested Handling", "Confidence"
        ])

    groups = groups.rename(columns={
        "Territory_Key": "Territory Key",
        "Territory_Name": "Territory Name",
        "Building_Key": "Building Key",
    })

    groups["Confidence"] = np.where(groups["Rows With Flags"].gt(0), "Medium", "High")
    groups["Suggested Handling"] = np.select(
        [
            groups["Total Units"].ge(50),
            groups["Total Units"].ge(20),
        ],
        [
            "Major apartment block; consider separate letter-writing handling",
            "Review as possible apartment-heavy territory adjustment",
        ],
        default="Review building grouping before action",
    )
    groups = groups.sort_values(["Total Units", "Territory Name"], ascending=[False, True])
    return groups


def calculate_counts(
    kml_gdf: gpd.GeoDataFrame,
    valid_gdf: gpd.GeoDataFrame,
    apartments_df: pd.DataFrame,
    min_goal: int,
    max_goal: int,
) -> pd.DataFrame:
    territory_base = kml_gdf[["Territory_Key", "Territory_Name", "Territory_Number", "Territory_CategoryCode"]].copy()

    counts = (
        valid_gdf.groupby(["Territory_Key", "Territory_Name"], observed=True)
        .size()
        .reset_index(name="# of Addresses")
        if not valid_gdf.empty else pd.DataFrame(columns=["Territory_Key", "Territory_Name", "# of Addresses"])
    )

    result = territory_base.merge(counts[["Territory_Key", "# of Addresses"]], on="Territory_Key", how="left")
    result["# of Addresses"] = result["# of Addresses"].fillna(0).astype(int)
    result["Status"] = result["# of Addresses"].map(lambda x: classify_count(x, min_goal, max_goal))
    result["Target Min"] = min_goal
    result["Target Max"] = max_goal
    result["Difference From Target Min"] = result["# of Addresses"] - min_goal
    result["Difference From Target Max"] = result["# of Addresses"] - max_goal

    if apartments_df is not None and not apartments_df.empty:
        apt_summary = (
            apartments_df.groupby("Territory Key")
            .agg(**{
                "Apartment Units": ("Total Units", "sum"),
                "Apartment Buildings": ("Building Key", "nunique"),
            })
            .reset_index()
        )
        result = result.merge(apt_summary, left_on="Territory_Key", right_on="Territory Key", how="left")
        result = result.drop(columns=[c for c in ["Territory Key"] if c in result.columns])
    else:
        result["Apartment Units"] = 0
        result["Apartment Buildings"] = 0

    result["Apartment Units"] = result["Apartment Units"].fillna(0).astype(int)
    result["Apartment Buildings"] = result["Apartment Buildings"].fillna(0).astype(int)
    result["Non-Apartment Address Count"] = (result["# of Addresses"] - result["Apartment Units"]).clip(lower=0)
    result["Apartment Heavy"] = np.where(result["Apartment Units"] >= max(10, int(max_goal * 0.25)), "Yes", "No")
    result["Suggested Action"] = result.apply(lambda row: suggested_action(row, min_goal, max_goal), axis=1)

    # Priority: oversized furthest over, empty, undersized furthest under, then ideal.
    def priority(row):
        status = row["Status"]
        count = row["# of Addresses"]
        if status == "Oversized":
            return 100000 + (count - max_goal)
        if status == "Empty":
            return 90000
        if status == "Undersized":
            return 80000 + (min_goal - count)
        return 0

    result["Priority Score"] = result.apply(priority, axis=1)
    result = natural_sort_df(result, "Territory_Name")
    result["Priority Rank"] = result["Priority Score"].rank(method="first", ascending=False).astype(int)

    result = result.rename(columns={
        "Territory_Key": "Territory Key",
        "Territory_Name": "Territory Name",
        "Territory_Number": "Territory Number",
        "Territory_CategoryCode": "Territory CategoryCode",
    })

    ordered_cols = [
        "Territory Key", "Territory Name", "Territory Number", "Territory CategoryCode",
        "# of Addresses", "Status", "Target Min", "Target Max",
        "Difference From Target Min", "Difference From Target Max",
        "Apartment Units", "Apartment Buildings", "Non-Apartment Address Count",
        "Apartment Heavy", "Suggested Action", "Priority Rank", "Priority Score"
    ]
    return result[ordered_cols]


def build_address_list(valid_gdf: gpd.GeoDataFrame) -> pd.DataFrame:
    if valid_gdf.empty:
        return pd.DataFrame(columns=[
            "Territory Name", "Territory DisplayName", "Territory Number", "Territory CategoryCode",
            "Mailable Address", "Full House Number", "HouseNo Raw", "HouseNo Sort",
            "Street PreDir", "Street Name", "Street Type", "Street Full", "Unit", "Muni", "State", "Zip_Code",
            "Latitude", "Longitude", "Source Record ID", "Address Key", "Building Key", "Unit Key",
            "Type Suggestion", "Raw Address Status", "Data Quality Flags", "Assignment Confidence", "Matched Territories"
        ])

    df = valid_gdf.copy()
    df["Type_Suggestion"] = "House"
    apt_keys = df.groupby("Territory_Building_Key")["Address_Key"].transform("size") >= DEFAULT_APARTMENT_THRESHOLD
    df.loc[apt_keys, "Type_Suggestion"] = "Apartment Candidate"

    df = df.sort_values(
        by=["Territory_Name", "Zip_Code", "Street_Name", "HouseNo_Sort", "HouseSx", "Unit_Normalized"],
        kind="mergesort",
    )

    out = pd.DataFrame({
        "Territory Name": df["Territory_Name"],
        "Territory DisplayName": df["Territory_DisplayName"],
        "Territory Number": df["Territory_Number"],
        "Territory CategoryCode": df["Territory_CategoryCode"],
        "Mailable Address": df["Mailable_Address"],
        "Full House Number": df["Full_HouseNo"],
        "HouseNo Raw": df["HouseNo_Raw"],
        "HouseNo Sort": df["HouseNo_Sort"],
        "HouseSx": df["HouseSx"],
        "Street PreDir": df["Street_PreDir"],
        "Street Name": df["Street_Name"],
        "Street Type": df["Street_Type"],
        "Street Full": df["Street_Full"],
        "Unit": df["Unit_Normalized"],
        "Muni": df["Muni"],
        "State": df["State"],
        "Zip_Code": df["Zip_Code"],
        "Latitude": df["Latitude"],
        "Longitude": df["Longitude"],
        "Source Record ID": df["Source_Record_ID"],
        "Address Key": df["Address_Key"],
        "Building Key": df["Building_Key"],
        "Unit Key": df["Unit_Key"],
        "Type Suggestion": df["Type_Suggestion"],
        "Raw Address Status": df["Raw_Address_Status"],
        "Data Quality Flags": df["Data_Quality_Flags"],
        "Assignment Confidence": df["Assignment_Confidence"],
        "Matched Territories": df["Matched_Territories"],
    })
    return out


def build_excluded_audit(excluded_gdf: gpd.GeoDataFrame) -> pd.DataFrame:
    if excluded_gdf.empty:
        return pd.DataFrame(columns=[
            "Territory Name", "Mailable Address", "Raw Status", "Normalized Exclusion Reason",
            "Exclusion Rule", "Source Record ID", "Full House Number", "Street Full", "Unit", "Muni",
            "Zip_Code", "Latitude", "Longitude", "Data Quality Flags", "Review Recommended"
        ])

    df = excluded_gdf.sort_values(
        by=["Territory_Name", "Street_Name", "HouseNo_Sort", "Unit_Normalized"],
        kind="mergesort",
    )

    return pd.DataFrame({
        "Territory Name": df["Territory_Name"],
        "Mailable Address": df["Mailable_Address"],
        "Raw Status": df["Raw_Address_Status"],
        "Normalized Exclusion Reason": df["Exclusion_Reason"],
        "Exclusion Rule": df["Exclusion_Rule"],
        "Source Record ID": df["Source_Record_ID"],
        "Full House Number": df["Full_HouseNo"],
        "Street Full": df["Street_Full"],
        "Unit": df["Unit_Normalized"],
        "Muni": df["Muni"],
        "Zip_Code": df["Zip_Code"],
        "Latitude": df["Latitude"],
        "Longitude": df["Longitude"],
        "Data Quality Flags": df["Data_Quality_Flags"],
        "Review Recommended": np.where(df["Review_Recommended"], "Yes", "No"),
    })


def build_spatial_qa_table(normalized_gdf: gpd.GeoDataFrame, spatial_qa_raw: pd.DataFrame) -> pd.DataFrame:
    qa = normalized_gdf[[
        "Source_Record_ID", "Territory_Name", "Mailable_Address", "Matched_Territory_Count",
        "Matched_Territories", "Data_Quality_Flags", "Assignment_Confidence"
    ]].copy()
    qa["QA Status"] = np.select(
        [
            qa["Territory_Name"].eq(""),
            qa["Matched_Territory_Count"].gt(1),
            qa["Data_Quality_Flags"].ne(""),
        ],
        ["Unassigned", "Matched Multiple Territories", "Review Flags"],
        default="Assigned",
    )
    qa = qa.rename(columns={
        "Source_Record_ID": "Source Record ID",
        "Territory_Name": "Chosen Territory",
        "Mailable_Address": "Mailable Address",
        "Matched_Territory_Count": "Matched Territory Count",
        "Matched_Territories": "Matched Territories",
        "Data_Quality_Flags": "Data Quality Flags",
        "Assignment_Confidence": "Assignment Confidence",
    })
    qa = qa[qa["QA Status"].ne("Assigned")].copy()
    return qa.sort_values(["QA Status", "Chosen Territory", "Mailable Address"], kind="mergesort")


def flag_counts_from_address_table(address_df: pd.DataFrame, excluded_df: pd.DataFrame, spatial_qa_df: pd.DataFrame) -> pd.DataFrame:
    flags = []
    for label, df, col in [
        ("Valid Address List", address_df, "Data Quality Flags"),
        ("Excluded Audit", excluded_df, "Data Quality Flags"),
        ("Spatial QA", spatial_qa_df, "Data Quality Flags"),
    ]:
        if df is None or df.empty or col not in df.columns:
            continue
        for value in df[col].fillna("").astype(str):
            for flag in [f.strip() for f in value.split(";") if f.strip()]:
                flags.append((label, flag))

    if not flags:
        return pd.DataFrame(columns=["Source", "Warning Type", "Count"])

    return (
        pd.DataFrame(flags, columns=["Source", "Warning Type"])
        .groupby(["Source", "Warning Type"])
        .size()
        .reset_index(name="Count")
        .sort_values(["Count", "Source", "Warning Type"], ascending=[False, True, True])
    )


def estimate_metric_crs(gdf: gpd.GeoDataFrame):
    try:
        return gdf.estimate_utm_crs()
    except Exception:
        return "EPSG:3857"


def generate_border_candidates(
    kml_gdf: gpd.GeoDataFrame,
    valid_gdf: gpd.GeoDataFrame,
    counts_df: pd.DataFrame,
    min_goal: int,
    max_goal: int,
    near_distance_m: int = DEFAULT_BORDER_DISTANCE_METERS,
) -> pd.DataFrame:
    if kml_gdf.empty or counts_df.empty:
        return pd.DataFrame(columns=[
            "Oversized Territory", "Oversized Count", "Undersized Neighbor", "Undersized Count",
            "Address Difference", "Shared Border Length (m)", "Candidate Addresses Near Border",
            "Nearby Streets", "Confidence", "Human Review Required", "Recommendation"
        ])

    counts_lookup = counts_df.set_index("Territory Name")["# of Addresses"].to_dict()
    status_lookup = counts_df.set_index("Territory Name")["Status"].to_dict()

    oversized_names = [name for name, status in status_lookup.items() if status == "Oversized"]
    undersized_names = [name for name, status in status_lookup.items() if status in {"Undersized", "Empty"}]

    if not oversized_names or not undersized_names:
        return pd.DataFrame(columns=[
            "Oversized Territory", "Oversized Count", "Undersized Neighbor", "Undersized Count",
            "Address Difference", "Shared Border Length (m)", "Candidate Addresses Near Border",
            "Nearby Streets", "Confidence", "Human Review Required", "Recommendation"
        ])

    metric_crs = estimate_metric_crs(kml_gdf)
    territories = kml_gdf[["Territory_Name", "geometry"]].to_crs(metric_crs).copy()
    territories = territories.set_index("Territory_Name", drop=False)

    if valid_gdf.empty:
        valid_metric = valid_gdf
    else:
        valid_metric = valid_gdf.to_crs(metric_crs).copy()

    suggestions = []
    sindex = territories.sindex

    for over_name in oversized_names:
        if over_name not in territories.index:
            continue

        over_geom = territories.loc[over_name, "geometry"]
        possible_idx = list(sindex.intersection(over_geom.bounds))
        candidates = territories.iloc[possible_idx]

        for _, under_row in candidates.iterrows():
            under_name = under_row["Territory_Name"]
            if under_name == over_name or under_name not in undersized_names:
                continue

            under_geom = under_row.geometry
            if not over_geom.intersects(under_geom):
                continue

            try:
                shared = over_geom.boundary.intersection(under_geom.boundary)
                shared_len = float(shared.length) if not shared.is_empty else 0.0
            except Exception:
                shared = None
                shared_len = 0.0

            if shared_len <= 0:
                # Intersections without shared boundary are usually overlaps/corner touches.
                confidence = "Low"
            elif shared_len < 25:
                confidence = "Medium"
            else:
                confidence = "High"

            nearby_count = 0
            nearby_streets = ""
            if shared is not None and not shared.is_empty and not valid_metric.empty:
                over_points = valid_metric[valid_metric["Territory_Name"].eq(over_name)].copy()
                if not over_points.empty:
                    try:
                        near_mask = over_points.geometry.distance(shared) <= near_distance_m
                        near_rows = over_points.loc[near_mask]
                        nearby_count = len(near_rows)
                        nearby_streets = safe_unique_join(near_rows.get("Street_Full", pd.Series(dtype=str)), 8)
                    except Exception:
                        nearby_count = 0
                        nearby_streets = ""

            over_count = int(counts_lookup.get(over_name, 0))
            under_count = int(counts_lookup.get(under_name, 0))
            diff = abs(over_count - under_count)

            recommendation = (
                f"Review candidate only: {over_name} is over target and touches {under_name}. "
                f"Consider whether addresses near the shared border could be reassigned."
            )

            suggestions.append({
                "Oversized Territory": over_name,
                "Oversized Count": over_count,
                "Undersized Neighbor": under_name,
                "Undersized Count": under_count,
                "Address Difference": diff,
                "Shared Border Length (m)": round(shared_len, 1),
                f"Candidate Addresses Within {near_distance_m}m": nearby_count,
                "Nearby Streets": nearby_streets,
                "Confidence": confidence,
                "Human Review Required": "Yes",
                "Recommendation": recommendation,
            })

    if not suggestions:
        return pd.DataFrame(columns=[
            "Oversized Territory", "Oversized Count", "Undersized Neighbor", "Undersized Count",
            "Address Difference", "Shared Border Length (m)", f"Candidate Addresses Within {near_distance_m}m",
            "Nearby Streets", "Confidence", "Human Review Required", "Recommendation"
        ])

    return pd.DataFrame(suggestions).sort_values(
        ["Confidence", "Address Difference", "Shared Border Length (m)"],
        ascending=[True, False, False],
        kind="mergesort",
    )


# ============================================================
# 9. EXCEL OUTPUT
# ============================================================


def write_df(writer, sheet_name: str, df: pd.DataFrame, table_name: Optional[str] = None, index: bool = False) -> None:
    safe_name = make_safe_sheet_name(sheet_name)
    if df is None or df.empty:
        df = pd.DataFrame(columns=df.columns if df is not None else ["Notice"])
    df.to_excel(writer, sheet_name=safe_name, index=index)
    ws = writer.sheets[safe_name]
    ws.freeze_panes = "A2"
    style_worksheet(ws)
    if table_name and ws.max_row >= 2 and ws.max_column >= 1:
        add_excel_table(ws, sanitize_table_name(table_name))


def style_worksheet(ws) -> None:
    thin = Side(style="thin", color="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)

    # Reasonable widths based on header names, capped for large sheets.
    for idx, cell in enumerate(ws[1], start=1):
        header = clean_scalar(cell.value)
        width = min(max(len(header) + 4, 12), 45)
        if header in {"Mailable Address", "Recommendation", "Suggested Action", "Data Quality Flags", "Matched Territories"}:
            width = 55
        elif "Key" in header or "ID" in header:
            width = 28
        elif header in {"Latitude", "Longitude"}:
            width = 14
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.auto_filter.ref = ws.dimensions

    # Text formatting for ZIPs and key/ID-like columns.
    text_headers = {"Zip_Code", "PostalCode", "Territory Number", "Source Record ID", "Address Key", "Building Key", "Unit Key"}
    for idx, cell in enumerate(ws[1], start=1):
        header = clean_scalar(cell.value)
        if header in text_headers or "Key" in header or "ID" in header:
            col_letter = get_column_letter(idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = "@"


def add_excel_table(ws, table_name: str) -> None:
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    try:
        ws.add_table(table)
    except ValueError:
        # If a name collision happens, append a short suffix.
        table.displayName = sanitize_table_name(f"{table_name}_{md5_short(table_name, 4)}")
        ws.add_table(table)


def style_counts_sheet(ws) -> None:
    headers = {clean_scalar(cell.value): idx for idx, cell in enumerate(ws[1], start=1)}
    status_col = headers.get("Status")
    if not status_col:
        return
    col_letter = get_column_letter(status_col)
    for row in range(2, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        color = STATUS_COLOR.get(clean_scalar(cell.value))
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def build_dashboard_sheet(
    writer,
    cong_name: str,
    run_config: Dict[str, object],
    counts_df: pd.DataFrame,
    address_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    apartments_df: pd.DataFrame,
    border_df: pd.DataFrame,
    spatial_qa_df: pd.DataFrame,
    flag_summary_df: pd.DataFrame,
    issues: List[ValidationIssue],
) -> None:
    wb = writer.book
    ws = wb.create_sheet("Dashboard", 0)

    ws["A1"] = f"Territory Analysis: {cong_name}"
    ws["A1"].font = Font(size=20, bold=True, color="1F4E3D")
    ws["A2"] = f"Generated {run_config['Generated At']} by {APP_NAME} ({APP_VERSION})"
    ws["A2"].font = Font(italic=True, color="666666")

    metrics = [
        ("County", run_config.get("County", "")),
        ("Goal Range Used", f"{run_config.get('Target Min')}–{run_config.get('Target Max')}"),
        ("Total Territories", len(counts_df)),
        ("Total Valid Addresses", len(address_df)),
        ("Excluded Addresses", len(excluded_df)),
        ("Apartment Groups", len(apartments_df)),
        ("Border Rewrite Candidates", len(border_df)),
        ("Spatial QA Rows", len(spatial_qa_df)),
        ("Fatal Errors", sum(1 for i in issues if i.severity == "ERROR")),
        ("Warnings", sum(1 for i in issues if i.severity == "WARNING")),
    ]

    status_counts = counts_df["Status"].value_counts().to_dict() if not counts_df.empty else {}
    metrics.extend([
        ("Ideal Territories", status_counts.get("Ideal", 0)),
        ("Undersized Territories", status_counts.get("Undersized", 0)),
        ("Oversized Territories", status_counts.get("Oversized", 0)),
        ("Empty Territories", status_counts.get("Empty", 0)),
    ])

    if not counts_df.empty:
        largest = counts_df.loc[counts_df["# of Addresses"].idxmax()]
        smallest = counts_df.loc[counts_df["# of Addresses"].idxmin()]
        metrics.extend([
            ("Largest Territory", f"{largest['Territory Name']} ({largest['# of Addresses']})"),
            ("Smallest Territory", f"{smallest['Territory Name']} ({smallest['# of Addresses']})"),
        ])

    start_row = 4
    ws.cell(row=start_row, column=1, value="Run Summary")
    ws.cell(row=start_row, column=1).font = Font(size=14, bold=True, color="1F4E3D")
    for i, (label, value) in enumerate(metrics, start=start_row + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=1).font = Font(bold=True)

    # Recommended next actions
    action_row = start_row + len(metrics) + 3
    ws.cell(row=action_row, column=1, value="Recommended Next Actions")
    ws.cell(row=action_row, column=1).font = Font(size=14, bold=True, color="1F4E3D")
    actions = [
        "Review the Counts tab by Priority Rank.",
        "Review apartment-heavy territories before making border changes.",
        "Review Data Quality Flags in Address List and Excluded Audit.",
        "Review Spatial QA for unassigned or multi-assigned address points.",
        "Treat Border Rewrites as review candidates, not automatic instructions.",
    ]
    for idx, action in enumerate(actions, start=action_row + 1):
        ws.cell(row=idx, column=1, value=idx - action_row)
        ws.cell(row=idx, column=2, value=action)

    # Top oversized / undersized / apartment-heavy blocks
    right_col = 4
    top_row = 4
    ws.cell(row=top_row, column=right_col, value="Top Oversized Territories")
    ws.cell(row=top_row, column=right_col).font = Font(size=14, bold=True, color="1F4E3D")
    oversized = counts_df[counts_df["Status"].eq("Oversized")].sort_values("# of Addresses", ascending=False).head(10)
    write_small_table(ws, oversized[["Territory Name", "# of Addresses", "Difference From Target Max", "Suggested Action"]], top_row + 1, right_col)

    under_row = top_row + 14
    ws.cell(row=under_row, column=right_col, value="Top Undersized / Empty Territories")
    ws.cell(row=under_row, column=right_col).font = Font(size=14, bold=True, color="1F4E3D")
    undersized = counts_df[counts_df["Status"].isin(["Undersized", "Empty"])].sort_values("# of Addresses", ascending=True).head(10)
    write_small_table(ws, undersized[["Territory Name", "# of Addresses", "Status", "Suggested Action"]], under_row + 1, right_col)

    apt_row = under_row + 14
    ws.cell(row=apt_row, column=right_col, value="Top Apartment Groups")
    ws.cell(row=apt_row, column=right_col).font = Font(size=14, bold=True, color="1F4E3D")
    apt_small = apartments_df.head(10) if apartments_df is not None and not apartments_df.empty else pd.DataFrame(columns=["Territory Name", "Base Address", "Total Units"])
    cols = [c for c in ["Territory Name", "Base Address", "Total Units", "Suggested Handling"] if c in apt_small.columns]
    write_small_table(ws, apt_small[cols], apt_row + 1, right_col)

    # Data quality summary
    dq_row = action_row + len(actions) + 4
    ws.cell(row=dq_row, column=1, value="Data Quality Summary")
    ws.cell(row=dq_row, column=1).font = Font(size=14, bold=True, color="1F4E3D")
    dq_small = flag_summary_df.head(15) if flag_summary_df is not None and not flag_summary_df.empty else pd.DataFrame(columns=["Source", "Warning Type", "Count"])
    write_small_table(ws, dq_small, dq_row + 1, 1)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 24 if col != 2 else 60

    ws.sheet_properties.tabColor = "1F4E3D"


def write_small_table(ws, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    if df is None or df.empty:
        ws.cell(row=start_row, column=start_col, value="No records")
        return
    for c_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for c_idx, value in enumerate(row.tolist(), start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def build_run_log_df(run_config: Dict[str, object], issues: List[ValidationIssue]) -> pd.DataFrame:
    rows = []
    for key, value in run_config.items():
        rows.append({"Section": "Config", "Severity": "INFO", "Code": key, "Message": str(value), "Count": "", "Details": ""})
    for issue in issues:
        rows.append({
            "Section": "Validation",
            "Severity": issue.severity,
            "Code": issue.code,
            "Message": issue.message,
            "Count": issue.count,
            "Details": issue.details,
        })
    return pd.DataFrame(rows)


def generate_excel_report(
    cong_name: str,
    run_config: Dict[str, object],
    counts_df: pd.DataFrame,
    address_df: pd.DataFrame,
    apartments_df: pd.DataFrame,
    border_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    spatial_qa_df: pd.DataFrame,
    flag_summary_df: pd.DataFrame,
    issues: List[ValidationIssue],
) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Create data sheets first; dashboard is inserted as first sheet later.
        write_df(writer, "Counts", counts_df.drop(columns=["Priority Score"], errors="ignore"), "CountsTable")
        style_counts_sheet(writer.sheets["Counts"])

        write_df(writer, "Address List", address_df, "AddressListTable")
        write_df(writer, "Apartments", apartments_df, "ApartmentsTable")
        write_df(writer, "Border Rewrites", border_df, "BorderRewritesTable")
        write_df(writer, "Excluded Audit", excluded_df, "ExcludedAuditTable")
        write_df(writer, "Spatial QA", spatial_qa_df, "SpatialQATable")
        write_df(writer, "Data Quality", flag_summary_df, "DataQualityTable")
        run_log_df = build_run_log_df(run_config, issues)
        write_df(writer, "Run Log", run_log_df, "RunLogTable")

        build_dashboard_sheet(
            writer,
            cong_name,
            run_config,
            counts_df,
            address_df,
            excluded_df,
            apartments_df,
            border_df,
            spatial_qa_df,
            flag_summary_df,
            issues,
        )

        # Tab colors
        tab_colors = {
            "Dashboard": "1F4E3D",
            "Counts": "70AD47",
            "Address List": "5B9BD5",
            "Apartments": "F4B183",
            "Border Rewrites": "C00000",
            "Excluded Audit": "808080",
            "Spatial QA": "7030A0",
            "Data Quality": "FFC000",
            "Run Log": "A5A5A5",
        }
        for sheet, color in tab_colors.items():
            if sheet in writer.sheets:
                writer.sheets[sheet].sheet_properties.tabColor = color

    output.seek(0)
    return output.getvalue()


# ============================================================
# 10. PIPELINE ORCHESTRATION
# ============================================================


def run_analysis_pipeline(
    uploaded_kml,
    county_name: str,
    congregation_name: str,
    min_goal: int,
    max_goal: int,
    apartment_threshold: int,
    border_distance_m: int,
) -> Tuple[Optional[bytes], Optional[str], Dict[str, object], List[ValidationIssue]]:
    issues: List[ValidationIssue] = []
    generated_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    issues.extend(validate_goal_range(min_goal, max_goal))
    if has_fatal_errors(issues):
        return None, None, {}, issues

    parcel_gdf = load_county_data(county_name)
    issues.extend(validate_county_schema(parcel_gdf))
    if has_fatal_errors(issues):
        return None, None, {}, issues

    kml_gdf = load_kml(uploaded_kml)
    issues.extend(validate_kml(kml_gdf))
    if has_fatal_errors(issues):
        return None, None, {}, issues

    address_points = prepare_address_points(parcel_gdf, issues)
    candidate_points, kml_for_join = subset_county_to_kml_extent(address_points, kml_gdf, issues)
    if has_fatal_errors(issues):
        return None, None, {}, issues

    assigned_single, spatial_qa_raw = assign_addresses_to_territories(candidate_points, kml_for_join, issues)
    normalized_gdf = normalize_milwaukee_addresses(assigned_single, issues)

    assigned_gdf = normalized_gdf[normalized_gdf["Territory_Name"].ne("")].copy()
    valid_gdf = assigned_gdf[~assigned_gdf["Is_Excluded"]].copy()
    excluded_gdf = assigned_gdf[assigned_gdf["Is_Excluded"]].copy()

    # Report zero-address territories explicitly through counts.
    apartments_df = detect_apartments(valid_gdf, apartment_threshold)
    counts_df = calculate_counts(kml_gdf, valid_gdf, apartments_df, min_goal, max_goal)

    # Enrich apartments with status/count context.
    if not apartments_df.empty:
        counts_context = counts_df[["Territory Key", "Status", "# of Addresses"]].rename(columns={
            "Status": "Territory Status",
            "# of Addresses": "Territory Count",
        })
        apartments_df = apartments_df.merge(counts_context, on="Territory Key", how="left")
        apartments_df["Count After Removing Building"] = (apartments_df["Territory Count"] - apartments_df["Total Units"]).clip(lower=0)
        apt_cols = [
            "Territory Key", "Territory Name", "Building Key", "Base Address", "Total Units",
            "Unit Examples", "Muni", "Zip_Code", "Rows With Flags", "Territory Status",
            "Territory Count", "Count After Removing Building", "Suggested Handling", "Confidence"
        ]
        apartments_df = apartments_df[[c for c in apt_cols if c in apartments_df.columns]]

    address_df = build_address_list(valid_gdf)
    # Recalculate type suggestion with actual selected threshold, not the default.
    if not address_df.empty and not valid_gdf.empty:
        actual_apt_keys = valid_gdf.groupby("Territory_Building_Key")["Address_Key"].transform("size") >= apartment_threshold
        type_series = pd.Series("House", index=valid_gdf.index)
        type_series.loc[actual_apt_keys] = "Apartment Candidate"
        type_map = dict(zip(valid_gdf["Address_Key"], type_series))
        address_df["Type Suggestion"] = address_df["Address Key"].map(type_map).fillna(address_df["Type Suggestion"])

    excluded_df = build_excluded_audit(excluded_gdf)
    spatial_qa_df = build_spatial_qa_table(normalized_gdf, spatial_qa_raw)
    border_df = generate_border_candidates(kml_for_join, valid_gdf, counts_df, min_goal, max_goal, border_distance_m)
    flag_summary_df = flag_counts_from_address_table(address_df, excluded_df, spatial_qa_df)

    # Extra validation summaries after processing.
    empty_territories = int(counts_df["Status"].eq("Empty").sum()) if not counts_df.empty else 0
    if empty_territories:
        add_issue(
            issues,
            "WARNING",
            "ZERO_ADDRESS_TERRITORIES",
            "Some territories have zero valid assigned addresses.",
            empty_territories,
            "Review Counts and Spatial QA tabs.",
        )

    rows_with_flags = int(address_df["Data Quality Flags"].fillna("").astype(str).str.strip().ne("").sum()) if not address_df.empty else 0
    if rows_with_flags:
        add_issue(
            issues,
            "WARNING",
            "ADDRESS_ROWS_WITH_FLAGS",
            "Some valid address rows have data quality flags.",
            rows_with_flags,
            "Review Address List and Data Quality tabs.",
        )

    run_config = {
        "App Version": APP_VERSION,
        "Generated At": generated_at,
        "Congregation": congregation_name,
        "County": county_name,
        "County Data": MILWAUKEE_PROFILE["county_data_label"],
        "KML Filename": getattr(uploaded_kml, "name", "uploaded.kml"),
        "Target Min": min_goal,
        "Target Max": max_goal,
        "Apartment Threshold": apartment_threshold,
        "Border Near Distance Meters": border_distance_m,
        "Spatial Predicate": MILWAUKEE_PROFILE["spatial_predicate"],
        "Geometry Contract": MILWAUKEE_PROFILE["geometry_contract"],
        "County CRS": str(parcel_gdf.crs),
        "KML CRS": str(kml_gdf.crs),
        "Candidate Records In Extent": len(candidate_points),
        "Assigned Source Records": len(assigned_gdf),
        "Valid Address Records": len(valid_gdf),
        "Excluded Assigned Records": len(excluded_gdf),
        "Unassigned Candidate Records": int(normalized_gdf["Territory_Name"].eq("").sum()),
    }

    excel_bytes = generate_excel_report(
        congregation_name,
        run_config,
        counts_df,
        address_df,
        apartments_df,
        border_df,
        excluded_df,
        spatial_qa_df,
        flag_summary_df,
        issues,
    )

    clean_name = re.sub(r"[^A-Za-z0-9_-]", "", congregation_name.replace(" ", "")) or "Congregation"
    filename = f"{clean_name}_{dt.datetime.now().strftime('%B%Y')}_TerritoryAnalysis.xlsx"

    # Reduce memory pressure in Streamlit Cloud.
    del parcel_gdf, address_points, candidate_points, assigned_single, normalized_gdf, assigned_gdf, valid_gdf, excluded_gdf
    gc.collect()

    return excel_bytes, filename, run_config, issues


# ============================================================
# 11. STREAMLIT UI
# ============================================================


def clear_stale_results_if_inputs_changed(current_key: str) -> None:
    previous_key = st.session_state.get("input_signature")
    if previous_key != current_key:
        for key in ["excel_data", "excel_filename", "run_config", "issues"]:
            st.session_state.pop(key, None)
        st.session_state["input_signature"] = current_key


def uploaded_file_signature(uploaded_file) -> str:
    if uploaded_file is None:
        return "NO_FILE"
    return f"{getattr(uploaded_file, 'name', '')}|{getattr(uploaded_file, 'size', '')}"


def render_issues(issues: List[ValidationIssue]) -> None:
    if not issues:
        return

    errors = [i for i in issues if i.severity == "ERROR"]
    warnings = [i for i in issues if i.severity == "WARNING"]
    infos = [i for i in issues if i.severity == "INFO"]

    if errors:
        st.error(f"{len(errors)} fatal issue(s) found. Report generation is blocked.")
        with st.expander("Fatal issues", expanded=True):
            st.dataframe(pd.DataFrame([i.to_dict() for i in errors]), use_container_width=True)

    if warnings:
        st.warning(f"{len(warnings)} warning(s) found. Review the workbook Run Log and Data Quality tabs.")
        with st.expander("Warnings", expanded=False):
            st.dataframe(pd.DataFrame([i.to_dict() for i in warnings]), use_container_width=True)

    if infos:
        with st.expander("Processing notes", expanded=False):
            st.dataframe(pd.DataFrame([i.to_dict() for i in infos]), use_container_width=True)


def run_streamlit_app() -> None:
    st.set_page_config(page_title="TerritoryToolbox Milwaukee Analyzer", layout="wide", page_icon="🗺️")

    st.title("🗺️ TerritoryToolbox Milwaukee Analyzer")
    st.caption(APP_VERSION)
    st.markdown(
        "Upload a KML territory map and generate a source-of-truth Excel workbook with "
        "territory counts, address lists, apartments, border rewrite candidates, excluded audit, and QA logs."
    )

    st.sidebar.header("Step 1: Configuration")
    congregation_name = st.sidebar.text_input("Congregation Name", "ExampleCongregation")
    selected_county = st.sidebar.selectbox("Select County Data", ["Milwaukee"])

    st.sidebar.subheader("Territory Balance Target")
    min_goal = st.sidebar.number_input("Minimum ideal mailable addresses", min_value=1, max_value=1000, value=DEFAULT_MIN_GOAL, step=1)
    max_goal = st.sidebar.number_input("Maximum ideal mailable addresses", min_value=1, max_value=1000, value=DEFAULT_MAX_GOAL, step=1)

    st.sidebar.subheader("Analysis Rules")
    apartment_threshold = st.sidebar.number_input(
        "Apartment grouping threshold",
        min_value=2,
        max_value=1000,
        value=DEFAULT_APARTMENT_THRESHOLD,
        step=1,
        help="Buildings with this many units or more appear in the Apartments tab.",
    )
    border_distance_m = st.sidebar.number_input(
        "Border candidate distance (meters)",
        min_value=10,
        max_value=500,
        value=DEFAULT_BORDER_DISTANCE_METERS,
        step=10,
        help="Counts oversized-territory addresses near a shared border for rewrite review.",
    )

    st.header("Step 2: Upload Territory Map")
    uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

    current_key = "|".join([
        uploaded_file_signature(uploaded_kml),
        selected_county,
        congregation_name,
        str(min_goal),
        str(max_goal),
        str(apartment_threshold),
        str(border_distance_m),
        APP_VERSION,
    ])
    clear_stale_results_if_inputs_changed(current_key)

    with st.expander("What this version checks", expanded=False):
        st.markdown(
            """
            This Milwaukee-specific version adds stronger safeguards before treating the workbook as a source of truth:

            - required county-column validation
            - KML name and geometry validation
            - zero-address territory visibility
            - unassigned and multi-assigned address QA
            - address normalization with stable keys
            - ZIP preservation as text
            - data quality flags
            - safer apartment grouping
            - border rewrite candidates with shared-border length and nearby-address counts
            - Run Log, Spatial QA, and Data Quality tabs
            """
        )

    if uploaded_kml:
        if min_goal > max_goal:
            st.error("Minimum ideal addresses cannot be greater than maximum ideal addresses.")
            return

        if st.button("Generate Territory Analysis", type="primary"):
            with st.spinner("Running Milwaukee GIS analysis and workbook QA..."):
                try:
                    excel_data, filename, run_config, issues = run_analysis_pipeline(
                        uploaded_kml=uploaded_kml,
                        county_name=selected_county,
                        congregation_name=congregation_name.strip() or "Congregation",
                        min_goal=int(min_goal),
                        max_goal=int(max_goal),
                        apartment_threshold=int(apartment_threshold),
                        border_distance_m=int(border_distance_m),
                    )

                    render_issues(issues)

                    if excel_data and filename:
                        st.session_state["excel_data"] = excel_data
                        st.session_state["excel_filename"] = filename
                        st.session_state["run_config"] = run_config
                        st.session_state["issues"] = [i.to_dict() for i in issues]
                        st.success("Analysis complete. Review the warnings, then download the workbook.")
                    else:
                        st.error("Analysis did not produce a workbook because fatal validation errors were found.")

                except Exception as exc:
                    st.error(f"An unexpected processing error occurred: {exc}")

    if "excel_data" in st.session_state:
        st.divider()
        st.subheader("Download")

        config = st.session_state.get("run_config", {})
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Target Min", config.get("Target Min", ""))
        c2.metric("Target Max", config.get("Target Max", ""))
        c3.metric("Valid Addresses", config.get("Valid Address Records", ""))
        c4.metric("Excluded", config.get("Excluded Assigned Records", ""))

        st.download_button(
            label="⬇️ Download Excel Analysis",
            data=st.session_state["excel_data"],
            file_name=st.session_state["excel_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    run_streamlit_app()
