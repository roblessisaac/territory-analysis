import streamlit as st
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.table import Table, TableStyleInfo
import fiona
import io
import datetime
import re
from fractions import Fraction

# Enable KML support in GeoPandas
fiona.drvsupport.supported_drivers['KML'] = 'rw'
fiona.drvsupport.supported_drivers['LIBKML'] = 'rw'

COUNTY_CONFIGS = {
    "Milwaukee": {
        "file_path": "zip://data/Milwaukee_Datapoints07072026.zip",
        "state": "WI",
        "metric_crs": "EPSG:3071",
        "native_source_id": "TAXKEY",
        "excluded_statuses": [
            "Undeveloped",
            "Parking Lot",
            "ROW",
            "Park or Recreational Facility",
            "Undeveloped Outlot",
            "Sliver or Remnant",
            "Non Addressable Assoc with Adj Parcel",
        ],
        "column_mapping": {
            "TAXKEY": "Canonical_Native_Source_ID",
            "HouseNo": "Canonical_HouseNo",
            "HouseSx": "Canonical_HouseSx",
            "Dir": "Canonical_Dir",
            "Street": "Canonical_Street",
            "StType": "Canonical_StType",
            "Muni": "Canonical_Muni",
            "Zip_Code": "Canonical_Zip_Code",
            "Unit": "Canonical_Unit",
            "Addr_Statu": "Canonical_Status",
        },
    }
}

REQUIRED_CANONICAL_COLUMNS = [
    "Canonical_HouseNo",
    "Canonical_HouseSx",
    "Canonical_Dir",
    "Canonical_Street",
    "Canonical_StType",
    "Canonical_Muni",
    "Canonical_Zip_Code",
    "Canonical_Unit",
    "Canonical_Status",
    "geometry",
]

# --- 1. CONFIGURATION & UI SETUP ---
st.set_page_config(page_title="Territory Audit Engine", layout="wide")

st.title("Congregation Territory Analysis Engine")
st.markdown("Upload your territories KML map to generate a complete, filtered address database & analysis.")

st.sidebar.header("Step 1: Configuration")
congregation_name = st.sidebar.text_input("Congregation Name (No Spaces)", "ExampleCongregation")
selected_county = st.sidebar.selectbox(
    "Select County Data",
    list(COUNTY_CONFIGS.keys()),
)
goal_range = st.sidebar.selectbox("Goal # of Addresses Per Territory", 
                                  ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"])

st.header("Step 2: Upload Territory Map")
uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

MIN_GOAL, MAX_GOAL = [int(x) for x in goal_range.split("-")]

# --- 2. DATA LOADING & CACHING ---
@st.cache_data
def load_county_data(county_name):
    county_config = COUNTY_CONFIGS[county_name]
    file_path = county_config["file_path"]
    try:
        return gpd.read_file(file_path)
    except Exception as error:
        st.error(
            "Error loading county shapefile. Check the configured file path. "
            f"Error: {error}"
        )
        return None

def natural_keys(text):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

# --- ADDRESS BUILDER + NORMALIZATION HELPERS ---
def clean_field(value):
    if pd.isna(value): return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}: return ""
    return text

def normalize_house_number(value):
    text = clean_field(value)
    if not text: return ""
    if re.fullmatch(r"[+-]?\d+\.0+", text): return text.split(".", 1)[0]
    return re.sub(r"\s+", " ", text)

def normalize_zip_code(value):
    text = clean_field(value)
    if not text: return ""
    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D", "", text)
    if not digits: return ""
    if len(digits) == 4: digits = digits.zfill(5)
    elif len(digits) == 8: digits = digits.zfill(9)
    if len(digits) == 5: return digits
    if len(digits) == 9: return f"{digits[:5]}-{digits[5:]}"
    if len(digits) > 9: return f"{digits[:5]}-{digits[5:9]}"
    return digits

def normalize_unit(value):
    text = clean_field(value)
    if not text: return ""
    if re.fullmatch(r"\d+\.0+", text): text = text.split(".", 1)[0]
    text = re.sub(r"\s+", " ", text).strip()
    descriptive_pattern = re.compile(r"^(?:apt(?:artment)?|unit|ste|suite|upper|lower|bsmt|basement|rear|front|floor|fl|building|bldg|room|rm)\b", flags=re.IGNORECASE)
    if descriptive_pattern.search(text): return text
    return f"Apt {text}"

def house_number_sort_parts(value):
    text = normalize_house_number(value).upper()
    if not text: return pd.Series([float("inf"), 9, ""])
    compact = re.sub(r"\s+", " ", text).strip()
    mixed_fraction = re.match(r"^(\d+)\s+(\d+)\s*/\s*(\d+)(.*)$", compact)
    if mixed_fraction:
        whole, numerator, denominator, suffix = mixed_fraction.groups()
        try: numeric_value = int(whole) + float(Fraction(int(numerator), int(denominator)))
        except (ValueError, ZeroDivisionError): numeric_value = float(int(whole))
        return pd.Series([numeric_value, 1, suffix.strip()])
    simple_fraction = re.match(r"^(\d+)\s*/\s*(\d+)(.*)$", compact)
    if simple_fraction:
        numerator, denominator, suffix = simple_fraction.groups()
        try: numeric_value = float(Fraction(int(numerator), int(denominator)))
        except (ValueError, ZeroDivisionError): numeric_value = float("inf")
        return pd.Series([numeric_value, 1, suffix.strip()])
    numeric_prefix = re.match(r"^(\d+(?:\.\d+)?)(.*)$", compact)
    if numeric_prefix:
        number, suffix = numeric_prefix.groups()
        return pd.Series([float(number), 0 if not suffix.strip() else 2, suffix.strip()])
    return pd.Series([float("inf"), 8, compact])

def build_addresses(row, state):
    house = normalize_house_number(row.get("Canonical_HouseNo"))
    house_sx = clean_field(row.get("Canonical_HouseSx"))
    direction = clean_field(row.get("Canonical_Dir"))
    street = clean_field(row.get("Canonical_Street"))
    st_type = clean_field(row.get("Canonical_StType"))
    muni = clean_field(row.get("Canonical_Muni"))
    zip_c = normalize_zip_code(row.get("Canonical_Zip_Code"))
    unit_str = normalize_unit(row.get("Canonical_Unit"))

    full_house_num = f"{house}{house_sx}".strip()
    full_street = " ".join(part for part in [direction, street, st_type] if part)
    base_addr_line = " ".join(part for part in [full_house_num, full_street] if part)
    locality = ", ".join(part for part in [muni, state] if part)
    if zip_c: locality = f"{locality} {zip_c}".strip()

    base_addr = ", ".join(part for part in [base_addr_line, locality] if part)
    mailable_addr = ", ".join(part for part in [" ".join([base_addr_line, unit_str]).strip(), locality] if part)

    return pd.Series([base_addr, mailable_addr], index=["Base_Address", "Mailable_Address"])

def evaluate_data_quality(row):
    issues = []
    house = normalize_house_number(row.get("Canonical_HouseNo"))
    street = clean_field(row.get("Canonical_Street"))
    municipality = clean_field(row.get("Canonical_Muni"))
    zip_code = normalize_zip_code(row.get("Canonical_Zip_Code"))
    base_address = clean_field(row.get("Base_Address"))
    mailable_address = clean_field(row.get("Mailable_Address"))

    if not street: issues.append("Missing Street")
    if not municipality: issues.append("Missing Municipality")
    if not zip_code: issues.append("Missing ZIP")
    if house and re.fullmatch(r"[+-]?0+(?:\.0+)?", house): issues.append("Zero House Number")

    zip_is_valid = not zip_code or bool(re.fullmatch(r"\d{5}(?:-\d{4})?", zip_code))
    if not house or not base_address or not mailable_address or ",," in base_address or ",," in mailable_address or not zip_is_valid:
        issues.append("Malformed Address")
    return " | ".join(issues)

# --- 3. EXCEL GENERATION ENGINE ---
def generate_excel_report(joined_gdf, kml_gdf, min_goal, max_goal, cong_name, county_config):
    output = io.BytesIO()
    run_timestamp = datetime.datetime.now()
    state = county_config["state"]
    metric_crs = county_config["metric_crs"]
    excluded_statuses = county_config["excluded_statuses"]

    joined_gdf = joined_gdf.copy()
    joined_gdf["Canonical_Zip_Code"] = joined_gdf["Canonical_Zip_Code"].map(normalize_zip_code)
    joined_gdf[["Base_Address", "Mailable_Address"]] = joined_gdf.apply(lambda row: build_addresses(row, state), axis=1)
    joined_gdf["Data_Quality_Flag"] = joined_gdf.apply(evaluate_data_quality, axis=1)
    flagged_record_count = joined_gdf["Data_Quality_Flag"].ne("").sum()

    normalized_excluded_statuses = {clean_field(status).upper() for status in excluded_statuses}
    joined_gdf["Canonical_Status_Normalized"] = joined_gdf["Canonical_Status"].map(clean_field).str.upper()
    exclusion_mask = joined_gdf["Canonical_Status_Normalized"].isin(normalized_excluded_statuses)
    excluded_gdf = joined_gdf[exclusion_mask].copy()
    valid_gdf = joined_gdf[~exclusion_mask].copy()

    unique_territories = valid_gdf["Territory_Name"].unique().tolist()
    unique_territories.sort(key=natural_keys)
    valid_gdf["Territory_Name"] = pd.Categorical(valid_gdf["Territory_Name"], categories=unique_territories, ordered=True)

    if not excluded_gdf.empty:
        excluded_unique = excluded_gdf["Territory_Name"].unique().tolist()
        excluded_unique.sort(key=natural_keys)
        excluded_gdf["Territory_Name"] = pd.Categorical(excluded_gdf["Territory_Name"], categories=excluded_unique, ordered=True)

    counts_df = valid_gdf.groupby("Territory_Name", observed=True).size().reset_index(name="Total_Addresses")
    counts_df = counts_df[counts_df["Total_Addresses"] > 0].copy()

    def get_category(count):
        if count < min_goal: return "Undersized"
        if count <= max_goal: return "Ideal"
        return "Oversized"

    counts_df["Category"] = counts_df["Total_Addresses"].apply(get_category)

    valid_gdf[["NWS_Category", "NWS_Number"]] = valid_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
    valid_gdf["NWS_Category"] = valid_gdf["NWS_Category"].fillna("UNK")
    valid_gdf["NWS_Number"] = valid_gdf["NWS_Number"].fillna("0")

    if not excluded_gdf.empty:
        excluded_gdf[["NWS_Category", "NWS_Number"]] = excluded_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
        excluded_gdf["NWS_Category"] = excluded_gdf["NWS_Category"].fillna("UNK")
        excluded_gdf["NWS_Number"] = excluded_gdf["NWS_Number"].fillna("0")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        
        # --- TAB 1: DASHBOARD ---
        total_territories = len(counts_df)
        total_addresses = counts_df["Total_Addresses"].sum()
        largest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmax()] if total_territories > 0 else None
        smallest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmin()] if total_territories > 0 else None
        ideal_pct = (len(counts_df[counts_df["Category"] == "Ideal"]) / total_territories * 100) if total_territories > 0 else 0

        largest_name = largest_terr["Territory_Name"] if largest_terr is not None else ""
        largest_count = largest_terr["Total_Addresses"] if largest_terr is not None else 0
        smallest_name = smallest_terr["Territory_Name"] if smallest_terr is not None else ""
        smallest_count = smallest_terr["Total_Addresses"] if smallest_terr is not None else 0

        dashboard_top = [
            [f"Territory Analysis: {cong_name}"],
            [f"Generated {run_timestamp.strftime('%B %Y')} by Territory Analysis Engine."],
            [""],
            [f"Total Territories: {total_territories}"],
            [f"Total Valid Addresses: {total_addresses}"],
            [f"Excluded Addresses (See Tab 6): {len(excluded_gdf)}"],
            [f"The largest territory has {largest_count} addresses in it ({largest_name})."],
            [f"The smallest territory has {smallest_count} addresses in it ({smallest_name})."],
            [""],
            [f"Goal Range: {min_goal}-{max_goal}"],
            [""]
        ]
        pd.DataFrame(dashboard_top).to_excel(writer, sheet_name="Dashboard", index=False, header=False)

        fixed_bins = [
            ("Under 25", counts_df["Total_Addresses"] < 25),
            ("25-49", counts_df["Total_Addresses"].between(25, 49, inclusive="both")),
            ("50-74", counts_df["Total_Addresses"].between(50, 74, inclusive="both")),
            ("75-99", counts_df["Total_Addresses"].between(75, 99, inclusive="both")),
            ("100-125", counts_df["Total_Addresses"].between(100, 125, inclusive="both")),
            ("126-150", counts_df["Total_Addresses"].between(126, 150, inclusive="both")),
            ("151-175", counts_df["Total_Addresses"].between(151, 175, inclusive="both")),
            ("Over 175", counts_df["Total_Addresses"] > 175)
        ]

        distribution = []
        for range_label, range_mask in fixed_bins:
            range_rows = counts_df.loc[range_mask]
            range_count = len(range_rows)
            range_categories = range_rows["Category"].dropna().astype(str).unique().tolist()
            
            if len(range_categories) == 1: range_category = range_categories[0]
            elif len(range_categories) > 1: range_category = "Mixed"
            else:
                if range_label == "Under 25": r_min, r_max = 0, 24
                elif range_label == "Over 175": r_min, r_max = 176, float("inf")
                else: r_min, r_max = [int(v) for v in range_label.split("-")]
                
                if r_max < min_goal: range_category = "Undersized"
                elif r_min > max_goal: range_category = "Oversized"
                elif r_min >= min_goal and r_max <= max_goal: range_category = "Ideal"
                else: range_category = "Mixed"
            distribution.append([range_category, range_label, range_count])

        pd.DataFrame(distribution, columns=["Category", "Range", "Count"]).to_excel(writer, sheet_name="Dashboard", startrow=11, index=False)

        ws1 = writer.sheets["Dashboard"]
        ws1.column_dimensions["A"].width = 18
        ws1.column_dimensions["B"].width = 25
        ws1.column_dimensions["C"].width = 12

        ws1["A1"].font = Font(size=20, bold=True)
        ws1["A2"].hyperlink = "http://www.territoryanalysis.com/"
        ws1["A2"].font = Font(color="0563C1", underline="single")

        bold_inline = InlineFont(b=True)
        ws1["A11"].value = CellRichText(["About ", TextBlock(bold_inline, f"{ideal_pct:.1f}%"), " of territories fall within this range."])

        header_fill = PatternFill(start_color="C7CDDB", end_color="C7CDDB", fill_type="solid")
        for col in range(1, 4):
            ws1.cell(row=12, column=col).fill = header_fill
            ws1.cell(row=12, column=col).font = Font(bold=True)

        distribution_end_row = 12 + len(distribution)
        for row_number in range(13, distribution_end_row + 1):
            if ws1.cell(row=row_number, column=1).value == "Ideal":
                for col in range(1, 4): ws1.cell(row=row_number, column=col).font = Font(bold=True)

        instruction_start = distribution_end_row + 2
        instructions = [
            CellRichText(["As a part of this analysis, every ", TextBlock(bold_inline, "address point"), " within your territory was collected & identified."]),
            "These addresses, with a little reformatting, can be added to NWS or other programs (Please see http://www.territoryanalysis.com/ to see if your system is supported.)",
            "It's suggested to export this file into a program you can easily edit, like excel or google sheets.",
            "That will allow you to expand cells to read easier, create custom filters to see specific data, and customize the sheet to make it more legible.",
            "",
            CellRichText(["The ", TextBlock(bold_inline, "DASHBOARD"), " tab displays basic statistics about the territory that was analyzed"]),
            CellRichText(["The ", TextBlock(bold_inline, "COUNTS"), " tab organizes territories by size. This is done by 'counting' workable addresses, not geographical size."]),
            CellRichText(["The ", TextBlock(bold_inline, "ADDRESS LIST"), " tab displays every workable address in your territory."]),
            CellRichText(["The ", TextBlock(bold_inline, "APARTMENTS"), " tab displays every multifamily above 5 units in your territory. Large units can be explanations for inflated door-to-door territories."]),
            CellRichText(["The ", TextBlock(bold_inline, "BORDER REWRITES"), " tab displays borders within your territory that may benefit from being redrawn. The intent is to shrink oversized territories adjacent to undersized territories. These are just suggestions."]),
            CellRichText(["The ", TextBlock(bold_inline, "EXCLUDED AUDIT"), " tab displays addresses that are NOT counted towards your territory. These are usually addresses of highways, vacant lots, parks, etc. This is included for confidence."])
        ]

        for offset, instruction in enumerate(instructions):
            row_number = instruction_start + offset
            cell = ws1.cell(row=row_number, column=1, value=instruction)
            if "http" in str(instruction):
                cell.hyperlink = "http://www.territoryanalysis.com/"
                cell.font = Font(color="0563C1", underline="single")

        technical_start = instruction_start + len(instructions) + 2
        ws1.cell(row=technical_start, column=1, value="Technical: Run Information").font = Font(bold=True, size=12)
        ws1.cell(row=technical_start, column=1).fill = header_fill
        ws1.cell(row=technical_start, column=2).fill = header_fill

        tech_info = [
            ("Run Timestamp", run_timestamp.strftime("%Y-%m-%d %I:%M:%S %p")),
            ("Goal Range Setting", f"{min_goal}-{max_goal} addresses"),
            ("Total Records Loaded", len(joined_gdf)),
            ("Valid Addresses Assigned", len(valid_gdf)),
            ("Excluded Address Count", len(excluded_gdf)),
            ("Records Flagged with Warnings", flagged_record_count)
        ]

        for offset, (label, value) in enumerate(tech_info, start=1):
            ws1.cell(row=technical_start + offset, column=1, value=label).font = Font(bold=True)
            ws1.cell(row=technical_start + offset, column=2, value=value)

        def add_excel_table(worksheet, dataframe, table_name):
            if dataframe.empty: return
            max_row, max_col = dataframe.shape
            table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row + 1}"
            tab = Table(displayName=table_name, ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            worksheet.add_table(tab)

        # --- TAB 2: COUNTS ---
        counts_df_sorted = counts_df.sort_values(by="Territory_Name").rename(columns={"Territory_Name": "Territory Name", "Total_Addresses": "# of Addresses"})
        counts_df_sorted.to_excel(writer, sheet_name="Counts", index=False)
        ws2 = writer.sheets["Counts"]
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 15
        add_excel_table(ws2, counts_df_sorted, "CountsTable")

        for row_number in range(2, len(counts_df_sorted) + 2):
            ws2[f"B{row_number}"].alignment = Alignment(horizontal="center")
            category_cell = ws2[f"C{row_number}"]
            if category_cell.value == "Ideal": category_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif category_cell.value == "Undersized": category_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif category_cell.value == "Oversized": category_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            category_cell.alignment = Alignment(horizontal="center")

        # --- TAB 3: ADDRESS LIST ---
        valid_gdf["Latitude"] = valid_gdf.geometry.y
        valid_gdf["Longitude"] = valid_gdf.geometry.x
        valid_gdf[["HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort"]] = valid_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
        valid_gdf["Unit_Sort"] = valid_gdf["Canonical_Unit"].map(clean_field).str.upper()

        address_list_df = valid_gdf.sort_values(by=["Territory_Name", "Canonical_Street", "HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort", "Unit_Sort"], kind="stable")
        
        export_df = address_list_df[[
            "Source_Record_ID", "Territory_Name", "NWS_Category", "NWS_Number", "Mailable_Address", "Data_Quality_Flag",
            "Canonical_HouseNo", "Canonical_HouseSx", "Canonical_Street", "Canonical_Unit", "Canonical_Muni", "Canonical_Zip_Code", "Latitude", "Longitude"
        ]].rename(columns={
            "Source_Record_ID": "Source Record ID", "Territory_Name": "Territory Name", "Mailable_Address": "Mailable Address", "Data_Quality_Flag": "Data Quality Flag",
            "Canonical_HouseNo": "HouseNo", "Canonical_HouseSx": "HouseSx", "Canonical_Street": "Street", "Canonical_Unit": "Unit", "Canonical_Muni": "Muni", "Canonical_Zip_Code": "Zip_Code"
        })

        export_df.to_excel(writer, sheet_name="Address List", index=False)
        ws3 = writer.sheets["Address List"]
        add_excel_table(ws3, export_df, "AddressListTable")

        for column_letter in ["G", "H", "I", "J", "K", "L", "M", "N"]: ws3.column_dimensions[column_letter].hidden = True
        ws3.column_dimensions["A"].width = 22
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 15
        ws3.column_dimensions["D"].width = 15
        ws3.column_dimensions["E"].width = 55
        ws3.column_dimensions["F"].width = 35

        # --- TAB 4: APARTMENTS ---
        apartment_source = valid_gdf[["Territory_Name", "Base_Address", "Canonical_Unit"]].copy()
        apartment_source["_Unit_Normalized"] = apartment_source["Canonical_Unit"].map(clean_field).str.upper().str.replace(r"\s+", " ", regex=True).str.strip()
        apartment_source = apartment_source[apartment_source["_Unit_Normalized"].ne("") & apartment_source["Base_Address"].map(clean_field).ne("")].copy()

        apt_groups = apartment_source.groupby(["Territory_Name", "Base_Address"], observed=True)["_Unit_Normalized"].nunique().reset_index(name="Total Units")
        apt_groups = apt_groups[apt_groups["Total Units"] >= 5].copy()

        if not counts_df.empty:
            category_mapping = counts_df.set_index("Territory_Name")["Category"].to_dict()
            apt_groups["Status"] = apt_groups["Territory_Name"].map(category_mapping)
        else: apt_groups["Status"] = "Unknown"

        def format_terr_name(row): return f"{row['Territory_Name']} [{row['Status']}]"

        if not apt_groups.empty:
            apt_groups["Territory Name"] = apt_groups.apply(format_terr_name, axis=1)
            apt_groups.rename(columns={"Base_Address": "Base Address"}, inplace=True)
            apt_export = apt_groups[["Territory Name", "Base Address", "Total Units"]]
        else:
            apt_export = pd.DataFrame(columns=["Territory Name", "Base Address", "Total Units"])

        apt_export.to_excel(writer, sheet_name="Apartments", index=False)
        ws4 = writer.sheets["Apartments"]
        add_excel_table(ws4, apt_export, "ApartmentsTable")
        ws4.column_dimensions["A"].width = 30
        ws4.column_dimensions["B"].width = 40
        ws4.column_dimensions["C"].width = 15
        for row_number in range(2, len(apt_export) + 2): ws4[f"C{row_number}"].alignment = Alignment(horizontal="center")

        # --- TAB 5: BORDER REWRITES ---
        oversized = counts_df.loc[counts_df["Category"].eq("Oversized"), "Territory_Name"].tolist() if not counts_df.empty else []
        undersized = counts_df.loc[counts_df["Category"].eq("Undersized"), "Territory_Name"].tolist() if not counts_df.empty else []

        terr_geoms = kml_gdf[["Territory_Name", "geometry_terr"]].dropna(subset=["Territory_Name", "geometry_terr"]).set_geometry("geometry_terr").dissolve(by="Territory_Name")
        terr_geoms["geometry_terr"] = terr_geoms.geometry.make_valid()
        terr_geoms = terr_geoms[terr_geoms.geometry.notna() & ~terr_geoms.geometry.is_empty].copy()
        
        terr_geoms_metric = terr_geoms.to_crs(metric_crs)
        territory_sindex = terr_geoms_metric.sindex

        count_lookup = counts_df.drop_duplicates("Territory_Name").set_index("Territory_Name")["Total_Addresses"].to_dict()
        undersized_set = set(undersized)
        seen_pairs = set()
        suggestions = []

        for over_name in oversized:
            if over_name not in terr_geoms_metric.index: continue
            over_geom = terr_geoms_metric.at[over_name, "geometry_terr"]
            over_count = count_lookup.get(over_name, 0)
            proximity_zone = over_geom.buffer(15.0)

            candidate_positions = territory_sindex.query(proximity_zone, predicate="intersects")
            for candidate_position in candidate_positions:
                under_name = terr_geoms_metric.index[candidate_position]
                if under_name == over_name or under_name not in undersized_set: continue
                
                pair_key = tuple(sorted((str(over_name), str(under_name))))
                if pair_key in seen_pairs: continue

                under_geom = terr_geoms_metric.iloc[candidate_position].geometry_terr
                if over_geom.distance(under_geom) > 15.0: continue

                seen_pairs.add(pair_key)
                under_count = count_lookup.get(under_name, 0)
                suggestions.append([over_name, over_count, under_name, under_count, ""])

        suggestion_df = pd.DataFrame(suggestions, columns=["Too Large", "Count", "Too Small", "Count ", "Recommendation"])
        suggestion_df.to_excel(writer, sheet_name="Border Rewrites", index=False)
        ws5 = writer.sheets["Border Rewrites"]
        add_excel_table(ws5, suggestion_df, "BorderRewritesTable")

        ws5.column_dimensions["A"].width = 18
        ws5.column_dimensions["C"].width = 18
        ws5.column_dimensions["E"].width = 85

        for row_number in range(2, len(suggestions) + 2):
            suggestion = suggestions[row_number - 2]
            difference = abs(suggestion[1] - suggestion[3])
            ws5.cell(row=row_number, column=5).value = CellRichText([
                "That is a ", TextBlock(bold_inline, f"{difference} address difference"),
                f". Shrink {suggestion[0]} & Expand {suggestion[2]}."
            ])

        # --- TAB 6: EXCLUDED AUDIT ---
        if not excluded_gdf.empty:
            excluded_gdf[["HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort"]] = excluded_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
            excluded_gdf["Unit_Sort"] = excluded_gdf["Canonical_Unit"].map(clean_field).str.upper()
            excluded_list_df = excluded_gdf.sort_values(by=["Territory_Name", "Canonical_Street", "HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort", "Unit_Sort"], kind="stable")
            
            export_ex_df = excluded_list_df[[
                "Source_Record_ID", "Territory_Name", "NWS_Category", "NWS_Number", "Mailable_Address", "Canonical_Status", "Data_Quality_Flag",
                "Canonical_HouseNo", "Canonical_Street", "Canonical_Unit", "Canonical_Zip_Code"
            ]].rename(columns={
                "Source_Record_ID": "Source Record ID", "Territory_Name": "Territory Name", "Mailable_Address": "Mailable Address", 
                "Canonical_Status": "Exclusion Reason", "Data_Quality_Flag": "Data Quality Flag",
                "Canonical_HouseNo": "HouseNo", "Canonical_Street": "Street", "Canonical_Unit": "Unit", "Canonical_Zip_Code": "Zip_Code"
            })
            export_ex_df.to_excel(writer, sheet_name="Excluded Audit", index=False)
            ws6 = writer.sheets["Excluded Audit"]
            add_excel_table(ws6, export_ex_df, "ExcludedAuditTable")

            for column_letter in ["H", "I", "J", "K"]: ws6.column_dimensions[column_letter].hidden = True
            ws6.column_dimensions["A"].width = 22
            ws6.column_dimensions["B"].width = 18
            ws6.column_dimensions["C"].width = 15
            ws6.column_dimensions["D"].width = 15
            ws6.column_dimensions["E"].width = 55
            ws6.column_dimensions["F"].width = 25
            ws6.column_dimensions["G"].width = 35
        else:
            empty_excluded_columns = ["Source Record ID", "Territory Name", "NWS_Category", "NWS_Number", "Mailable Address", "Exclusion Reason", "Data Quality Flag", "HouseNo", "Street", "Unit", "Zip_Code"]
            pd.DataFrame(columns=empty_excluded_columns).to_excel(writer, sheet_name="Excluded Audit", index=False)
            writer.sheets["Excluded Audit"].cell(row=2, column=1, value="No addresses were excluded in this map area.")

        # --- EXCEL UX POLISH ---
        for tab_name in ["Counts", "Address List", "Apartments", "Border Rewrites", "Excluded Audit"]:
            ws = writer.sheets[tab_name]
            ws.freeze_panes = "A2"

        writer.sheets["Dashboard"].sheet_properties.tabColor = "1E90FF"
        writer.sheets["Counts"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Address List"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Apartments"].sheet_properties.tabColor = "FF8C00"
        writer.sheets["Border Rewrites"].sheet_properties.tabColor = "FF0000"
        writer.sheets["Excluded Audit"].sheet_properties.tabColor = "808080"

    output.seek(0)
    return output

# --- 4. EXECUTION FLOW ---
if "last_uploaded_kml" not in st.session_state:
    st.session_state["last_uploaded_kml"] = None

if uploaded_kml != st.session_state["last_uploaded_kml"]:
    if "excel_data" in st.session_state:
        del st.session_state["excel_data"]
    st.session_state["last_uploaded_kml"] = uploaded_kml

if uploaded_kml:
    if st.button("Generate Territory Analysis"):
        with st.spinner(f"Loading Master {selected_county} County Data..."):
            parcel_gdf = load_county_data(selected_county)

        if parcel_gdf is not None:
            county_config = COUNTY_CONFIGS[selected_county]
            parcel_gdf = parcel_gdf.reset_index(drop=True).copy()
            parcel_gdf = parcel_gdf.rename(columns=county_config["column_mapping"], errors="ignore")

            missing_columns = [
                column for column in REQUIRED_CANONICAL_COLUMNS if column not in parcel_gdf.columns
            ]
            if missing_columns:
                st.error("County data failed preflight validation. Missing required canonical columns: " + ", ".join(missing_columns))
                st.stop()

            # Ensure we default to Canonical_Native_Source_ID, or fall back immediately if it's completely missing
            native_id_col = "Canonical_Native_Source_ID"
            county_id_prefix = re.sub(r"[^A-Za-z0-9]+", "_", selected_county.upper()).strip("_")
            fallback_ids = pd.Series([f"{county_id_prefix}-FALLBACK-{row_number:09d}" for row_number in range(1, len(parcel_gdf) + 1)], index=parcel_gdf.index, dtype="string")
            
            if native_id_col in parcel_gdf.columns:
                native_ids = parcel_gdf[native_id_col].map(clean_field)
                parcel_gdf["Source_Record_ID"] = native_ids.where(native_ids.ne(""), fallback_ids)
            else:
                parcel_gdf["Source_Record_ID"] = fallback_ids

            duplicate_native_ids = parcel_gdf["Source_Record_ID"].duplicated(keep=False)
            if duplicate_native_ids.any():
                duplicate_sequence = (parcel_gdf.groupby("Source_Record_ID").cumcount() + 1).astype(str)
                parcel_gdf.loc[duplicate_native_ids, "Source_Record_ID"] = parcel_gdf.loc[duplicate_native_ids, "Source_Record_ID"] + "-DUP-" + duplicate_sequence.loc[duplicate_native_ids]

            with st.spinner("Parsing KML Territories & Executing Spatial Join..."):
                try:
                    kml_gdf = gpd.read_file(uploaded_kml, driver="KML")
                    if kml_gdf.crs is None: kml_gdf = kml_gdf.set_crs("EPSG:4326", allow_override=True)
                    if parcel_gdf.crs is None: raise ValueError("The parcel dataset has no CRS. Assign the correct source CRS before processing.")

                    kml_gdf = kml_gdf.copy()
                    kml_gdf["geometry"] = kml_gdf.geometry.make_valid()
                    kml_gdf = kml_gdf[kml_gdf.geometry.notna() & ~kml_gdf.geometry.is_empty].copy()

                    fallback_names = "Territory_" + kml_gdf.index.to_series().astype(str)
                    if "Name" in kml_gdf.columns: kml_gdf["Territory_Name"] = kml_gdf["Name"].replace(r"^\s*$", pd.NA, regex=True).fillna(fallback_names)
                    elif "Description" in kml_gdf.columns: kml_gdf["Territory_Name"] = kml_gdf["Description"].replace(r"^\s*$", pd.NA, regex=True).fillna(fallback_names)
                    else: kml_gdf["Territory_Name"] = fallback_names

                    if parcel_gdf.crs != kml_gdf.crs: parcel_gdf = parcel_gdf.to_crs(kml_gdf.crs)

                    parcel_gdf = parcel_gdf.copy()
                    parcel_gdf["geometry"] = parcel_gdf.geometry.make_valid()
                    parcel_gdf = parcel_gdf[parcel_gdf.geometry.notna() & ~parcel_gdf.geometry.is_empty].copy()

                    territory_envelope = kml_gdf.geometry.union_all().envelope
                    parcel_gdf = parcel_gdf[parcel_gdf.
