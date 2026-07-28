import streamlit as st
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.table import Table, TableStyleInfo
import fiona
import io
import datetime
import re
import random
from fractions import Fraction

# Enable KML support in GeoPandas
fiona.drvsupport.supported_drivers['KML'] = 'rw'
fiona.drvsupport.supported_drivers['LIBKML'] = 'rw'

COUNTY_CONFIGS = {
    "Milwaukee": {
        "file_path": "zip://data/Milwaukee_Datapoints07072026.zip",
        "state": "WI",
        "metric_crs": "EPSG:3071",
        "native_source_id": "TAXKEY",
        "excluded_statuses": [
            "Undeveloped", "Parking Lot", "ROW", "Park or Recreational Facility",
            "Undeveloped Outlot", "Sliver or Remnant", "Non Addressable Assoc with Adj Parcel",
        ],
        "column_mapping": {
            "TAXKEY": "Canonical_Native_Source_ID",
            "HouseNo": "Canonical_HouseNo",
            "HouseSx": "Canonical_HouseSx",
            "Dir": "Canonical_Dir",
            "Street": "Canonical_Street",
            "StType": "Canonical_StType",
            "Muni": "Canonical_Muni",
            "Zip_Code": "Canonical_Zip_Code",
            "Unit": "Canonical_Unit",
            "Addr_Statu": "Canonical_Status",
        },
    }
}

REQUIRED_CANONICAL_COLUMNS = [
    "Canonical_HouseNo", "Canonical_HouseSx", "Canonical_Dir", "Canonical_Street",
    "Canonical_StType", "Canonical_Muni", "Canonical_Zip_Code", "Canonical_Unit",
    "Canonical_Status", "geometry",
]

# --- 1. CONFIGURATION & UI SETUP ---
st.set_page_config(page_title="Territory Audit Engine", layout="wide")

st.title("Congregation Territory Analysis Engine")
st.markdown("Upload your territories KML map to generate a complete, filtered address database & analysis.")

st.sidebar.header("Step 1: Configuration")
congregation_name = st.sidebar.text_input("Congregation Name (No Spaces)", "ExampleCongregation")
selected_county = st.sidebar.selectbox("Select County Data", list(COUNTY_CONFIGS.keys()))
goal_range = st.sidebar.selectbox("Goal # of Addresses Per Territory", ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"])
apartment_threshold = st.sidebar.selectbox("Apartment Grouping Threshold", [4, 5, 6], index=1)

st.header("Step 2: Upload Territory Map")
uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

MIN_GOAL, MAX_GOAL = [int(x) for x in goal_range.split("-")]

# --- 2. DATA LOADING & CACHING ---
@st.cache_data
def load_county_data(county_name):
    county_config = COUNTY_CONFIGS[county_name]
    try:
        return gpd.read_file(county_config["file_path"])
    except Exception as error:
        st.error(f"Error loading county shapefile. Check the configured file path. Error: {error}")
        return None

def natural_keys(text):
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(text))]

# --- ADDRESS BUILDER + NORMALIZATION HELPERS ---
def clean_field(value):
    if pd.isna(value): return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}: return ""
    return text

def normalize_house_number(value):
    text = clean_field(value)
    if not text: return ""
    if re.fullmatch(r"[+-]?\d+\.0+", text): return text.split(".", 1)[0]
    return re.sub(r"\s+", " ", text)

def normalize_zip_code(value):
    text = clean_field(value)
    if not text: return ""
    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D", "", text)
    if not digits: return ""
    if len(digits) == 4: digits = digits.zfill(5)
    elif len(digits) == 8: digits = digits.zfill(9)
    if len(digits) == 5: return digits
    if len(digits) == 9: return f"{digits[:5]}-{digits[5:]}"
    if len(digits) > 9: return f"{digits[:5]}-{digits[5:9]}"
    return digits

def normalize_unit(value):
    text = clean_field(value)
    if not text: return ""
    if re.fullmatch(r"\d+\.0+", text): text = text.split(".", 1)[0]
    text = re.sub(r"\s+", " ", text).strip()
    descriptive_pattern = re.compile(r"^(?:apt(?:artment)?|unit|ste|suite|upper|lower|bsmt|basement|rear|front|floor|fl|building|bldg|room|rm)\b", flags=re.IGNORECASE)
    if descriptive_pattern.search(text): return text
    return f"Apt {text}"

def house_number_sort_parts(value):
    text = normalize_house_number(value).upper()
    if not text: return pd.Series([float("inf"), 9, ""])
    compact = re.sub(r"\s+", " ", text).strip()
    mixed_fraction = re.match(r"^(\d+)\s+(\d+)\s*/\s*(\d+)(.*)$", compact)
    if mixed_fraction:
        whole, numerator, denominator, suffix = mixed_fraction.groups()
        try: numeric_value = int(whole) + float(Fraction(int(numerator), int(denominator)))
        except (ValueError, ZeroDivisionError): numeric_value = float(int(whole))
        return pd.Series([numeric_value, 1, suffix.strip()])
    simple_fraction = re.match(r"^(\d+)\s*/\s*(\d+)(.*)$", compact)
    if simple_fraction:
        numerator, denominator, suffix = simple_fraction.groups()
        try: numeric_value = float(Fraction(int(numerator), int(denominator)))
        except (ValueError, ZeroDivisionError): numeric_value = float("inf")
        return pd.Series([numeric_value, 1, suffix.strip()])
    numeric_prefix = re.match(r"^(\d+(?:\.\d+)?)(.*)$", compact)
    if numeric_prefix:
        number, suffix = numeric_prefix.groups()
        return pd.Series([float(number), 0 if not suffix.strip() else 2, suffix.strip()])
    return pd.Series([float("inf"), 8, compact])

def build_addresses(row, state):
    house = normalize_house_number(row.get("Canonical_HouseNo"))
    house_sx = clean_field(row.get("Canonical_HouseSx"))
    direction = clean_field(row.get("Canonical_Dir"))
    street = clean_field(row.get("Canonical_Street"))
    st_type = clean_field(row.get("Canonical_StType"))
    muni = clean_field(row.get("Canonical_Muni"))
    normalized_zip = normalize_zip_code(row.get("Canonical_Zip_Code"))
    zip_c = normalized_zip[:5] if normalized_zip else ""
    unit_str = normalize_unit(row.get("Canonical_Unit"))

    full_house_num = f"{house}{house_sx}".strip()
    full_street = " ".join(part for part in [direction, street, st_type] if part)
    base_addr_line = " ".join(part for part in [full_house_num, full_street] if part)
    locality = ", ".join(part for part in [muni, state] if part)
    if zip_c: locality = f"{locality} {zip_c}".strip()

    base_addr = ", ".join(part for part in [base_addr_line, locality] if part)
    mailable_addr = ", ".join(part for part in [" ".join([base_addr_line, unit_str]).strip(), locality] if part)

    return pd.Series([base_addr, mailable_addr], index=["Base_Address", "Mailable_Address"])

def evaluate_data_quality(row):
    issues = []
    house = normalize_house_number(row.get("Canonical_HouseNo"))
    street = clean_field(row.get("Canonical_Street"))
    municipality = clean_field(row.get("Canonical_Muni"))
    zip_code = normalize_zip_code(row.get("Canonical_Zip_Code"))
    base_address = clean_field(row.get("Base_Address"))
    mailable_address = clean_field(row.get("Mailable_Address"))

    if not street:
        issues.append("Missing Street")
    if not municipality:
        issues.append("Missing Municipality")
    if not zip_code:
        issues.append("Missing ZIP")
    if house and re.fullmatch(r"[+-]?0+(?:\.0+)?", house):
        issues.append("Zero House Number")

    zip_is_valid = not zip_code or bool(
        re.fullmatch(r"\d{5}(?:-\d{4})?", zip_code)
    )
    if (
        not house
        or not base_address
        or not mailable_address
        or ",," in base_address
        or ",," in mailable_address
        or not zip_is_valid
    ):
        issues.append("Malformed Address")

    return " | ".join(issues)


def parse_house_number_components(full_house_number):
    text = normalize_house_number(full_house_number).upper()
    if not text:
        return "", "", ""

    text = re.sub(r"\s+", " ", text).strip()
    directional_prefix = ""
    house_number_main = ""
    house_suffix = ""

    directional_match = re.fullmatch(
        r"([NSEW])\s*(\d+(?:\s+\d+/\d+|\.\d+)?)([A-Z]?)",
        text,
    )
    if directional_match:
        directional_prefix, house_number_main, house_suffix = (
            directional_match.groups()
        )
        return directional_prefix, house_number_main, house_suffix

    standard_match = re.fullmatch(
        r"(\d+(?:\s+\d+/\d+|\.\d+)?)([A-Z]?)",
        text,
    )
    if standard_match:
        house_number_main, house_suffix = standard_match.groups()
        return "", house_number_main, house_suffix

    numeric_match = re.search(r"\d+(?:\s+\d+/\d+|\.\d+)?", text)
    if numeric_match:
        house_number_main = numeric_match.group(0)
        prefix_text = text[:numeric_match.start()].strip()
        suffix_text = text[numeric_match.end():].strip()
        if prefix_text in {"N", "S", "E", "W"}:
            directional_prefix = prefix_text
        if re.fullmatch(r"[A-Z]", suffix_text):
            house_suffix = suffix_text

    return directional_prefix, house_number_main, house_suffix


def parse_mailable_address(row, state):
    mailable_address = clean_field(row.get("Mailable_Address"))
    address_parts = [part.strip() for part in mailable_address.split(",")]

    street_line = address_parts[0] if address_parts else ""
    municipality = (
        address_parts[1]
        if len(address_parts) > 1
        else clean_field(row.get("Canonical_Muni"))
    )
    state_zip = address_parts[2] if len(address_parts) > 2 else ""

    state_value = state
    normalized_zip = normalize_zip_code(row.get("Canonical_Zip_Code"))
    zip_code = normalized_zip[:5] if normalized_zip else ""
    zip4_code = (
        normalized_zip.split("-", 1)[1]
        if "-" in normalized_zip
        else ""
    )
    state_zip_match = re.fullmatch(
        r"([A-Za-z]{2})(?:\s+(\d{5})(?:-(\d{4}))?)?",
        state_zip,
    )
    if state_zip_match:
        state_value = state_zip_match.group(1).upper()
        if state_zip_match.group(2):
            zip_code = state_zip_match.group(2)
        if state_zip_match.group(3):
            zip4_code = state_zip_match.group(3)

    unit_type = ""
    unit_value = clean_field(row.get("Canonical_Unit"))
    normalized_unit = normalize_unit(unit_value)
    if normalized_unit:
        unit_match = re.match(
            r"^(APT(?:ARTMENT)?|UNIT|STE|SUITE|UPPER|LOWER|BSMT|BASEMENT|"
            r"REAR|FRONT|FLOOR|FL|BUILDING|BLDG|ROOM|RM)\b\s*(.*)$",
            normalized_unit,
            flags=re.IGNORECASE,
        )
        if unit_match:
            unit_type = unit_match.group(1).upper()
            unit_value = unit_match.group(2).strip()
        else:
            unit_value = normalized_unit

    street_without_unit = street_line
    if normalized_unit:
        unit_suffix_pattern = re.compile(
            rf"\s+{re.escape(normalized_unit)}$",
            flags=re.IGNORECASE,
        )
        street_without_unit = unit_suffix_pattern.sub("", street_line).strip()

    full_house_number = (
        f"{normalize_house_number(row.get('Canonical_HouseNo'))}"
        f"{clean_field(row.get('Canonical_HouseSx'))}"
    ).strip()
    if not full_house_number:
        house_match = re.match(r"^(\S+)\s+(.*)$", street_without_unit)
        if house_match:
            full_house_number = house_match.group(1)

    house_prefix, house_main, house_suffix = parse_house_number_components(
        full_house_number
    )

    street_prefix = clean_field(row.get("Canonical_Dir")).upper()
    street_name = clean_field(row.get("Canonical_Street"))
    street_type = clean_field(row.get("Canonical_StType")).upper()
    full_street = " ".join(
        part for part in [street_prefix, street_name, street_type] if part
    )

    if not full_street and street_without_unit:
        remaining_street = street_without_unit
        if full_house_number and remaining_street.upper().startswith(
            full_house_number.upper()
        ):
            remaining_street = remaining_street[len(full_house_number):].strip()

        street_match = re.fullmatch(
            r"(?:(N|S|E|W|NE|NW|SE|SW)\s+)?(.+?)"
            r"(?:\s+(ST|STREET|AVE|AVENUE|RD|ROAD|BLVD|BOULEVARD|DR|DRIVE|"
            r"LN|LANE|CT|COURT|PL|PLACE|PKWY|PARKWAY|HWY|HIGHWAY|WAY|TER|"
            r"TERRACE|CIR|CIRCLE))?",
            remaining_street,
            flags=re.IGNORECASE,
        )
        if street_match:
            street_prefix = clean_field(street_match.group(1)).upper()
            street_name = clean_field(street_match.group(2))
            street_type = clean_field(street_match.group(3)).upper()
            full_street = " ".join(
                part for part in [street_prefix, street_name, street_type] if part
            )

    return pd.Series(
        {
            "FullHouNumber": full_house_number,
            "FullStreet": full_street,
            "Municipality": municipality,
            "State": state_value,
            "ZipCode": zip_code,
            "ZIP4Code": zip4_code,
            "HouseNoPrefix": house_prefix,
            "HouseNoMain": house_main,
            "HouseSx": house_suffix,
            "StreetPrefixDir": street_prefix,
            "StreetName": street_name,
            "StreetType": street_type,
            "UnitType": unit_type,
            "Unit": unit_value,
        }
    )

# --- 3. EXCEL GENERATION ENGINE ---
def generate_excel_report(joined_gdf, kml_gdf, min_goal, max_goal, cong_name, county_config, apt_threshold, kml_filename, county_filename, duplicate_assignment_count, unassigned_address_count):
    output = io.BytesIO()
    run_timestamp = datetime.datetime.now()
    state = county_config["state"]
    metric_crs = county_config["metric_crs"]
    excluded_statuses = county_config["excluded_statuses"]

    joined_gdf = joined_gdf.copy()
    joined_gdf["Canonical_Zip_Code"] = joined_gdf["Canonical_Zip_Code"].map(normalize_zip_code)
    joined_gdf[["Base_Address", "Mailable_Address"]] = joined_gdf.apply(lambda row: build_addresses(row, state), axis=1)
    joined_gdf["Data_Quality_Flag"] = joined_gdf.apply(evaluate_data_quality, axis=1)
    flagged_record_count = joined_gdf["Data_Quality_Flag"].ne("").sum()

    normalized_excluded_statuses = {clean_field(status).upper() for status in excluded_statuses}
    joined_gdf["Canonical_Status_Normalized"] = joined_gdf["Canonical_Status"].map(clean_field).str.upper()
    exclusion_mask = joined_gdf["Canonical_Status_Normalized"].isin(normalized_excluded_statuses)
    excluded_gdf = joined_gdf[exclusion_mask].copy()
    valid_gdf = joined_gdf[~exclusion_mask].copy()

    unique_territories = valid_gdf["Territory_Name"].unique().tolist()
    unique_territories.sort(key=natural_keys)
    valid_gdf["Territory_Name"] = pd.Categorical(valid_gdf["Territory_Name"], categories=unique_territories, ordered=True)

    if not excluded_gdf.empty:
        excluded_unique = excluded_gdf["Territory_Name"].unique().tolist()
        excluded_unique.sort(key=natural_keys)
        excluded_gdf["Territory_Name"] = pd.Categorical(excluded_gdf["Territory_Name"], categories=excluded_unique, ordered=True)

    counts_df = valid_gdf.groupby("Territory_Name", observed=True).size().reset_index(name="Total_Addresses")
    counts_df = counts_df[counts_df["Total_Addresses"] > 0].copy()

    def get_category(count):
        if count < min_goal: return "Undersized"
        if count <= max_goal: return "Ideal"
        return "Oversized"

    counts_df["Category"] = counts_df["Total_Addresses"].apply(get_category)

    valid_gdf[["NWS_Category", "NWS_Number"]] = valid_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
    valid_gdf["NWS_Category"] = valid_gdf["NWS_Category"].fillna("UNK")
    valid_gdf["NWS_Number"] = valid_gdf["NWS_Number"].fillna("0")

    if not excluded_gdf.empty:
        excluded_gdf[["NWS_Category", "NWS_Number"]] = excluded_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
        excluded_gdf["NWS_Category"] = excluded_gdf["NWS_Category"].fillna("UNK")
        excluded_gdf["NWS_Number"] = excluded_gdf["NWS_Number"].fillna("0")

    apartment_source = valid_gdf[
        ["Territory_Name", "Base_Address", "Canonical_Unit"]
    ].copy()
    apartment_source["_Unit_Normalized"] = (
        apartment_source["Canonical_Unit"]
        .map(clean_field)
        .str.upper()
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    apartment_source = apartment_source[
        apartment_source["Base_Address"].map(clean_field).ne("")
    ].copy()
    apartment_source["_Has_Nonblank_Unit"] = apartment_source[
        "_Unit_Normalized"
    ].ne("")

    apt_groups = (
        apartment_source.groupby(
            ["Territory_Name", "Base_Address"],
            observed=True,
        )
        .agg(
            **{
                "Source Rows": ("_Unit_Normalized", "size"),
                "Nonblank Unit Rows": ("_Has_Nonblank_Unit", "sum"),
                "Unique Normalized Units": (
                    "_Unit_Normalized",
                    lambda values: values[values.ne("")].nunique(),
                ),
            }
        )
        .reset_index()
    )
    apt_groups["Blank Parent Rows"] = (
        apt_groups["Source Rows"] - apt_groups["Nonblank Unit Rows"]
    )
    apt_groups["Duplicate Units"] = (
        apt_groups["Nonblank Unit Rows"]
        - apt_groups["Unique Normalized Units"]
    ).clip(lower=0)
    apt_groups["Reported County Unit Count"] = ""
    apt_groups["Total Units"] = apt_groups["Unique Normalized Units"]

    def get_apartment_confidence(row):
        if row["Unique Normalized Units"] < apt_threshold:
            return "Below Threshold"
        if row["Duplicate Units"] == 0 and row["Blank Parent Rows"] <= 1:
            return "High"
        if row["Duplicate Units"] <= 2:
            return "Medium"
        return "Low"

    apt_groups["Apartment Confidence"] = apt_groups.apply(
        get_apartment_confidence,
        axis=1,
    )
    apt_groups["Detection Reason"] = apt_groups.apply(
        lambda row: (
            f"{int(row['Unique Normalized Units'])} unique nonblank unit "
            f"identifier(s) met the threshold of {apt_threshold}."
        ),
        axis=1,
    )
    apt_groups = apt_groups[
        apt_groups["Total Units"] >= apt_threshold
    ].copy()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        
        # --- TAB 1: DASHBOARD ---
        total_territories = len(counts_df)
        total_addresses = counts_df["Total_Addresses"].sum()
        largest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmax()] if total_territories > 0 else None
        smallest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmin()] if total_territories > 0 else None
        ideal_pct = (len(counts_df[counts_df["Category"] == "Ideal"]) / total_territories * 100) if total_territories > 0 else 0

        largest_name = largest_terr["Territory_Name"] if largest_terr is not None else ""
        largest_count = largest_terr["Total_Addresses"] if largest_terr is not None else 0
        smallest_name = smallest_terr["Territory_Name"] if smallest_terr is not None else ""
        smallest_count = smallest_terr["Total_Addresses"] if smallest_terr is not None else 0

        dashboard_top = [
            [f"TERRITORY ANALYSIS: {cong_name}".upper()],
            [f"Generated on {run_timestamp.strftime('%Y-%m-%d %H:%M')} by TerritoryToolbox (using the analysis tool)"],
            [""],
            [f"Total Territories: {total_territories}"],
            [f"Total Valid Addresses: {total_addresses}"],
            [f"Excluded Addresses (See Tab 6): {len(excluded_gdf)}"],
            [f"The largest territory has {largest_count} addresses in it ({largest_name})."],
            [f"The smallest territory has {smallest_count} addresses in it ({smallest_name})."],
            [""],
            [f"Ideal Address Range: ({min_goal}-{max_goal})"],
            [""]
        ]
        pd.DataFrame(dashboard_top).to_excel(writer, sheet_name="Dashboard", index=False, header=False)

        fixed_bins = [
            ("Under 25", counts_df["Total_Addresses"] < 25),
            ("25-49", counts_df["Total_Addresses"].between(25, 49, inclusive="both")),
            ("50-74", counts_df["Total_Addresses"].between(50, 74, inclusive="both")),
            ("75-99", counts_df["Total_Addresses"].between(75, 99, inclusive="both")),
            ("100-125", counts_df["Total_Addresses"].between(100, 125, inclusive="both")),
            ("126-150", counts_df["Total_Addresses"].between(126, 150, inclusive="both")),
            ("151-175", counts_df["Total_Addresses"].between(151, 175, inclusive="both")),
            ("Over 175", counts_df["Total_Addresses"] > 175)
        ]

        distribution = []
        for range_label, range_mask in fixed_bins:
            range_rows = counts_df.loc[range_mask]
            range_count = len(range_rows)
            range_categories = range_rows["Category"].dropna().astype(str).unique().tolist()
            
            if len(range_categories) == 1: range_category = range_categories[0]
            elif len(range_categories) > 1: range_category = "Mixed"
            else:
                if range_label == "Under 25": r_min, r_max = 0, 24
                elif range_label == "Over 175": r_min, r_max = 176, float("inf")
                else: r_min, r_max = [int(v) for v in range_label.split("-")]
                
                if r_max < min_goal: range_category = "Undersized"
                elif r_min > max_goal: range_category = "Oversized"
                elif r_min >= min_goal and r_max <= max_goal: range_category = "Ideal"
                else: range_category = "Mixed"
            distribution.append([range_category, range_label, range_count])

        pd.DataFrame(distribution, columns=["Category", "Range", "Count"]).to_excel(writer, sheet_name="Dashboard", startrow=11, index=False)

        ws1 = writer.sheets["Dashboard"]
        ws1.column_dimensions["A"].width = 18
        ws1.column_dimensions["B"].width = 18
        ws1.column_dimensions["C"].width = 18

        ws1["A1"].font = Font(size=20, bold=True, color="0D6B31")
        ws1["A2"].hyperlink = None
        ws1["A2"].font = Font(size=10, italic=True, color="0D6B31")

        bold_inline = InlineFont(b=True)
        ws1["A10"].value = CellRichText([TextBlock(bold_inline, "Ideal Address Range"), f": ({min_goal}-{max_goal})"])
        ws1["A11"].value = CellRichText(["About ", TextBlock(bold_inline, f"{ideal_pct:.1f}%"), " of territories fall within this range."])

        header_fill = PatternFill(start_color="C7CDDB", end_color="C7CDDB", fill_type="solid")
        for col in range(1, 4):
            ws1.cell(row=12, column=col).fill = header_fill
            ws1.cell(row=12, column=col).font = Font(bold=True)

        distribution_end_row = 12 + len(distribution)
        for row_number in range(13, distribution_end_row + 1):
            if ws1.cell(row=row_number, column=1).value == "Ideal":
                for col in range(1, 4): ws1.cell(row=row_number, column=col).font = Font(bold=True)

        instruction_start = distribution_end_row + 2
        volunteer_instructions = [
            "The addresses in this analysis, with a little reformatting, can be added to NWS or other supported programs (Visit https://territorytoolbox.com for details)",
            "It's suggested to export this file into a program you can easily edit, like excel or google sheets.",
            "That will allow you to expand cells to read easier, create custom filters to see specific data, and customize the sheet to make it more legible.",
        ]

        for offset, instruction in enumerate(volunteer_instructions):
            row_number = instruction_start + offset
            cell = ws1.cell(row=row_number, column=1, value=instruction)
            if "https://territorytoolbox.com" in instruction:
                cell.hyperlink = "https://territorytoolbox.com"
                cell.font = Font(color="0563C1", underline="single")

        features_start = instruction_start + len(volunteer_instructions) + 2
        ws1.cell(row=features_start, column=1, value="Features Of This Spreadsheet").font = Font(bold=True, size=12)
        for col in range(1, 4): ws1.cell(row=features_start, column=col).fill = header_fill

        feature_instructions = [
            CellRichText(["The ", TextBlock(bold_inline, "DASHBOARD"), " tab displays basic statistics about the territory that was analyzed"]),
            CellRichText(["The ", TextBlock(bold_inline, "COUNTS"), " tab organizes territories by size. This is done by 'counting' workable addresses, not geographical size."]),
            CellRichText(["The ", TextBlock(bold_inline, "ADDRESS LIST"), " tab displays every workable address in your territory."]),
            CellRichText(["The ", TextBlock(bold_inline, "APARTMENTS"), f" tab displays every multifamily at or above {apt_threshold} units in your territory. Large units can be explanations for inflated door-to-door territories."]),
            CellRichText(["The ", TextBlock(bold_inline, "TERRITORY BALANCING"), " tab provides reduction, consolidation, and border-shift actions for balancing territories."]),
            CellRichText(["The ", TextBlock(bold_inline, "EXCLUDED AUDIT"), " tab displays addresses that are NOT counted towards your territory. These are usually addresses of highways, vacant lots, parks, etc. This is included for confidence."]),
        ]

        for offset, instruction in enumerate(feature_instructions, start=1):
            ws1.cell(row=features_start + offset, column=1, value=instruction)

        technical_start = features_start + len(feature_instructions) + 2
        ws1.cell(row=technical_start, column=1, value="Technical: Run Information").font = Font(bold=True, size=12)
        for col in range(1, 4): ws1.cell(row=technical_start, column=col).fill = header_fill

        tech_info = [
            ("Run Timestamp", run_timestamp.strftime("%Y-%m-%d %H:%M")),
            ("Ideal Address Range Setting", f"{min_goal}-{max_goal} addresses"),
            ("Apartment Grouping Threshold", f"{apt_threshold} units"),
            ("Address Records Loaded", len(joined_gdf)),
            ("Valid Addresses Assigned", len(valid_gdf)),
            ("Excluded Address Count", len(excluded_gdf)),
            ("Records Flagged with Warnings", flagged_record_count),
            ("KML Filename", kml_filename),
            ("County Source Filename", county_filename),
            ("Duplicate Boundary Assignments", duplicate_assignment_count),
            ("Unassigned Address Records", unassigned_address_count),
        ]

        for offset, (label, value) in enumerate(tech_info, start=1):
            row_number = technical_start + offset
            ws1.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=2,
            )
            label_cell = ws1.cell(row=row_number, column=1, value=label)
            value_cell = ws1.cell(row=row_number, column=3, value=value)

            label_cell.font = Font(bold=False)
            value_cell.font = Font(bold=False)
            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
            value_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=False,
            )

        for row in ws1.iter_rows(
            min_col=2,
            max_col=3,
            min_row=1,
            max_row=ws1.max_row,
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical=cell.alignment.vertical or "top",
                    wrap_text=cell.alignment.wrap_text,
                )

        ws1.delete_cols(17, 10)

        def add_excel_table(worksheet, dataframe, table_name, show_stripes=False):
            if dataframe.empty: return
            max_row, max_col = dataframe.shape
            table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row + 1}"
            tab = Table(displayName=table_name, ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=show_stripes, showColumnStripes=False)
            worksheet.add_table(tab)

        # --- TAB 2: COUNTS ---
        apartment_units_by_territory = (
            apt_groups.groupby("Territory_Name", observed=True)["Total Units"]
            .sum()
            .reset_index(name="# of Apartment Units")
        )
        counts_df_sorted = counts_df.merge(
            apartment_units_by_territory,
            on="Territory_Name",
            how="left",
        )
        counts_df_sorted["# of Apartment Units"] = (
            counts_df_sorted["# of Apartment Units"]
            .fillna(0)
            .astype(int)
        )
        counts_df_sorted["Addresses With Apartments Removed"] = (
            counts_df_sorted["Total_Addresses"]
            - counts_df_sorted["# of Apartment Units"]
        ).clip(lower=0)
        counts_df_sorted["Potential Status"] = counts_df_sorted[
            "Addresses With Apartments Removed"
        ].apply(get_category)

        suggested_actions = {
            ("Oversized", "Ideal"): (
                "Consider turning apartment units into letter writing"
            ),
            ("Ideal", "Undersized"): (
                "This is an apartment heavy territory. If you intend on "
                "transforming apartments into letter writing, consider a "
                "border adjustment to add more door-to-door territory."
            ),
            ("Oversized", "Undersized"): (
                "Consider turning some (but not all) apartment units into "
                "letter writing, or consider a border adjustment"
            ),
            ("Ideal", "Ideal"): "No action needed",
            ("Undersized", "Undersized"): (
                "Consider a border adjustment to add more door-to-door territory"
            ),
            ("Oversized", "Oversized"): (
                "Consider a border adjustment to subtract the amount of "
                "door-to-door needed to cover the territory"
            ),
            ("Undersized", "Ideal"): "This is impossible.",
            ("Ideal", "Oversized"): "This is impossible.",
            ("Undersized", "Oversized"): "This is impossible.",
        }
        counts_df_sorted["Suggested Action"] = counts_df_sorted.apply(
            lambda row: suggested_actions.get(
                (row["Category"], row["Potential Status"]),
                "Review territory manually",
            ),
            axis=1,
        )

        counts_df_sorted = (
            counts_df_sorted.sort_values(by="Territory_Name")
            .rename(
                columns={
                    "Territory_Name": "Territory Name",
                    "Total_Addresses": "Total Address Count",
                    "Category": "Current Status",
                    "Addresses With Apartments Removed": (
                        "Total Address Count Without Apartments"
                    ),
                }
            )[
                [
                    "Territory Name",
                    "Total Address Count",
                    "Current Status",
                    "# of Apartment Units",
                    "Total Address Count Without Apartments",
                    "Potential Status",
                    "Suggested Action",
                ]
            ]
        )

        counts_df_sorted.to_excel(
            writer,
            sheet_name="Counts",
            index=False,
        )
        ws2 = writer.sheets["Counts"]
        ws2.freeze_panes = "B2"

        counts_widths = {
            "A": 14,
            "B": 14,
            "C": 18,
            "D": 14,
            "E": 18,
            "F": 18,
            "G": 92,
        }
        for column_letter, width in counts_widths.items():
            ws2.column_dimensions[column_letter].width = width

        header_fill_counts = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        green_status_fill = PatternFill(
            start_color="C4EFD0",
            end_color="C4EFD0",
            fill_type="solid",
        )
        green_data_fill = PatternFill(
            start_color="E1F2DB",
            end_color="E1F2DB",
            fill_type="solid",
        )
        red_current_status_fill = PatternFill(
            start_color="EA9D9C",
            end_color="EA9D9C",
            fill_type="solid",
        )
        red_potential_status_fill = PatternFill(
            start_color="DFC2D0",
            end_color="DFC2D0",
            fill_type="solid",
        )
        red_data_fill = PatternFill(
            start_color="E3D5DC",
            end_color="E3D5DC",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        bottom_border = Border(
            bottom=Side(style="thin", color="666666")
        )

        for cell in ws2[1]:
            cell.fill = header_fill_counts
            cell.font = Font(bold=True, color="EAECEB")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = bottom_border

        for row_number in range(2, len(counts_df_sorted) + 2):
            current_status = ws2.cell(row=row_number, column=3).value
            potential_status = ws2.cell(row=row_number, column=6).value
            suggested_action = ws2.cell(row=row_number, column=7).value

            alternate_fill = (
                white_fill if row_number % 2 == 0 else stripe_fill
            )
            ws2.cell(row=row_number, column=1).fill = alternate_fill

            current_is_ideal = current_status == "Ideal"
            potential_is_ideal = potential_status == "Ideal"

            current_status_fill = (
                green_status_fill
                if current_is_ideal
                else red_current_status_fill
            )
            current_data_fill = (
                green_data_fill
                if current_is_ideal
                else red_data_fill
            )
            potential_status_fill = (
                green_status_fill
                if potential_is_ideal
                else red_potential_status_fill
            )

            for column_number in [2, 4, 5]:
                ws2.cell(
                    row=row_number,
                    column=column_number,
                ).fill = current_data_fill

            ws2.cell(
                row=row_number,
                column=3,
            ).fill = current_status_fill
            ws2.cell(
                row=row_number,
                column=6,
            ).fill = potential_status_fill
            suggested_action_fill = (
                green_data_fill if potential_is_ideal else red_data_fill
            )
            ws2.cell(
                row=row_number,
                column=7,
            ).fill = suggested_action_fill

            for column_number in range(1, 8):
                cell = ws2.cell(
                    row=row_number,
                    column=column_number,
                )
                cell.font = Font(
                    bold=column_number == 2,
                    italic=column_number in {4, 5, 6},
                    color="000000",
                )
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_number in {1, 7} else "center"
                    ),
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = bottom_border

        # --- TAB 3: ADDRESS LIST ---
        valid_gdf["Latitude"] = valid_gdf.geometry.y.astype(float)
        valid_gdf["Longitude"] = valid_gdf.geometry.x.astype(float)
        valid_gdf[
            [
                "HouseNum_Sort",
                "HouseNum_Suffix_Rank",
                "HouseNum_Text_Sort",
            ]
        ] = valid_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
        valid_gdf["Unit_Sort"] = (
            valid_gdf["Canonical_Unit"].map(clean_field).str.upper()
        )

        address_list_df = valid_gdf.sort_values(
            by=[
                "Territory_Name",
                "Canonical_Street",
                "HouseNum_Sort",
                "HouseNum_Suffix_Rank",
                "HouseNum_Text_Sort",
                "Unit_Sort",
            ],
            kind="stable",
        ).copy()

        parsed_address_df = address_list_df.apply(
            lambda row: parse_mailable_address(row, state),
            axis=1,
        )
        address_list_df = pd.concat(
            [address_list_df, parsed_address_df],
            axis=1,
        )

        export_df = address_list_df[
            [
                "Territory_Name",
                "Mailable_Address",
                "FullHouNumber",
                "FullStreet",
                "Municipality",
                "State",
                "ZipCode",
                "ZIP4Code",
                "HouseNoPrefix",
                "HouseNoMain",
                "HouseSx",
                "StreetPrefixDir",
                "StreetName",
                "StreetType",
                "UnitType",
                "Unit",
                "Latitude",
                "Longitude",
                "Source_Record_ID",
                "Data_Quality_Flag",
            ]
        ].rename(
            columns={
                "Territory_Name": "Territory Name",
                "Mailable_Address": "Mailable Address",
                "Source_Record_ID": "Source record ID",
                "Data_Quality_Flag": "Data Quality Flag",
            }
        )

        export_df.to_excel(
            writer,
            sheet_name="Address List",
            index=False,
        )
        ws3 = writer.sheets["Address List"]
        ws3.freeze_panes = "C2"

        address_table = Table(
            displayName="AddressListTable",
            ref=(
                f"A1:{openpyxl.utils.get_column_letter(len(export_df.columns))}"
                f"{len(export_df) + 1}"
            ),
        )
        ws3.add_table(address_table)

        hidden_address_columns = [
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
        ]
        for column_letter in hidden_address_columns:
            ws3.column_dimensions[column_letter].hidden = True

        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 57
        ws3.column_dimensions["Q"].width = 25
        ws3.column_dimensions["R"].width = 25
        ws3.column_dimensions["S"].width = 36
        ws3.column_dimensions["T"].width = 38

        header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        quality_warning_fill = PatternFill(
            start_color="EA9F9D",
            end_color="EA9F9D",
            fill_type="solid",
        )
        address_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="EAECEB")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )

        source_record_column = (
            export_df.columns.get_loc("Source record ID") + 1
        )
        latitude_column = export_df.columns.get_loc("Latitude") + 1
        longitude_column = export_df.columns.get_loc("Longitude") + 1
        quality_flag_column = (
            export_df.columns.get_loc("Data Quality Flag") + 1
        )

        for row_number in range(2, len(export_df) + 2):
            row_fill = (
                white_fill if row_number % 2 == 0 else stripe_fill
            )
            for column_number in range(1, len(export_df.columns) + 1):
                ws3.cell(
                    row=row_number,
                    column=column_number,
                ).fill = row_fill

            ws3.cell(
                row=row_number,
                column=1,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            ws3.cell(
                row=row_number,
                column=2,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            source_id_cell = ws3.cell(
                row=row_number,
                column=source_record_column,
            )
            source_id_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=False,
            )

            for coordinate_column in [latitude_column, longitude_column]:
                coordinate_cell = ws3.cell(
                    row=row_number,
                    column=coordinate_column,
                )
                if coordinate_cell.value not in {None, ""}:
                    coordinate_cell.value = float(coordinate_cell.value)
                    coordinate_cell.number_format = "0.################"

            quality_flag = ws3.cell(
                row=row_number,
                column=quality_flag_column,
            ).value
            if clean_field(quality_flag):
                for column_number in range(1, len(export_df.columns) + 1):
                    ws3.cell(
                        row=row_number,
                        column=column_number,
                    ).fill = quality_warning_fill

            for column_number in [1, 2, source_record_column]:
                ws3.cell(
                    row=row_number,
                    column=column_number,
                ).border = address_border

        # --- TAB 4: APARTMENTS ---
        if not counts_df.empty:
            category_mapping = counts_df.set_index("Territory_Name")[
                "Category"
            ].to_dict()
            address_count_mapping = counts_df.set_index("Territory_Name")[
                "Total_Addresses"
            ].to_dict()
            apt_groups["Current Territory Status"] = apt_groups[
                "Territory_Name"
            ].map(category_mapping)
            apt_groups["Total Addresses in Territory"] = apt_groups[
                "Territory_Name"
            ].map(address_count_mapping).fillna(0).astype(int)
        else:
            apt_groups["Current Territory Status"] = "Unknown"
            apt_groups["Total Addresses in Territory"] = 0

        apt_groups["_Potential Address Count"] = (
            apt_groups["Total Addresses in Territory"]
            - apt_groups["Total Units"]
        ).clip(lower=0)
        apt_groups["_Potential Status"] = apt_groups[
            "_Potential Address Count"
        ].apply(get_category)

        def get_apartment_action_code(row):
            units = int(row["Total Units"])
            current_status = row["Current Territory Status"]
            potential_status = row["_Potential Status"]
            if units >= 10:
                return "TEN_PLUS"
            if current_status == "Undersized":
                return "CURRENT_UNDERSIZED"
            if potential_status == "Undersized":
                return "POTENTIAL_UNDERSIZED"
            if current_status == "Ideal" and potential_status == "Ideal":
                return "IDEAL_TO_IDEAL"
            if current_status == "Oversized" and potential_status == "Oversized":
                return "OVERSIZED_TO_OVERSIZED"
            if current_status == "Oversized" and potential_status == "Ideal":
                return "OVERSIZED_TO_IDEAL"
            return "MANUAL"

        apartment_action_text = {
            "TEN_PLUS": (
                "Ideal for letter writing. This building has 10 or more units, "
                "which is difficult to cover if not easily accessible."
            ),
            "CURRENT_UNDERSIZED": (
                "Keep as door-to-door. The territory is already undersized, "
                "removing address would further shrink it"
            ),
            "POTENTIAL_UNDERSIZED": (
                "Keep as door-to-door if accessible from street level. "
                "Removing these units would shrink the territory to undersized."
            ),
            "IDEAL_TO_IDEAL": (
                "Indifferent. The territory remains in the target range with "
                "or without its units. If a border is adjusted, reconsider."
            ),
            "OVERSIZED_TO_OVERSIZED": (
                "Consider letter writing. Even without these units, the "
                "territory remains oversized. Further adjustments would be needed."
            ),
            "OVERSIZED_TO_IDEAL": (
                "Ideal candidate for letter writing. Removing this building "
                "brings the territory into the target range."
            ),
            "MANUAL": "Review building manually",
        }
        apt_groups["_Action_Code"] = apt_groups.apply(
            get_apartment_action_code,
            axis=1,
        )
        apt_groups["Suggested Action"] = apt_groups["_Action_Code"].map(
            apartment_action_text
        )

        if not apt_groups.empty:
            apt_export = apt_groups.rename(
                columns={
                    "Territory_Name": "Territory Name",
                    "Base_Address": "Base Address",
                    "Total Units": "Units",
                }
            )[
                [
                    "Base Address",
                    "Units",
                    "Territory Name",
                    "Total Addresses in Territory",
                    "Current Territory Status",
                    "Suggested Action",
                    "Source Rows",
                    "Nonblank Unit Rows",
                    "Unique Normalized Units",
                    "Blank Parent Rows",
                    "Duplicate Units",
                    "Reported County Unit Count",
                    "Apartment Confidence",
                    "Detection Reason",
                ]
            ]
        else:
            apt_export = pd.DataFrame(
                columns=[
                    "Base Address",
                    "Units",
                    "Territory Name",
                    "Total Addresses in Territory",
                    "Current Territory Status",
                    "Suggested Action",
                    "Source Rows",
                    "Nonblank Unit Rows",
                    "Unique Normalized Units",
                    "Blank Parent Rows",
                    "Duplicate Units",
                    "Reported County Unit Count",
                    "Apartment Confidence",
                    "Detection Reason",
                ]
            )

        apt_export.to_excel(writer, sheet_name="Apartments", index=False)
        ws4 = writer.sheets["Apartments"]
        ws4.freeze_panes = "C2"

        apartment_table = Table(
            displayName="ApartmentsTable",
            ref=(
                f"A1:{openpyxl.utils.get_column_letter(len(apt_export.columns))}"
                f"{len(apt_export) + 1}"
            ),
        )
        ws4.add_table(apartment_table)

        apartment_widths = {
            "A": 43,
            "B": 11,
            "C": 14,
            "D": 18,
            "E": 18,
            "F": 107,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 18,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 57,
        }
        for column_letter, width in apartment_widths.items():
            ws4.column_dimensions[column_letter].width = width
        for column_letter in ["G", "H", "I", "J", "K", "L", "M", "N"]:
            ws4.column_dimensions[column_letter].hidden = True
        ws4.delete_cols(15, 12)

        header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        apartment_bottom_border = Border(
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="EAECEB")
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )

        address_row_lookup = {}
        for dataframe_position, (_, address_row) in enumerate(
            address_list_df.iterrows(),
            start=2,
        ):
            base_address_value = clean_field(address_row.get("Base_Address"))
            if base_address_value and base_address_value not in address_row_lookup:
                address_row_lookup[base_address_value] = dataframe_position

        apartment_bold_phrases = {
            "TEN_PLUS": "Ideal for letter writing.",
            "CURRENT_UNDERSIZED": "Keep as door-to-door.",
            "POTENTIAL_UNDERSIZED": (
                "Keep as door-to-door if accessible from street level."
            ),
            "IDEAL_TO_IDEAL": "Indifferent.",
            "OVERSIZED_TO_OVERSIZED": "Consider letter writing.",
            "OVERSIZED_TO_IDEAL": "Ideal candidate for letter writing.",
            "MANUAL": "Review building manually",
        }

        for row_number in range(2, len(apt_export) + 2):
            row_fill = white_fill if row_number % 2 == 0 else stripe_fill
            for column_number in range(1, len(apt_export.columns) + 1):
                cell = ws4.cell(row=row_number, column=column_number)
                cell.fill = row_fill
                cell.border = apartment_bottom_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_number in {1, 3, 6, 14} else "center"
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            base_address = clean_field(ws4.cell(row=row_number, column=1).value)
            units_cell = ws4.cell(row=row_number, column=2)
            address_list_row = address_row_lookup.get(base_address)
            if address_list_row is not None:
                units_value = units_cell.value
                units_cell.value = (
                    f'=HYPERLINK("#\'Address List\'!A{address_list_row}", '
                    f'"{units_value}")'
                )
                units_cell.font = Font(color="0563C1", underline="single")

            action_code = apt_groups.iloc[row_number - 2]["_Action_Code"]
            full_text = apartment_action_text[action_code]
            bold_phrase = apartment_bold_phrases[action_code]
            remainder = full_text[len(bold_phrase):]
            ws4.cell(row=row_number, column=6).value = CellRichText(
                [TextBlock(InlineFont(b=True), bold_phrase), remainder]
            )

        # --- TAB 5: TERRITORY BALANCING ---
        balancing_columns = [
            "Territory",
            "Action Type",
            "Target Territory",
            "Priority",
            "Addresses Affected",
            "Projected Statuses",
            "Why",
        ]

        def distance_from_ideal(count, goal_minimum, goal_maximum):
            if count < goal_minimum:
                return goal_minimum - count
            if count > goal_maximum:
                return count - goal_maximum
            return 0

        def status_with_count(count, category_function):
            integer_count = int(count)
            return f"{integer_count} ({category_function(integer_count)})"

        def materially_improved(original_distance, projected_distance):
            if original_distance <= 0:
                return False

            improvement = original_distance - projected_distance
            proportional_improvement = improvement / original_distance

            return improvement >= 10 or proportional_improvement >= 0.25

        def evaluate_spatial_shift(
            source_count,
            target_count,
            goal_minimum,
            goal_maximum,
            category_function,
        ):
            equalizing_shift = max(
                int((source_count - target_count) // 2),
                0,
            )

            if equalizing_shift < 1:
                return None

            minimum_for_source_to_be_ideal = max(
                source_count - goal_maximum,
                0,
            )
            minimum_for_target_to_be_ideal = max(
                goal_minimum - target_count,
                0,
            )
            minimum_to_make_both_ideal = max(
                minimum_for_source_to_be_ideal,
                minimum_for_target_to_be_ideal,
            )

            maximum_that_keeps_source_ideal = max(
                source_count - goal_minimum,
                0,
            )
            maximum_that_keeps_target_ideal = max(
                goal_maximum - target_count,
                0,
            )
            maximum_for_both_ideal = min(
                maximum_that_keeps_source_ideal,
                maximum_that_keeps_target_ideal,
                equalizing_shift,
            )

            if (
                minimum_to_make_both_ideal >= 1
                and minimum_to_make_both_ideal <= maximum_for_both_ideal
            ):
                shift = int(minimum_to_make_both_ideal)
                projected_source = source_count - shift
                projected_target = target_count + shift

                return {
                    "shift": shift,
                    "priority": "High",
                    "projected_source": projected_source,
                    "projected_target": projected_target,
                    "why": (
                        f"Shifting {shift} addresses brings both territories "
                        "into the ideal range."
                    ),
                }

            original_source_distance = distance_from_ideal(
                source_count,
                goal_minimum,
                goal_maximum,
            )
            original_target_distance = distance_from_ideal(
                target_count,
                goal_minimum,
                goal_maximum,
            )

            evaluated_shifts = []

            for shift in range(1, equalizing_shift + 1):
                projected_source = source_count - shift
                projected_target = target_count + shift

                projected_source_status = category_function(
                    projected_source
                )
                projected_target_status = category_function(
                    projected_target
                )

                projected_source_distance = distance_from_ideal(
                    projected_source,
                    goal_minimum,
                    goal_maximum,
                )
                projected_target_distance = distance_from_ideal(
                    projected_target,
                    goal_minimum,
                    goal_maximum,
                )

                source_is_ideal = projected_source_status == "Ideal"
                target_is_ideal = projected_target_status == "Ideal"
                ideal_count = int(source_is_ideal) + int(target_is_ideal)

                source_materially_improved = materially_improved(
                    original_source_distance,
                    projected_source_distance,
                )
                target_materially_improved = materially_improved(
                    original_target_distance,
                    projected_target_distance,
                )

                one_is_ideal = ideal_count == 1
                other_materially_improved = (
                    source_is_ideal and target_materially_improved
                ) or (
                    target_is_ideal and source_materially_improved
                )

                total_distance_improvement = (
                    original_source_distance
                    + original_target_distance
                    - projected_source_distance
                    - projected_target_distance
                )

                if one_is_ideal and other_materially_improved:
                    priority = "Medium"
                    priority_rank = 2
                else:
                    priority = "Low"
                    priority_rank = 1

                evaluated_shifts.append(
                    {
                        "shift": shift,
                        "priority": priority,
                        "priority_rank": priority_rank,
                        "ideal_count": ideal_count,
                        "projected_source": projected_source,
                        "projected_target": projected_target,
                        "projected_source_distance": (
                            projected_source_distance
                        ),
                        "projected_target_distance": (
                            projected_target_distance
                        ),
                        "total_distance_improvement": (
                            total_distance_improvement
                        ),
                    }
                )

            if not evaluated_shifts:
                return None

            evaluated_shifts.sort(
                key=lambda result: (
                    -result["priority_rank"],
                    -result["ideal_count"],
                    -result["total_distance_improvement"],
                    (
                        result["projected_source_distance"]
                        + result["projected_target_distance"]
                    ),
                    result["shift"],
                )
            )

            best_result = evaluated_shifts[0]

            if best_result["priority"] == "Medium":
                best_result["why"] = (
                    f"Shifting {best_result['shift']} addresses brings one "
                    "territory into the ideal range and materially improves "
                    "the other."
                )
            else:
                best_result["why"] = (
                    f"Shifting {best_result['shift']} addresses improves the "
                    "imbalance but does not fully resolve both territories."
                )

            return best_result

        apartment_units_by_territory = (
            apt_groups.groupby(
                "Territory_Name",
                observed=True,
            )["Total Units"]
            .sum()
            .to_dict()
        )

        territory_metrics = counts_df[
            [
                "Territory_Name",
                "Total_Addresses",
                "Category",
            ]
        ].copy()

        territory_metrics["Apartment_Units"] = (
            territory_metrics["Territory_Name"]
            .map(apartment_units_by_territory)
            .fillna(0)
            .astype(int)
        )

        territory_metrics["Potential_Count"] = (
            territory_metrics["Total_Addresses"]
            - territory_metrics["Apartment_Units"]
        ).clip(lower=0)

        territory_metrics["Potential_Status"] = territory_metrics[
            "Potential_Count"
        ].apply(get_category)

        territory_metrics["Shift_Baseline_Count"] = territory_metrics[
            "Total_Addresses"
        ]
        territory_metrics["Resolved"] = False

        territory_metrics = territory_metrics.set_index(
            "Territory_Name"
        )

        balancing_rows = []

        # Phase 1: Reductions.
        for territory_name, territory_row in territory_metrics.iterrows():
            raw_count = int(territory_row["Total_Addresses"])
            current_status = territory_row["Category"]
            apartment_units = int(territory_row["Apartment_Units"])
            potential_count = int(territory_row["Potential_Count"])
            potential_status = territory_row["Potential_Status"]

            if current_status != "Oversized" or apartment_units <= 0:
                continue

            if potential_status == "Ideal":
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Reduction",
                        "Target Territory": "Internal",
                        "Priority": "High",
                        "Addresses Affected": apartment_units,
                        "Projected Statuses": (
                            f"{raw_count} (Oversized with Apartments) -> "
                            f"{potential_count} (Ideal W/O Apartments)"
                        ),
                        "Why": (
                            "Converting all apartments to letter writing "
                            "brings this territory into the ideal range "
                            "without border adjustments."
                        ),
                    }
                )
                territory_metrics.at[
                    territory_name,
                    "Resolved",
                ] = True

            elif potential_status == "Oversized":
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Reduction",
                        "Target Territory": "Internal",
                        "Priority": "Medium",
                        "Addresses Affected": apartment_units,
                        "Projected Statuses": (
                            f"{raw_count} (Oversized) -> "
                            f"{potential_count} (Oversized)"
                        ),
                        "Why": (
                            "Converting all apartments reduces bloat, but "
                            "the territory remains oversized. A border shift "
                            "is still required."
                        ),
                    }
                )
                territory_metrics.at[
                    territory_name,
                    "Shift_Baseline_Count",
                ] = potential_count

            elif potential_status == "Undersized":
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Review Warning",
                        "Target Territory": "Internal",
                        "Priority": "Low",
                        "Addresses Affected": "Review",
                        "Projected Statuses": (
                            f"{raw_count} (Oversized) -> "
                            f"{potential_count} (Undersized)"
                        ),
                        "Why": (
                            "Warning: Converting all apartments drops this "
                            "territory below the minimum goal. To avoid a "
                            "partial conversion (e.g., converting 2 buildings "
                            "but leaving 1), consider a border shift instead."
                        ),
                    }
                )

        terr_geoms = (
            kml_gdf[
                [
                    "Territory_Name",
                    "geometry_terr",
                ]
            ]
            .dropna(
                subset=[
                    "Territory_Name",
                    "geometry_terr",
                ]
            )
            .set_geometry("geometry_terr")
            .dissolve(by="Territory_Name")
        )

        terr_geoms["geometry_terr"] = (
            terr_geoms.geometry.make_valid()
        )

        terr_geoms = terr_geoms[
            terr_geoms.geometry.notna()
            & ~terr_geoms.geometry.is_empty
        ].copy()

        terr_geoms_metric = terr_geoms.to_crs(metric_crs)
        territory_sindex = terr_geoms_metric.sindex

        # Phase 2: Consolidations.
        unresolved_undersized = set(
            territory_metrics.index[
                territory_metrics["Category"].eq("Undersized")
                & ~territory_metrics["Resolved"]
            ]
        )

        consolidation_pairs = set()

        for territory_name in sorted(
            unresolved_undersized,
            key=natural_keys,
        ):
            if (
                territory_name not in terr_geoms_metric.index
                or territory_name not in unresolved_undersized
            ):
                continue

            territory_geom = terr_geoms_metric.at[
                territory_name,
                "geometry_terr",
            ]

            candidate_positions = territory_sindex.query(
                territory_geom.buffer(45.0),
                predicate="intersects",
            )

            for candidate_position in candidate_positions:
                neighbor_name = terr_geoms_metric.index[
                    candidate_position
                ]

                if (
                    neighbor_name == territory_name
                    or neighbor_name not in unresolved_undersized
                ):
                    continue

                pair_key = tuple(
                    sorted(
                        (
                            str(territory_name),
                            str(neighbor_name),
                        )
                    )
                )

                if pair_key in consolidation_pairs:
                    continue

                neighbor_geom = terr_geoms_metric.iloc[
                    candidate_position
                ].geometry_terr

                if territory_geom.distance(neighbor_geom) > 45.0:
                    continue

                combined_count = int(
                    territory_metrics.at[
                        territory_name,
                        "Total_Addresses",
                    ]
                    + territory_metrics.at[
                        neighbor_name,
                        "Total_Addresses",
                    ]
                )

                if get_category(combined_count) != "Ideal":
                    continue

                consolidation_pairs.add(pair_key)

                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Consolidation",
                        "Target Territory": neighbor_name,
                        "Priority": "High",
                        "Addresses Affected": "Merge",
                        "Projected Statuses": (
                            f"{territory_name}: "
                            f"{int(territory_metrics.at[territory_name, 'Total_Addresses'])} + "
                            f"{neighbor_name}: "
                            f"{int(territory_metrics.at[neighbor_name, 'Total_Addresses'])} = "
                            f"{combined_count} (Ideal when combined)"
                        ),
                        "Why": (
                            "Merging these adjacent undersized territories "
                            "creates a single ideal territory and reduces "
                            "map bloat."
                        ),
                    }
                )

                territory_metrics.at[
                    territory_name,
                    "Resolved",
                ] = True
                territory_metrics.at[
                    neighbor_name,
                    "Resolved",
                ] = True

                unresolved_undersized.discard(territory_name)
                unresolved_undersized.discard(neighbor_name)
                break

        # Phase 3: Border Shifts.
        unresolved_territories = territory_metrics.index[
            ~territory_metrics["Resolved"]
        ]

        shift_sources = [
            territory_name
            for territory_name in unresolved_territories
            if get_category(
                int(
                    territory_metrics.at[
                        territory_name,
                        "Shift_Baseline_Count",
                    ]
                )
            )
            == "Oversized"
        ]

        seen_shift_pairs = set()

        for source_name in shift_sources:
            if source_name not in terr_geoms_metric.index:
                continue

            source_count = int(
                territory_metrics.at[
                    source_name,
                    "Shift_Baseline_Count",
                ]
            )

            source_geom = terr_geoms_metric.at[
                source_name,
                "geometry_terr",
            ]

            candidate_positions = territory_sindex.query(
                source_geom.buffer(45.0),
                predicate="intersects",
            )

            for candidate_position in candidate_positions:
                target_name = terr_geoms_metric.index[
                    candidate_position
                ]

                if (
                    target_name == source_name
                    or target_name not in unresolved_territories
                ):
                    continue

                target_count = int(
                    territory_metrics.at[
                        target_name,
                        "Shift_Baseline_Count",
                    ]
                )

                if get_category(target_count) != "Undersized":
                    continue

                pair_key = tuple(
                    sorted(
                        (
                            str(source_name),
                            str(target_name),
                        )
                    )
                )

                if pair_key in seen_shift_pairs:
                    continue

                target_geom = terr_geoms_metric.iloc[
                    candidate_position
                ].geometry_terr

                if source_geom.distance(target_geom) > 45.0:
                    continue

                seen_shift_pairs.add(pair_key)

                shift_result = evaluate_spatial_shift(
                    source_count,
                    target_count,
                    min_goal,
                    max_goal,
                    get_category,
                )

                if shift_result is None:
                    continue

                projected_source = int(
                    shift_result["projected_source"]
                )
                projected_target = int(
                    shift_result["projected_target"]
                )

                balancing_rows.append(
                    {
                        "Territory": source_name,
                        "Action Type": "Border Shift",
                        "Target Territory": target_name,
                        "Priority": shift_result["priority"],
                        "Addresses Affected": int(
                            shift_result["shift"]
                        ),
                        "Projected Statuses": (
                            f"{source_name}: "
                            f"{status_with_count(projected_source, get_category)}"
                            " | "
                            f"{target_name}: "
                            f"{status_with_count(projected_target, get_category)}"
                        ),
                        "Why": (
                            "(Candidate territories are within the configured "
                            "45m boundary tolerance). "
                            f"{shift_result['why']}"
                        ),
                    }
                )

        territory_balancing_df = pd.DataFrame(
            balancing_rows,
            columns=balancing_columns,
        )

        count_lookup = (
            counts_df.set_index("Territory_Name")["Total_Addresses"].to_dict()
            if not counts_df.empty
            else {}
        )

        if not territory_balancing_df.empty:
            territory_balancing_df["Target Territory"] = (
                territory_balancing_df["Target Territory"]
                .replace({"Internal": "Internal (Apartments)"})
            )
            territory_balancing_df["OT Count"] = (
                territory_balancing_df["Territory"]
                .map(count_lookup)
                .fillna(0)
                .astype(int)
            )
            territory_balancing_df["TT Count"] = (
                territory_balancing_df["Target Territory"].map(
                    lambda target: (
                        "N/A"
                        if target == "Internal (Apartments)"
                        else int(count_lookup.get(target, 0))
                    )
                )
            )
            originating_order = sorted(
                territory_balancing_df["Territory"].astype(str).unique(),
                key=natural_keys,
            )
            originating_rank = {
                territory_name: rank
                for rank, territory_name in enumerate(originating_order)
            }
            territory_balancing_df["_Originating_Sort"] = (
                territory_balancing_df["Territory"]
                .astype(str)
                .map(originating_rank)
            )
            territory_balancing_df = (
                territory_balancing_df.sort_values(
                    by="_Originating_Sort",
                    kind="stable",
                )
                .drop(columns=["_Originating_Sort"])
                .reset_index(drop=True)
            )
        else:
            territory_balancing_df["OT Count"] = pd.Series(dtype="int64")
            territory_balancing_df["TT Count"] = pd.Series(dtype="object")

        territory_balancing_df = (
            territory_balancing_df.rename(
                columns={
                    "Territory": "Originating Territory",
                    "Target Territory": "Targeted Territory",
                    "Action Type": "Balancing Method",
                    "Addresses Affected": "Est. Addresses Affected",
                    "Why": "Why?/Comments",
                }
            )[
                [
                    "Originating Territory",
                    "OT Count",
                    "Targeted Territory",
                    "TT Count",
                    "Balancing Method",
                    "Priority",
                    "Est. Addresses Affected",
                    "Projected Statuses",
                    "Why?/Comments",
                ]
            ]
        )

        territory_balancing_df.to_excel(
            writer,
            sheet_name="Territory Balancing",
            index=False,
        )

        ws5 = writer.sheets["Territory Balancing"]
        ws5.freeze_panes = "D2"

        balancing_widths = {
            "A": 14,
            "B": 7,
            "C": 21,
            "D": 7,
            "E": 14,
            "F": 14,
            "G": 16,
            "H": 64,
            "I": 92,
        }

        for column_letter, width in balancing_widths.items():
            ws5.column_dimensions[column_letter].width = width

        balancing_header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        balancing_stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        balancing_white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        high_priority_fill = PatternFill(
            start_color="EA9D9C",
            end_color="EA9D9C",
            fill_type="solid",
        )
        medium_priority_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        balancing_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws5[1]:
            cell.fill = balancing_header_fill
            cell.font = Font(
                bold=True,
                color="EAECEB",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="top",
                wrap_text=True,
            )
            cell.border = balancing_border

        for row_number in range(
            2,
            len(territory_balancing_df) + 2,
        ):
            row_fill = (
                balancing_white_fill
                if row_number % 2 == 0
                else balancing_stripe_fill
            )

            for column_number in range(1, 10):
                cell = ws5.cell(
                    row=row_number,
                    column=column_number,
                )
                cell.fill = row_fill
                cell.border = balancing_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_number in {1, 3, 8, 9}
                        else "center"
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            balancing_method_cell = ws5.cell(
                row=row_number,
                column=5,
            )
            priority_cell = ws5.cell(
                row=row_number,
                column=6,
            )

            if balancing_method_cell.value == "Review Warning":
                balancing_method_cell.fill = high_priority_fill

            if priority_cell.value == "High":
                priority_cell.fill = high_priority_fill
            elif priority_cell.value == "Medium":
                priority_cell.fill = medium_priority_fill

        # --- TAB 6: EXCLUDED AUDIT ---
        if not excluded_gdf.empty:
            excluded_gdf[["HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort"]] = excluded_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
            excluded_gdf["Unit_Sort"] = excluded_gdf["Canonical_Unit"].map(clean_field).str.upper()
            excluded_list_df = excluded_gdf.sort_values(by=["Territory_Name", "Canonical_Street", "HouseNum_Sort", "HouseNum_Suffix_Rank", "HouseNum_Text_Sort", "Unit_Sort"], kind="stable")
            
            export_ex_df = excluded_list_df[[
                "Source_Record_ID", "Territory_Name", "NWS_Category", "NWS_Number", "Mailable_Address", "Canonical_Status", "Data_Quality_Flag",
                "Canonical_HouseNo", "Canonical_Street", "Canonical_Unit", "Canonical_Zip_Code"
            ]].rename(columns={
                "Source_Record_ID": "Source Record ID", "Territory_Name": "Territory Name", "Mailable_Address": "Mailable Address", 
                "Canonical_Status": "Exclusion Reason", "Data_Quality_Flag": "Data Quality Flag",
                "Canonical_HouseNo": "HouseNo", "Canonical_Street": "Street", "Canonical_Unit": "Unit", "Canonical_Zip_Code": "Zip_Code"
            })
            export_ex_df.to_excel(writer, sheet_name="Excluded Audit", index=False)
            ws6 = writer.sheets["Excluded Audit"]
            add_excel_table(ws6, export_ex_df, "ExcludedAuditTable", show_stripes=True)

            for column_letter in ["H", "I", "J", "K"]: ws6.column_dimensions[column_letter].hidden = True
            ws6.column_dimensions["A"].width = 22
            ws6.column_dimensions["B"].width = 18
            ws6.column_dimensions["C"].width = 15
            ws6.column_dimensions["D"].width = 15
            ws6.column_dimensions["E"].width = 55
            ws6.column_dimensions["F"].width = 25
            ws6.column_dimensions["G"].width = 35
        else:
            empty_excluded_columns = ["Source Record ID", "Territory Name", "NWS_Category", "NWS_Number", "Mailable Address", "Exclusion Reason", "Data Quality Flag", "Canonical_HouseNo", "Canonical_Street", "Canonical_Unit", "Canonical_Zip_Code"]
            pd.DataFrame(columns=empty_excluded_columns).to_excel(writer, sheet_name="Excluded Audit", index=False)
            writer.sheets["Excluded Audit"].cell(row=2, column=1, value="No addresses were excluded in this map area.")

        # --- EXCEL UX POLISH ---
        writer.sheets["Territory Balancing"].freeze_panes = "D2"
        writer.sheets["Excluded Audit"].freeze_panes = "A2"

        writer.sheets["Dashboard"].sheet_properties.tabColor = "1E90FF"
        writer.sheets["Counts"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Address List"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Apartments"].sheet_properties.tabColor = "FF8C00"
        writer.sheets["Territory Balancing"].sheet_properties.tabColor = "FF0000"
        writer.sheets["Excluded Audit"].sheet_properties.tabColor = "808080"

    output.seek(0)
    return output

# --- 4. EXECUTION FLOW ---
if "last_uploaded_kml" not in st.session_state:
    st.session_state["last_uploaded_kml"] = None

if uploaded_kml != st.session_state["last_uploaded_kml"]:
    if "excel_data" in st.session_state:
        del st.session_state["excel_data"]
    st.session_state["last_uploaded_kml"] = uploaded_kml

if uploaded_kml:
    if st.button("Generate Territory Analysis"):
        status_placeholder = st.empty()
        status_placeholder.info("Analysis engine stops for coffee…")
        
        parcel_gdf = load_county_data(selected_county)

        if parcel_gdf is not None:
            status_placeholder.info("Analysis engine gets a new call…")
            
            county_config = COUNTY_CONFIGS[selected_county]
            parcel_gdf = parcel_gdf.reset_index(drop=True).copy()
            parcel_gdf = parcel_gdf.rename(columns=county_config["column_mapping"], errors="ignore")

            missing_columns = [
                column for column in REQUIRED_CANONICAL_COLUMNS if column not in parcel_gdf.columns
            ]
            if missing_columns:
                status_placeholder.empty()
                st.error("County data failed preflight validation. Missing required canonical columns: " + ", ".join(missing_columns))
                st.stop()

            native_id_col = "Canonical_Native_Source_ID"
            county_id_prefix = re.sub(r"[^A-Za-z0-9]+", "_", selected_county.upper()).strip("_")
            fallback_ids = pd.Series([f"{county_id_prefix}-FALLBACK-{row_number:09d}" for row_number in range(1, len(parcel_gdf) + 1)], index=parcel_gdf.index, dtype="string")
            
            if native_id_col in parcel_gdf.columns:
                native_ids = parcel_gdf[native_id_col].map(clean_field)
                parcel_gdf["Source_Record_ID"] = native_ids.where(native_ids.ne(""), fallback_ids)
            else:
                parcel_gdf["Source_Record_ID"] = fallback_ids

            duplicate_native_ids = parcel_gdf["Source_Record_ID"].duplicated(keep=False)
            if duplicate_native_ids.any():
                duplicate_sequence = (parcel_gdf.groupby("Source_Record_ID").cumcount() + 1).astype(str)
                parcel_gdf.loc[duplicate_native_ids, "Source_Record_ID"] = parcel_gdf.loc[duplicate_native_ids, "Source_Record_ID"] + "-DUP-" + duplicate_sequence.loc[duplicate_native_ids]

            status_placeholder.info("Analysis engine makes a return visit…")

            try:
                kml_gdf = gpd.read_file(uploaded_kml, driver="KML")
                if kml_gdf.crs is None: kml_gdf = kml_gdf.set_crs("EPSG:4326", allow_override=True)
                if parcel_gdf.crs is None: raise ValueError("The parcel dataset has no CRS. Assign the correct source CRS before processing.")

                kml_gdf = kml_gdf.copy()
                kml_gdf["geometry"] = kml_gdf.geometry.make_valid()
                kml_gdf = kml_gdf[kml_gdf.geometry.notna() & ~kml_gdf.geometry.is_empty].copy()

                fallback_names = "Territory_" + kml_gdf.index.to_series().astype(str)
                if "Name" in kml_gdf.columns: kml_gdf["Territory_Name"] = kml_gdf["Name"].replace(r"^\s*$", pd.NA, regex=True).fillna(fallback_names)
                elif "Description" in kml_gdf.columns: kml_gdf["Territory_Name"] = kml_gdf["Description"].replace(r"^\s*$", pd.NA, regex=True).fillna(fallback_names)
                else: kml_gdf["Territory_Name"] = fallback_names

                if parcel_gdf.crs != kml_gdf.crs: parcel_gdf = parcel_gdf.to_crs(kml_gdf.crs)

                parcel_gdf = parcel_gdf.copy()
                parcel_gdf["geometry"] = parcel_gdf.geometry.make_valid()
                parcel_gdf = parcel_gdf[parcel_gdf.geometry.notna() & ~parcel_gdf.geometry.is_empty].copy()

                territory_envelope = kml_gdf.geometry.union_all().envelope
                parcel_gdf = parcel_gdf[parcel_gdf.geometry.intersects(territory_envelope)].copy()

                parcel_gdf["_join_point"] = parcel_gdf.geometry.representative_point()
                parcel_join_gdf = parcel_gdf.set_geometry("_join_point")
                kml_gdf = kml_gdf.rename(columns={"geometry": "geometry_terr"}).set_geometry("geometry_terr")

                joined_gdf = gpd.sjoin(
                    parcel_join_gdf,
                    kml_gdf[["Territory_Name", "geometry_terr"]],
                    how="inner",
                    predicate="covered_by",
                )
                joined_gdf = joined_gdf.dropna(
                    subset=["Territory_Name"]
                ).copy()
                assigned_source_count = joined_gdf[
                    "Source_Record_ID"
                ].nunique()
                unassigned_address_count = max(
                    len(parcel_join_gdf) - assigned_source_count,
                    0,
                )

                joined_gdf["_Territory_Sort"] = joined_gdf["Territory_Name"].astype(str).str.upper()
                joined_gdf = joined_gdf.sort_values(by=["Source_Record_ID", "_Territory_Sort"], kind="stable")
                status_placeholder.info("Analysis engine stamps a letter…")

                duplicate_assignment_count = joined_gdf.duplicated(subset=["Source_Record_ID"], keep="first").sum()
                joined_gdf = joined_gdf.drop_duplicates(subset=["Source_Record_ID"], keep="first").copy()
                
                joined_gdf = joined_gdf.drop(columns=["_Territory_Sort"], errors="ignore")
                joined_gdf = joined_gdf.set_geometry("geometry")
                joined_gdf = joined_gdf.drop(columns=["_join_point", "index_right"], errors="ignore")

                if duplicate_assignment_count > 0:
                    st.warning(f"{duplicate_assignment_count:,} duplicate boundary assignment(s) were resolved by retaining the first territory match for each Source_Record_ID.")

                status_placeholder.info(random.choice([
                    "Analysis engine stops for coffee…",
                    "Analysis engine gets a new call…",
                    "Analysis engine makes a return visit…",
                    "Analysis engine stamps a letter…",
                ]))

                excel_file = generate_excel_report(
                    joined_gdf, kml_gdf, MIN_GOAL, MAX_GOAL, congregation_name.replace(" ", ""), 
                    county_config,
                    apt_threshold=apartment_threshold,
                    kml_filename=uploaded_kml.name,
                    county_filename=county_config["file_path"],
                    duplicate_assignment_count=duplicate_assignment_count,
                    unassigned_address_count=unassigned_address_count,
                )
                
                safe_congregation_name = re.sub(
                    r"[^A-Za-z0-9_-]+",
                    "",
                    congregation_name.replace(" ", ""),
                ) or "Congregation"
                filename = (
                    f"{safe_congregation_name}-TerritoryAnalysis-"
                    f"{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
                )
                st.session_state["excel_data"] = excel_file.getvalue()
                st.session_state["excel_filename"] = filename
                
                status_placeholder.success("Analysis Complete! Download the generated file below")

            except Exception as error:
                status_placeholder.empty()
                st.error(f"An error occurred during processing: {error}")

if "excel_data" in st.session_state:
    st.download_button(
        label="⬇️ Download Excel Analysis",
        data=st.session_state["excel_data"],
        file_name=st.session_state["excel_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
