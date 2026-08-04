import streamlit as st
import geopandas as gpd
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.worksheet.table import Table, TableStyleInfo
import fiona
import io
import datetime
import hashlib
import json
import os
import re
import tempfile
from pathlib import Path
from fractions import Fraction
from itertools import combinations

import requests
from shapely.geometry import box

# Enable KML support in GeoPandas.
fiona.drvsupport.supported_drivers["KML"] = "rw"
fiona.drvsupport.supported_drivers["LIBKML"] = "rw"

st.set_page_config(
    page_title="TerritoryToolbox's Analysis Engine",
    layout="wide",
)

# --- CLOUDFLARE R2 RELEASE CONFIGURATION ---
R2_PUBLIC_BASE_URL = "https://data.territorytoolbox.com"
R2_RELEASE_POINTER_PATH = "current/release.json"
R2_METADATA_CACHE_TTL_SECONDS = 600
R2_DOWNLOAD_TIMEOUT_SECONDS = 120
R2_BBOX_PADDING_DEGREES = 0.001
R2_CACHE_DIRECTORY = Path(tempfile.gettempdir()) / "territorytoolbox_r2_cache"
DEFAULT_STATE = "WI"
DEFAULT_METRIC_CRS = "EPSG:3071"

WISCONSIN_COUNTY_ORDER = (
    "Adams", "Ashland", "Barron", "Bayfield", "Brown", "Buffalo",
    "Burnett", "Calumet", "Chippewa", "Clark", "Columbia", "Crawford",
    "Dane", "Dodge", "Door", "Douglas", "Dunn", "Eau Claire", "Florence",
    "Fond du Lac", "Forest", "Grant", "Green", "Green Lake", "Iowa",
    "Iron", "Jackson", "Jefferson", "Juneau", "Kenosha", "Kewaunee",
    "La Crosse", "Lafayette", "Langlade", "Lincoln", "Manitowoc",
    "Marathon", "Marinette", "Marquette", "Menominee", "Milwaukee",
    "Monroe", "Oconto", "Oneida", "Outagamie", "Ozaukee", "Pepin",
    "Pierce", "Polk", "Portage", "Price", "Racine", "Richland", "Rock",
    "Rusk", "St. Croix", "Sauk", "Sawyer", "Shawano", "Sheboygan",
    "Taylor", "Trempealeau", "Vernon", "Vilas", "Walworth", "Washburn",
    "Washington", "Waukesha", "Waupaca", "Waushara", "Winnebago", "Wood",
)

# Milwaukee remains on its existing county-specific source. All other available
# counties are determined dynamically from the active R2 manifest.
COUNTY_OVERRIDE_CONFIGS = {
    "Milwaukee": {
        "file_path": "zip://data/Milwaukee_Datapoints07072026.zip",
        "state": "WI",
        "metric_crs": "EPSG:3071",
        "native_source_id": "TAXKEY",
        "excluded_statuses": [
            "Undeveloped",
            "Parking Lot",
            "ROW",
            "Park or Recreational Facility",
            "Undeveloped Outlot",
            "Sliver or Remnant",
            "Non Addressable Assoc with Adj Parcel",
        ],
        "default_excluded_statuses": [
            "Undeveloped",
            "Parking Lot",
            "ROW",
            "Park or Recreational Facility",
            "Undeveloped Outlot",
            "Sliver or Remnant",
            "Non Addressable Assoc with Adj Parcel",
        ],
        "column_mapping": {
            "TAXKEY": "Canonical_Native_Source_ID",
            "HouseNo": "Canonical_HouseNo",
            "HouseSx": "Canonical_HouseSx",
            "Dir": "Canonical_Dir",
            "Street": "Canonical_Street",
            "StType": "Canonical_StType",
            "Muni": "Canonical_Muni",
            "Zip_Code": "Canonical_Zip_Code",
            "Unit": "Canonical_Unit",
            "Addr_Statu": "Canonical_Status",
        },
    },
}

STATEWIDE_REVIEW_STATUS_OPTIONS = [
    "Building Parent",
    "Unknown Unit Address",
]
STATEWIDE_DEFAULT_REVIEW_EXCLUSIONS = ["Building Parent"]

REQUIRED_CANONICAL_COLUMNS = [
    "Canonical_HouseNo", "Canonical_HouseSx", "Canonical_Dir", "Canonical_Street",
    "Canonical_StType", "Canonical_Muni", "Canonical_Zip_Code", "Canonical_Unit",
    "Canonical_Status", "geometry",
]

CROSS_COUNTY_DUPLICATE_TOLERANCE_METERS = 5.0
BOUNDARY_AUDIT_BUFFER_METERS = 45.0


def clean_field(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}:
        return ""
    return text


def natural_keys(text):
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", str(text))
    ]


def assign_territory_names(kml_gdf):
    kml_gdf = kml_gdf.copy()
    fallback_names = "Territory_" + kml_gdf.index.to_series().astype(str)

    if "Name" in kml_gdf.columns:
        names = kml_gdf["Name"]
    elif "Description" in kml_gdf.columns:
        names = kml_gdf["Description"]
    else:
        names = fallback_names

    if isinstance(names, pd.Series):
        names = names.replace(r"^\s*$", pd.NA, regex=True).fillna(
            fallback_names
        )

    kml_gdf["Territory_Name"] = names.astype(str).str.strip()
    return kml_gdf


def detect_territory_group(territory_name):
    name = clean_field(territory_name)
    if not name:
        return "Residential"

    group_name = re.sub(
        r"(?:[\s_-]+)?\d+[A-Za-z]?$",
        "",
        name,
    ).strip(" -_")
    group_name = re.sub(r"[\s_-]+", " ", group_name).strip()

    if not group_name or group_name.lower() in {
        "territory",
        "imported",
        "(imported)",
    }:
        return "Residential"
    return group_name


def apply_territory_groups(kml_gdf, group_overrides=None):
    group_overrides = group_overrides or {}
    kml_gdf = kml_gdf.copy()
    kml_gdf["Detected_Territory_Group"] = kml_gdf[
        "Territory_Name"
    ].map(detect_territory_group)
    kml_gdf["Territory_Group"] = kml_gdf[
        "Detected_Territory_Group"
    ].map(lambda group: clean_field(group_overrides.get(group, group)))
    kml_gdf["Territory_Group"] = kml_gdf["Territory_Group"].replace(
        "",
        "Residential",
    )
    return kml_gdf


def build_territory_order(kml_gdf):
    territory_records = (
        kml_gdf[["Territory_Name", "Territory_Group"]]
        .dropna(subset=["Territory_Name"])
        .drop_duplicates(subset=["Territory_Name"], keep="first")
    )
    records = territory_records.to_dict("records")
    records.sort(
        key=lambda record: (
            natural_keys(record["Territory_Group"]),
            natural_keys(record["Territory_Name"]),
        )
    )
    territory_order = [record["Territory_Name"] for record in records]
    territory_rank = {
        territory_name: rank
        for rank, territory_name in enumerate(territory_order)
    }
    return territory_order, territory_rank


def territory_group_priority(territory_group):
    """Return the cross-group assignment priority for a territory group."""
    normalized_group = re.sub(
        r"[^a-z0-9]+",
        "",
        clean_field(territory_group).lower(),
    )
    if normalized_group.startswith("letterwriting"):
        return 0
    if normalized_group.startswith("residential"):
        return 1
    if normalized_group.startswith("business"):
        return 2
    return 3


def normalize_county_prefix(county_name):
    """Create a stable county prefix for globally unique source IDs."""
    return re.sub(
        r"[^A-Za-z0-9]+",
        "_",
        clean_field(county_name).upper(),
    ).strip("_")


def build_public_r2_url(relative_path):
    """Build a public HTTPS URL without accepting private endpoints."""
    clean_path = clean_field(relative_path).lstrip("/")
    if not clean_path:
        raise ValueError("The active data release contains an empty file path.")
    return f"{R2_PUBLIC_BASE_URL}/{clean_path}"


def _response_json(response, resource_name):
    try:
        return response.json()
    except ValueError as exc:
        raise ValueError(
            f"The {resource_name} response was not valid JSON."
        ) from exc


def _request_json(url, resource_name):
    try:
        response = requests.get(
            url,
            timeout=(15, R2_DOWNLOAD_TIMEOUT_SECONDS),
            headers={"User-Agent": "TerritoryToolbox-Analyzer/1.0"},
        )
        response.raise_for_status()
    except requests.RequestException as exc:
        raise ConnectionError(
            f"TerritoryToolbox could not load the {resource_name}. "
            "Please retry in a few minutes."
        ) from exc
    return _response_json(response, resource_name)


def validate_release_metadata(release):
    if not isinstance(release, dict):
        raise ValueError("The active release pointer has an invalid structure.")
    required_fields = (
        "release_version",
        "release_status",
        "manifest_path",
        "runtime_base_path",
        "runtime_schema_version",
        "manifest_schema_version",
    )
    missing = [field for field in required_fields if not clean_field(release.get(field))]
    if missing:
        raise ValueError(
            "The active release pointer is missing required fields: "
            + ", ".join(missing)
        )
    if clean_field(release.get("release_status")).lower() != "complete":
        raise ValueError("The active Wisconsin data release is not marked complete.")
    return release


@st.cache_data(ttl=R2_METADATA_CACHE_TTL_SECONDS, show_spinner=False)
def load_active_release():
    """Load and validate the stable current-release pointer."""
    release_url = build_public_r2_url(R2_RELEASE_POINTER_PATH)
    return validate_release_metadata(
        _request_json(release_url, "active Wisconsin data release")
    )


def _manifest_rows(manifest):
    if isinstance(manifest, dict):
        rows = manifest.get("counties")
    else:
        rows = manifest
    if not isinstance(rows, list):
        raise ValueError("The county manifest does not contain a counties list.")
    return rows


def _as_manifest_bool(value):
    if isinstance(value, bool):
        return value
    return clean_field(value).lower() in {"true", "1", "yes"}


def validate_manifest(manifest, release):
    rows = _manifest_rows(manifest)
    required_fields = {
        "canonical_county",
        "technical_validation_status",
        "production_source_status",
        "public_availability_status",
        "coverage_confidence_status",
        "publication_eligible",
        "analyzer_enabled",
        "included_in_publish_package",
        "published_runtime_relative_path",
        "runtime_schema_version",
    }
    seen_counties = set()
    for position, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"County manifest row {position} is invalid.")
        missing = required_fields.difference(row)
        if missing:
            raise ValueError(
                f"County manifest row {position} is missing: "
                + ", ".join(sorted(missing))
            )
        county_name = clean_field(row.get("canonical_county"))
        if not county_name:
            raise ValueError(f"County manifest row {position} has no county name.")
        if county_name in seen_counties:
            raise ValueError(f"The county manifest contains duplicate rows for {county_name}.")
        seen_counties.add(county_name)

        enabled = (
            _as_manifest_bool(row.get("analyzer_enabled"))
            and _as_manifest_bool(row.get("included_in_publish_package"))
            and _as_manifest_bool(row.get("publication_eligible"))
        )
        if enabled:
            if not clean_field(row.get("published_runtime_relative_path")):
                raise ValueError(
                    f"{county_name} is enabled but has no published runtime path."
                )
            if not clean_field(row.get("published_runtime_sha256")):
                raise ValueError(
                    f"{county_name} is enabled but has no published runtime hash."
                )
            if not row.get("published_runtime_byte_size"):
                raise ValueError(
                    f"{county_name} is enabled but has no published runtime size."
                )
            row_schema = clean_field(row.get("runtime_schema_version"))
            release_schema = clean_field(release.get("runtime_schema_version"))
            if row_schema and row_schema != release_schema:
                raise ValueError(
                    f"{county_name} runtime schema {row_schema} does not match "
                    f"the active release schema {release_schema}."
                )
    return manifest


@st.cache_data(ttl=R2_METADATA_CACHE_TTL_SECONDS, show_spinner=False)
def load_county_manifest(release_version, manifest_path, runtime_schema_version):
    """Load the release-specific manifest using cache identity from the release."""
    release_stub = {
        "runtime_schema_version": runtime_schema_version,
    }
    manifest_url = build_public_r2_url(manifest_path)
    manifest = _request_json(manifest_url, "Wisconsin county manifest")
    return validate_manifest(manifest, release_stub)


def get_manifest_county_lookup(manifest):
    return {
        clean_field(row.get("canonical_county")): row
        for row in _manifest_rows(manifest)
    }


def _manifest_row_is_enabled(row):
    return bool(
        row
        and _as_manifest_bool(row.get("analyzer_enabled"))
        and _as_manifest_bool(row.get("included_in_publish_package"))
        and _as_manifest_bool(row.get("publication_eligible"))
        and clean_field(row.get("published_runtime_relative_path"))
    )


def get_available_counties(manifest_lookup):
    enabled = {
        county_name
        for county_name, row in manifest_lookup.items()
        if _manifest_row_is_enabled(row)
    }
    enabled.update(COUNTY_OVERRIDE_CONFIGS)
    return [county for county in WISCONSIN_COUNTY_ORDER if county in enabled]


def get_county_source_strategy(county_name, manifest_lookup):
    if county_name in COUNTY_OVERRIDE_CONFIGS:
        return "county_override"
    row = manifest_lookup.get(county_name)
    if _manifest_row_is_enabled(row):
        return "statewide_runtime"
    raise ValueError(f"{county_name} County is not available in the active release.")


def get_county_exclusion_settings(county_name, manifest_lookup):
    strategy = get_county_source_strategy(county_name, manifest_lookup)
    if strategy == "county_override":
        config = COUNTY_OVERRIDE_CONFIGS[county_name]
        return (
            list(config.get("excluded_statuses", [])),
            list(config.get("default_excluded_statuses", [])),
        )
    return (
        list(STATEWIDE_REVIEW_STATUS_OPTIONS),
        list(STATEWIDE_DEFAULT_REVIEW_EXCLUSIONS),
    )


def validate_selected_counties(selected_counties, manifest_lookup):
    """Return one compatible state and metric CRS for selected sources."""
    if not selected_counties:
        raise ValueError("Select at least one county before generating an analysis.")

    states = set()
    metric_crs_values = set()
    for county_name in selected_counties:
        strategy = get_county_source_strategy(county_name, manifest_lookup)
        if strategy == "county_override":
            config = COUNTY_OVERRIDE_CONFIGS[county_name]
            states.add(clean_field(config.get("state")).upper())
            metric_crs_values.add(clean_field(config.get("metric_crs")).upper())
        else:
            states.add(DEFAULT_STATE)
            metric_crs_values.add(DEFAULT_METRIC_CRS)

    if len(states) != 1:
        raise ValueError(
            "The selected counties do not share one state. Multi-state analyses "
            "are not supported by this version."
        )
    if len(metric_crs_values) != 1:
        raise ValueError(
            "The selected counties do not share one compatible analysis CRS."
        )
    return states.pop(), metric_crs_values.pop()


def summarize_selected_county_confidence(selected_counties, manifest_lookup):
    validated = []
    provisional = []
    overrides = []
    confidence_by_county = {}
    disclosure = ""

    for county_name in selected_counties:
        if county_name in COUNTY_OVERRIDE_CONFIGS:
            overrides.append(county_name)
            confidence_by_county[county_name] = "county_specific"
            continue
        row = manifest_lookup[county_name]
        status = clean_field(row.get("coverage_confidence_status")).lower()
        confidence_by_county[county_name] = status or "provisional"
        if status == "validated":
            validated.append(county_name)
        else:
            provisional.append(county_name)
            if not disclosure:
                disclosure = clean_field(row.get("confidence_disclosure"))

    parts = []
    if validated:
        parts.append("Validated statewide coverage: " + ", ".join(validated))
    if provisional:
        parts.append("Provisional statewide coverage: " + ", ".join(provisional))
    if overrides:
        parts.append("County-specific source: " + ", ".join(overrides))
    if provisional and not disclosure:
        disclosure = (
            "Provisional counties passed automated technical validation but "
            "have not been independently compared with county-maintained sources."
        )
    return " | ".join(parts), disclosure, confidence_by_county


def _safe_cache_component(value):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", clean_field(value)).strip("_")


def _sha256_file(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        while True:
            chunk = handle.read(8 * 1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def verify_downloaded_file(path, county_name, expected_size, expected_sha256):
    path = Path(path)
    if not path.is_file():
        raise IOError(f"{county_name} County runtime data was not saved correctly.")
    if expected_size and path.stat().st_size != int(expected_size):
        raise IOError(
            f"{county_name} County runtime data failed its size verification."
        )
    if expected_sha256:
        actual_hash = _sha256_file(path)
        if actual_hash.lower() != clean_field(expected_sha256).lower():
            raise IOError(
                f"{county_name} County runtime data failed its integrity verification."
            )
    return path


@st.cache_resource(show_spinner=False)
def download_runtime_file(
    release_version,
    county_name,
    relative_path,
    expected_size,
    expected_sha256,
):
    """Stream one versioned county file into a deterministic local cache."""
    R2_CACHE_DIRECTORY.mkdir(parents=True, exist_ok=True)
    hash_prefix = clean_field(expected_sha256)[:16] or "nohash"
    filename = (
        f"{_safe_cache_component(release_version)}__"
        f"{_safe_cache_component(county_name)}__{hash_prefix}.parquet"
    )
    destination = R2_CACHE_DIRECTORY / filename

    if destination.exists():
        try:
            return str(
                verify_downloaded_file(
                    destination,
                    county_name,
                    expected_size,
                    expected_sha256,
                )
            )
        except IOError:
            destination.unlink(missing_ok=True)

    url = build_public_r2_url(relative_path)
    temporary = destination.with_name(
        f".{destination.name}.{os.getpid()}.part"
    )
    temporary.unlink(missing_ok=True)
    digest = hashlib.sha256()
    byte_count = 0
    try:
        with requests.get(
            url,
            stream=True,
            timeout=(15, R2_DOWNLOAD_TIMEOUT_SECONDS),
            headers={"User-Agent": "TerritoryToolbox-Analyzer/1.0"},
        ) as response:
            response.raise_for_status()
            with temporary.open("wb") as handle:
                for chunk in response.iter_content(chunk_size=1024 * 1024):
                    if not chunk:
                        continue
                    handle.write(chunk)
                    digest.update(chunk)
                    byte_count += len(chunk)
    except requests.RequestException as exc:
        temporary.unlink(missing_ok=True)
        raise ConnectionError(
            f"TerritoryToolbox could not download {county_name} County data. "
            "Please retry later or report the issue."
        ) from exc
    except OSError as exc:
        temporary.unlink(missing_ok=True)
        raise IOError(
            f"TerritoryToolbox could not save {county_name} County data locally."
        ) from exc

    if expected_size and byte_count != int(expected_size):
        temporary.unlink(missing_ok=True)
        raise IOError(
            f"{county_name} County runtime data failed its size verification."
        )
    if expected_sha256 and digest.hexdigest().lower() != clean_field(expected_sha256).lower():
        temporary.unlink(missing_ok=True)
        raise IOError(
            f"{county_name} County runtime data failed its integrity verification."
        )
    os.replace(temporary, destination)
    return str(destination)


def get_county_runtime_file(release, county_row):
    county_name = clean_field(county_row.get("canonical_county"))
    return Path(
        download_runtime_file(
            clean_field(release.get("release_version")),
            county_name,
            clean_field(county_row.get("published_runtime_relative_path")),
            int(county_row.get("published_runtime_byte_size") or 0),
            clean_field(county_row.get("published_runtime_sha256")),
        )
    )


def _padded_wgs84_bbox(kml_bounds, kml_crs):
    envelope = gpd.GeoSeries([box(*kml_bounds)], crs=kml_crs)
    if envelope.crs is None:
        envelope = envelope.set_crs("EPSG:4326", allow_override=True)
    envelope = envelope.to_crs("EPSG:4326")
    minx, miny, maxx, maxy = envelope.total_bounds
    return (
        float(minx - R2_BBOX_PADDING_DEGREES),
        float(miny - R2_BBOX_PADDING_DEGREES),
        float(maxx + R2_BBOX_PADDING_DEGREES),
        float(maxy + R2_BBOX_PADDING_DEGREES),
    )


def read_runtime_county_bbox(local_path, kml_bounds, kml_crs):
    """Use GeoParquet bbox pushdown when available, with a safe fallback."""
    read_bbox = _padded_wgs84_bbox(kml_bounds, kml_crs)
    try:
        frame = gpd.read_parquet(local_path, bbox=read_bbox)
        bbox_pushdown_used = True
    except Exception as exc:
        message = str(exc).lower()
        expected_bbox_failure = any(
            token in message
            for token in (
                "bbox",
                "covering",
                "unexpected keyword",
                "not supported",
                "geoparquet 1.1",
            )
        ) or isinstance(exc, (TypeError, ValueError, NotImplementedError))
        if not expected_bbox_failure:
            raise
        frame = gpd.read_parquet(local_path)
        if frame.crs is None:
            raise ValueError("The statewide runtime file has no CRS.")
        frame_wgs84 = frame.to_crs("EPSG:4326")
        frame = frame.loc[frame_wgs84.geometry.intersects(box(*read_bbox))].copy()
        bbox_pushdown_used = False
    if frame.crs is None:
        raise ValueError("The statewide runtime file has no CRS.")
    if frame.crs.to_epsg() != 4326:
        frame = frame.to_crs("EPSG:4326")
    return frame, bbox_pushdown_used


def _ensure_internal_columns(frame):
    defaults = {
        "Canonical_HouseSx": "",
        "Canonical_Dir": "",
        "Canonical_StType": "",
        "Canonical_SuffixDir": "",
        "Canonical_UnitType": "",
        "Canonical_Unit": "",
        "Canonical_Full_Address": "",
        "Canonical_Full_House_Number": "",
        "Canonical_Full_Street": "",
        "Canonical_Mailable_Address": "",
        "Canonical_Subaddress": "",
        "Canonical_Postal_City": "",
        "Canonical_ZIP4": "",
        "Canonical_Full_ZIP": "",
        "Canonical_Quality_Flags": "",
        "Canonical_Record_Role": "Standalone Address",
        "Canonical_Occupancy_Category": "",
        "Canonical_Occupancy_Confidence": "",
        "Canonical_Occupancy_Reason": "",
        "Canonical_Analyzer_Handling": "include_standard",
        "Canonical_Exclusion_Category": "none",
        "Canonical_Analyzer_Eligible": True,
        "Potential_Parent_Record": False,
        "Potential_Child_Record": False,
        "Potential_Double_Count_Flag": False,
    }
    for column, default in defaults.items():
        if column not in frame.columns:
            frame[column] = default
    return frame


def normalize_statewide_runtime_source(county_name, county_gdf, analysis_crs):
    """Normalize the stable statewide runtime contract for the legacy engine."""
    frame = county_gdf.reset_index(drop=True).copy()
    frame = _ensure_internal_columns(frame)
    missing_columns = [
        column for column in REQUIRED_CANONICAL_COLUMNS
        if column not in frame.columns
    ]
    if missing_columns:
        raise ValueError(
            f"{county_name} County runtime data is missing required fields: "
            + ", ".join(missing_columns)
        )
    if "Source_Record_ID" not in frame.columns:
        raise ValueError(
            f"{county_name} County runtime data has no stable Source_Record_ID."
        )
    if frame.crs is None:
        raise ValueError(f"{county_name} County runtime data has no CRS.")

    frame["geometry"] = frame.geometry.make_valid()
    frame = frame[frame.geometry.notna() & ~frame.geometry.is_empty].copy()
    frame["Source_County"] = county_name
    frame["Source_State"] = frame.get(
        "Canonical_State",
        pd.Series(DEFAULT_STATE, index=frame.index),
    ).map(clean_field).replace("", DEFAULT_STATE)

    source_ids = frame["Source_Record_ID"].map(clean_field)
    if source_ids.eq("").any() or not source_ids.is_unique:
        raise ValueError(
            f"{county_name} County runtime data contains blank or duplicate source IDs."
        )
    frame["Canonical_Analyzer_Eligible"] = frame[
        "Canonical_Analyzer_Eligible"
    ].fillna(False).astype(bool)
    for boolean_column in (
        "Potential_Parent_Record",
        "Potential_Child_Record",
        "Potential_Double_Count_Flag",
    ):
        frame[boolean_column] = frame[boolean_column].fillna(False).astype(bool)
    return frame.to_crs(analysis_crs)


@st.cache_data(show_spinner=False)
def load_county_override_data(county_name, kml_bounds=None, kml_crs=None):
    """Load the existing county-specific override with an early bbox filter."""
    county_config = COUNTY_OVERRIDE_CONFIGS[county_name]
    county_path = county_config["file_path"]
    read_bbox = None
    if kml_bounds and kml_crs:
        try:
            with fiona.open(county_path) as county_source:
                source_crs = county_source.crs_wkt or county_source.crs
            if source_crs:
                kml_envelope = gpd.GeoSeries(
                    [box(*kml_bounds)],
                    crs=kml_crs,
                ).to_crs(source_crs)
                read_bbox = tuple(float(value) for value in kml_envelope.total_bounds)
        except (ValueError, TypeError, fiona.errors.FionaError):
            read_bbox = None
    if read_bbox is None:
        return gpd.read_file(county_path)
    return gpd.read_file(county_path, bbox=read_bbox)


def normalize_county_override_source(
    county_name,
    county_gdf,
    analysis_crs,
):
    """Normalize Milwaukee's existing county-specific schema."""
    county_config = COUNTY_OVERRIDE_CONFIGS[county_name]
    frame = county_gdf.reset_index(drop=True).copy()
    frame = frame.rename(columns=county_config["column_mapping"], errors="ignore")
    frame = _ensure_internal_columns(frame)
    missing_columns = [
        column for column in REQUIRED_CANONICAL_COLUMNS
        if column not in frame.columns
    ]
    if missing_columns:
        raise ValueError(
            f"{county_name} County data failed preflight validation. Missing "
            "required canonical columns: " + ", ".join(missing_columns)
        )
    if frame.crs is None:
        raise ValueError(f"{county_name} County data has no CRS.")

    frame["geometry"] = frame.geometry.make_valid()
    frame = frame[frame.geometry.notna() & ~frame.geometry.is_empty].copy()
    frame["Source_County"] = county_name
    frame["Source_State"] = county_config["state"]
    frame["Canonical_Analyzer_Eligible"] = True

    county_prefix = normalize_county_prefix(county_name)
    fallback_ids = pd.Series(
        [f"FALLBACK-{row_number:09d}" for row_number in range(1, len(frame) + 1)],
        index=frame.index,
        dtype="string",
    )
    native_ids = frame.get(
        "Canonical_Native_Source_ID",
        pd.Series("", index=frame.index),
    ).map(clean_field)
    source_ids = native_ids.where(native_ids.ne(""), fallback_ids)
    frame["Source_Record_ID"] = county_prefix + "-" + source_ids.astype(str)
    duplicate_ids = frame["Source_Record_ID"].duplicated(keep=False)
    if duplicate_ids.any():
        duplicate_sequence = (frame.groupby("Source_Record_ID").cumcount() + 1).astype(str)
        frame.loc[duplicate_ids, "Source_Record_ID"] = (
            frame.loc[duplicate_ids, "Source_Record_ID"]
            + "-DUP-"
            + duplicate_sequence.loc[duplicate_ids]
        )
    return frame.to_crs(analysis_crs)


def prepare_county_data(
    county_name,
    kml_bounds,
    kml_crs,
    analysis_crs,
    release,
    manifest_lookup,
):
    """Load only the selected county and normalize it to one internal schema."""
    strategy = get_county_source_strategy(county_name, manifest_lookup)
    if strategy == "county_override":
        source = load_county_override_data(
            county_name,
            kml_bounds=kml_bounds,
            kml_crs=kml_crs,
        )
        normalized = normalize_county_override_source(
            county_name,
            source,
            analysis_crs,
        )
        source_description = f"{county_name} County-specific dataset"
        return normalized, source_description, False

    county_row = manifest_lookup[county_name]
    local_path = get_county_runtime_file(release, county_row)
    source, bbox_pushdown_used = read_runtime_county_bbox(
        local_path,
        kml_bounds,
        kml_crs,
    )
    normalized = normalize_statewide_runtime_source(
        county_name,
        source,
        analysis_crs,
    )
    confidence = clean_field(
        county_row.get("coverage_confidence_status")
    ).capitalize()
    source_description = (
        f"Wisconsin NG911 Runtime {release['release_version']} "
        f"({confidence or 'Provisional'})"
    )
    return normalized, source_description, bbox_pushdown_used


@st.cache_data
def inspect_kml_territory_groups(kml_bytes):
    preview_gdf = gpd.read_file(io.BytesIO(kml_bytes), driver="KML")
    preview_gdf = assign_territory_names(preview_gdf)
    preview_gdf = apply_territory_groups(preview_gdf)
    detected_groups = preview_gdf["Detected_Territory_Group"].unique().tolist()
    detected_groups.sort(key=natural_keys)
    return detected_groups, len(preview_gdf["Territory_Name"].unique())


def resolve_overlapping_assignments(
    joined_gdf,
    kml_gdf,
    metric_crs,
    territory_rank,
):
    """Resolve all polygon matches to one final territory per source record.

    Cross-group matches follow this priority: Letter Writing, Residential,
    Business, then all other groups. If several territories remain inside the
    winning group, the address is assigned to the territory whose internal
    representative point is nearest. Natural territory order breaks ties.
    Every rejected match is retained in the Overlap Audit.
    """
    if joined_gdf.empty:
        return joined_gdf.copy(), pd.DataFrame(), 0

    matches = joined_gdf.reset_index(drop=True).copy()
    matches["_Match_ID"] = matches.index
    matches["_Territory_Rank"] = (
        matches["Territory_Name"].map(territory_rank).fillna(float("inf"))
    )
    matches["_Group_Priority"] = matches["Territory_Group"].map(
        territory_group_priority
    )
    group_order = sorted(
        matches["Territory_Group"].map(clean_field).unique(),
        key=natural_keys,
    )
    group_rank = {
        territory_group: rank
        for rank, territory_group in enumerate(group_order)
    }
    matches["_Group_Rank"] = (
        matches["Territory_Group"].map(group_rank).fillna(float("inf"))
    )

    territory_reference_gdf = (
        kml_gdf[["Territory_Name", "geometry_terr"]]
        .dropna(subset=["Territory_Name", "geometry_terr"])
        .set_geometry("geometry_terr")
        .dissolve(by="Territory_Name")
        .to_crs(metric_crs)
    )
    territory_reference_gdf["_Reference_Point"] = (
        territory_reference_gdf.geometry.representative_point()
    )
    reference_lookup = territory_reference_gdf[
        "_Reference_Point"
    ].to_dict()

    match_points_metric = gpd.GeoDataFrame(
        matches.copy(),
        geometry="_join_point",
        crs=joined_gdf.crs,
    ).to_crs(metric_crs)
    reference_points = gpd.GeoSeries(
        matches["Territory_Name"].map(reference_lookup),
        index=matches.index,
        crs=metric_crs,
    )
    matches["_Assignment_Distance"] = (
        match_points_metric.geometry.distance(reference_points, align=False)
    ).fillna(float("inf"))

    ordered_matches = matches.sort_values(
        by=[
            "Source_Record_ID",
            "_Group_Priority",
            "_Group_Rank",
            "_Assignment_Distance",
            "_Territory_Rank",
            "_Match_ID",
        ],
        kind="stable",
    )
    selected_matches = ordered_matches.drop_duplicates(
        subset=["Source_Record_ID"],
        keep="first",
    ).copy()

    source_match_counts = matches.groupby("Source_Record_ID").size()
    overlap_match_count = int(
        (source_match_counts[source_match_counts > 1] - 1).sum()
    )

    additional_info = matches[
        [
            "Source_Record_ID",
            "_Match_ID",
            "Territory_Name",
            "Territory_Group",
            "_Group_Priority",
        ]
    ].rename(
        columns={
            "_Match_ID": "_Additional_Match_ID",
            "Territory_Name": "Additional_Territory",
            "Territory_Group": "Additional_Group",
            "_Group_Priority": "_Additional_Group_Priority",
        }
    )
    overlap_audit_df = selected_matches.merge(
        additional_info,
        on="Source_Record_ID",
        how="inner",
    )
    overlap_audit_df = overlap_audit_df[
        overlap_audit_df["_Match_ID"]
        != overlap_audit_df["_Additional_Match_ID"]
    ].copy()

    if not overlap_audit_df.empty:
        overlap_audit_df["Assigned_Territory"] = overlap_audit_df[
            "Territory_Name"
        ]
        overlap_audit_df["Assigned_Group"] = overlap_audit_df[
            "Territory_Group"
        ]
        same_group_mask = overlap_audit_df["Assigned_Group"].eq(
            overlap_audit_df["Additional_Group"]
        )
        overlap_audit_df["Overlap_Type"] = (
            "Cross-Group Priority Assignment"
        )
        overlap_audit_df.loc[
            same_group_mask,
            "Overlap_Type",
        ] = "Same-Group Review"

        resolution_values = []
        resolution_fields = overlap_audit_df[
            [
                "Assigned_Territory",
                "Additional_Territory",
                "Assigned_Group",
                "Additional_Group",
                "_Group_Priority",
                "_Additional_Group_Priority",
            ]
        ].itertuples(index=False, name=None)
        for (
            assigned_territory,
            additional_territory,
            assigned_group,
            additional_group,
            assigned_priority,
            additional_priority,
        ) in resolution_fields:
            if assigned_group == additional_group:
                resolution_values.append(
                    f"Assigned to {assigned_territory} using the nearest "
                    "territory reference point; "
                    f"{additional_territory} requires map review."
                )
            elif assigned_priority < additional_priority:
                resolution_values.append(
                    f"Assigned to {assigned_territory}. "
                    f"{additional_territory} was not counted because "
                    f"{assigned_group} has higher assignment priority."
                )
            else:
                resolution_values.append(
                    f"Assigned to {assigned_territory}. "
                    f"{additional_territory} was not counted because the "
                    "groups share the same fallback priority and "
                    f"{assigned_group} won the natural group-order "
                    "tie-breaker."
                )
        overlap_audit_df["Resolution"] = resolution_values

    audit_helper_columns = [
        "_Additional_Match_ID",
        "_Additional_Group_Priority",
    ]
    overlap_audit_df = overlap_audit_df.drop(
        columns=audit_helper_columns,
        errors="ignore",
    )

    selected_helper_columns = [
        "_Match_ID",
        "_Territory_Rank",
        "_Group_Priority",
        "_Group_Rank",
        "_Assignment_Distance",
        "_join_point",
        "index_right",
    ]
    selected_matches = selected_matches.drop(
        columns=selected_helper_columns,
        errors="ignore",
    )
    selected_matches = gpd.GeoDataFrame(
        selected_matches,
        geometry="geometry",
        crs=joined_gdf.crs,
    )
    return selected_matches, overlap_audit_df, overlap_match_count


def show_loading_status(placeholder, message=None):
    """Display a loading wheel with messages rotating every five seconds."""
    messages = [
        "Loading the active county data release…",
        "Downloading selected county records…",
        "Matching addresses to territory boundaries…",
        "Building your analysis workbook…",
    ]
    if message in messages:
        start_index = messages.index(message)
        messages = messages[start_index:] + messages[:start_index]

    message_spans = "".join(
        f'<span class="territory-loading-message territory-loading-message-{i}">' 
        f'{text}</span>'
        for i, text in enumerate(messages)
    )
    placeholder.markdown(
        f"""
        <div class="territory-loading-row">
            <span class="territory-loading-wheel"></span>
            <span class="territory-loading-messages">{message_spans}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


# --- 1. CONFIGURATION & UI SETUP ---

st.markdown(
    """
    <style>
    .block-container {
        width: 100%;
        max-width: 980px;
        margin-left: auto;
        margin-right: auto;
        padding-left: 3.5rem;
        padding-right: 3.5rem;
        padding-top: 2.5rem;
    }
    h1 {
        text-align: center;
    }
    .territory-engine-intro {
        text-align: center;
        margin: 0 auto 2rem auto;
        max-width: 760px;
    }
    @media (max-width: 768px) {
        .block-container {
            max-width: 100%;
            padding-left: 1rem;
            padding-right: 1rem;
            padding-top: 1.25rem;
        }
        h1 {
            font-size: 2rem !important;
            line-height: 1.15 !important;
        }
        .territory-engine-intro {
            margin-bottom: 1.5rem;
        }
        .territory-loading-messages {
            min-width: 0 !important;
            width: 100%;
        }
        .territory-loading-message {
            white-space: normal !important;
        }
    }
    .territory-loading-row {
        display: flex;
        align-items: center;
        gap: 0.65rem;
        padding: 0.75rem 1rem;
        border-radius: 0.5rem;
        background: rgba(128, 128, 128, 0.10);
    }
    .territory-loading-wheel {
        width: 1rem;
        height: 1rem;
        border: 0.16rem solid rgba(90, 90, 90, 0.25);
        border-top-color: rgb(90, 90, 90);
        border-radius: 50%;
        animation: territory-spin 0.8s linear infinite;
        flex: 0 0 auto;
    }
    .territory-loading-messages {
        position: relative;
        display: inline-block;
        min-height: 1.5rem;
        min-width: 20rem;
    }
    .territory-loading-message {
        position: absolute;
        inset: 0 auto auto 0;
        opacity: 0;
        animation: territory-message-cycle 20s linear infinite;
        white-space: nowrap;
    }
    .territory-loading-message-0 { animation-delay: 0s; }
    .territory-loading-message-1 { animation-delay: 5s; }
    .territory-loading-message-2 { animation-delay: 10s; }
    .territory-loading-message-3 { animation-delay: 15s; }
    .territory-guidance {
        padding: 0.8rem 0.9rem;
        border: 1px solid color-mix(in srgb, var(--text-color) 18%, transparent);
        border-radius: 0.4rem;
        background: var(--secondary-background-color);
        color: var(--text-color);
        font-size: 0.9rem;
        line-height: 1.5;
    }
    .territory-guidance p {
        margin: 0;
    }
    .territory-guidance p + p {
        margin-top: 0.75rem;
    }
    [data-testid="stMultiSelect"] [data-baseweb="tag"],
    .stMultiSelect [data-baseweb="tag"],
    div[data-baseweb="tag"] {
        background-color: #6B7280 !important;
        border-color: #6B7280 !important;
        color: #FFFFFF !important;
    }
    [data-testid="stMultiSelect"] [data-baseweb="tag"] span,
    [data-testid="stMultiSelect"] [data-baseweb="tag"] svg,
    .stMultiSelect [data-baseweb="tag"] span,
    .stMultiSelect [data-baseweb="tag"] svg,
    div[data-baseweb="tag"] span,
    div[data-baseweb="tag"] svg {
        color: #FFFFFF !important;
        fill: #FFFFFF !important;
    }
    div[data-testid="stDownloadButton"] button {
        background-color: #0D6B31 !important;
        border-color: #0D6B31 !important;
        color: white !important;
    }
    @keyframes territory-spin {
        to { transform: rotate(360deg); }
    }
    @keyframes territory-message-cycle {
        0%, 24.9% { opacity: 1; }
        25%, 100% { opacity: 0; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("TerritoryToolbox's Analysis Engine")
st.markdown(
    """
    <div class="territory-engine-intro">
        Upload your territories KML map to generate a complete, filtered
        address database &amp; analysis.
    </div>
    """,
    unsafe_allow_html=True,
)

try:
    active_release = load_active_release()
    county_manifest = load_county_manifest(
        active_release["release_version"],
        active_release["manifest_path"],
        active_release["runtime_schema_version"],
    )
    manifest_county_lookup = get_manifest_county_lookup(county_manifest)
    county_options = get_available_counties(manifest_county_lookup)
except Exception as metadata_error:
    st.error(
        "TerritoryToolbox could not load the current Wisconsin county data "
        f"release. {metadata_error}"
    )
    st.stop()

st.header("Step 1: Enter Your Analysis Details")
congregation_name = st.text_input(
    "Congregation Name (No Spaces)",
    "ExampleCongregation",
)
default_counties = ["Milwaukee"] if "Milwaukee" in county_options else county_options[:1]
selected_counties = st.multiselect(
    "Counties Included In This Analysis",
    options=county_options,
    default=default_counties,
)

confidence_summary, confidence_disclosure, county_confidence_by_name = (
    summarize_selected_county_confidence(
        selected_counties,
        manifest_county_lookup,
    )
)
if confidence_summary:
    st.caption(confidence_summary)
if confidence_disclosure:
    st.caption(confidence_disclosure)

goal_range = st.selectbox(
    "Goal # of Addresses Per Territory",
    ["25-50", "50-75", "75-100", "100-125", "125-150", "150-175"],
    index=3,
)

selected_excluded_statuses = {}
with st.expander("Advanced Settings"):
    apartment_threshold = st.selectbox(
        "Apartment Grouping Threshold",
        [4, 5, 6],
        index=1,
    )
    for county_name in selected_counties:
        county_excluded_statuses, county_default_exclusions = (
            get_county_exclusion_settings(
                county_name,
                manifest_county_lookup,
            )
        )
        selected_excluded_statuses[county_name] = st.multiselect(
            f"{county_name} Excluded Audit Controls",
            options=county_excluded_statuses,
            default=county_default_exclusions,
            key=f"excluded_audit_controls_{county_name}",
        )

st.header("Step 2: Upload Your Territory Map")
uploaded_kml = st.file_uploader("Upload Territory KML File", type=["kml"])

territory_group_overrides = {}
if uploaded_kml:
    try:
        detected_groups, detected_territory_count = (
            inspect_kml_territory_groups(uploaded_kml.getvalue())
        )
        st.header("Step 3: Confirm Territory Groups")
        st.caption(
            f"Detected {detected_territory_count:,} unique territories."
        )
        st.markdown(
            """
            <div class="territory-guidance">
                <p><strong>Tip:</strong> It’s best practice to analyze your territory
                types (door-to-door/residential, business, letter-writing, etc.)
                using separate KML files.</p>
                <p>Combined files are supported, but if your maps overlap (ex:
                business territories physically overlapping with house-to-house
                territories), the engine will be forced to prioritize where shared
                addresses are assigned.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        group_label_to_internal = {
            "Door-to-Door (Residential)": "Residential",
            "Letter Writing": "Letter Writing",
            "Business": "Business",
            "Other": "Other",
        }
        group_options = list(group_label_to_internal.keys())
        with st.expander("Review Detected Territory Groups", expanded=True):
            for detected_group in detected_groups:
                normalized_detected = clean_field(detected_group).lower()
                if "letter" in normalized_detected:
                    default_label = "Letter Writing"
                elif "business" in normalized_detected:
                    default_label = "Business"
                elif "residential" in normalized_detected:
                    default_label = "Door-to-Door (Residential)"
                else:
                    default_label = "Door-to-Door (Residential)"
                selected_group_label = st.selectbox(
                    f'Territories detected as "{detected_group}" should be considered:',
                    options=group_options,
                    index=group_options.index(default_label),
                    key=(
                        "territory_group_"
                        + re.sub(
                            r"[^A-Za-z0-9]+",
                            "_",
                            detected_group,
                        ).strip("_")
                    ),
                )
                territory_group_overrides[detected_group] = (
                    group_label_to_internal[selected_group_label]
                )
    except Exception as error:
        st.error(f"Unable to inspect territory groups: {error}")

MIN_GOAL, MAX_GOAL = [int(x) for x in goal_range.split("-")]
group_signature = tuple(sorted(territory_group_overrides.items()))
county_signature = tuple(selected_counties)
exclusion_signature = tuple(
    (county_name, tuple(selected_excluded_statuses.get(county_name, [])))
    for county_name in selected_counties
)

# --- 2. DATA LOADING & CACHING ---

# --- ADDRESS BUILDER + NORMALIZATION HELPERS ---
def normalize_house_number(value):
    text = clean_field(value)
    if not text: return ""
    if re.fullmatch(r"[+-]?\d+\.0+", text): return text.split(".", 1)[0]
    return re.sub(r"\s+", " ", text)

def normalize_zip_code(value):
    text = clean_field(value)
    if not text: return ""
    text = re.sub(r"\.0+$", "", text)
    digits = re.sub(r"\D", "", text)
    if not digits: return ""
    if len(digits) == 4: digits = digits.zfill(5)
    elif len(digits) == 8: digits = digits.zfill(9)
    if len(digits) == 5: return digits
    if len(digits) == 9: return f"{digits[:5]}-{digits[5:]}"
    if len(digits) > 9: return f"{digits[:5]}-{digits[5:9]}"
    return digits

def normalize_unit(value, unit_type=None):
    text = clean_field(value)
    explicit_type = clean_field(unit_type)

    if text and re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]

    text = re.sub(r"\s+", " ", text).strip()
    explicit_type = re.sub(r"\s+", " ", explicit_type).strip()

    if explicit_type:
        if not text:
            return explicit_type
        if re.match(rf"^{re.escape(explicit_type)}\b", text, re.IGNORECASE):
            return text
        return f"{explicit_type} {text}"

    if not text:
        return ""

    descriptive_pattern = re.compile(
        r"^(?:apt(?:artment)?|unit|ste|suite|upper|lower|bsmt|basement|"
        r"rear|front|floor|fl|building|bldg|room|rm)\b",
        re.IGNORECASE,
    )
    return text if descriptive_pattern.search(text) else f"Apt {text}"


def combine_house_number(row):
    preferred_house = normalize_house_number(
        row.get("Canonical_Full_House_Number")
    )
    house = preferred_house or normalize_house_number(
        row.get("Canonical_HouseNo")
    )
    suffix = clean_field(row.get("Canonical_HouseSx"))

    if not house or not suffix or house.upper().endswith(suffix.upper()):
        return house

    return f"{house}{suffix}"


def build_canonical_street(row):
    preferred_street = clean_field(row.get("Canonical_Full_Street"))
    if preferred_street:
        return preferred_street
    fields = [
        row.get("Canonical_Dir"),
        row.get("Canonical_Street"),
        row.get("Canonical_StType"),
        row.get("Canonical_SuffixDir"),
    ]
    return " ".join(clean_field(value) for value in fields if clean_field(value))


def normalize_full_address(value):
    text = clean_field(value)
    if not text:
        return ""

    text = re.sub(r"\s+", " ", text)
    return re.sub(r"\s*,\s*", ", ", text).strip(" ,")


def is_usable_full_address(value):
    street_line = normalize_full_address(value).split(",", 1)[0].strip()
    tokens = street_line.split()

    if len(tokens) < 2:
        return False

    return bool(re.search(r"\d", tokens[0])) and bool(
        re.search(r"[A-Za-z]", " ".join(tokens[1:]))
    )


def house_number_sort_parts(value):
    text = normalize_house_number(value).upper()
    if not text:
        return pd.Series([float("inf"), 9, ""])

    compact = re.sub(r"\s+", " ", text).strip()
    grid_match = re.fullmatch(
        r"([NSEW])(\d+)([NSEW])(\d+)([A-Z]?)",
        compact,
    )
    if grid_match:
        first_dir, first_num, second_dir, number, suffix = grid_match.groups()
        text_sort = f"{first_dir}{int(first_num):06d}{second_dir}{suffix}"
        return pd.Series([float(number), 3, text_sort])

    mixed_fraction = re.match(
        r"^(\d+)\s+(\d+)\s*/\s*(\d+)(.*)$",
        compact,
    )
    if mixed_fraction:
        whole, numerator, denominator, suffix = mixed_fraction.groups()
        try:
            numeric_value = int(whole) + float(
                Fraction(int(numerator), int(denominator))
            )
        except (ValueError, ZeroDivisionError):
            numeric_value = float(int(whole))
        return pd.Series([numeric_value, 1, suffix.strip()])

    simple_fraction = re.match(r"^(\d+)\s*/\s*(\d+)(.*)$", compact)
    if simple_fraction:
        numerator, denominator, suffix = simple_fraction.groups()
        try:
            numeric_value = float(Fraction(int(numerator), int(denominator)))
        except (ValueError, ZeroDivisionError):
            numeric_value = float("inf")
        return pd.Series([numeric_value, 1, suffix.strip()])

    numeric_prefix = re.match(r"^(\d+(?:\.\d+)?)(.*)$", compact)
    if numeric_prefix:
        number, suffix = numeric_prefix.groups()
        suffix_rank = 0 if not suffix.strip() else 2
        return pd.Series([float(number), suffix_rank, suffix.strip()])

    return pd.Series([float("inf"), 8, compact])


def build_addresses(row, state):
    full_house_number = combine_house_number(row)
    full_street = build_canonical_street(row)
    parsed_is_usable = bool(full_house_number and full_street)

    fallback = normalize_full_address(row.get("Canonical_Full_Address"))
    fallback_line = fallback.split(",", 1)[0].strip()
    if parsed_is_usable:
        base_line = " ".join([full_house_number, full_street]).strip()
    elif fallback_line:
        base_line = fallback_line
    else:
        base_line = " ".join([full_house_number, full_street]).strip()

    municipality = (
        clean_field(row.get("Canonical_Postal_City"))
        or clean_field(row.get("Canonical_Muni"))
    )
    preferred_zip = (
        clean_field(row.get("Canonical_Full_ZIP"))
        or clean_field(row.get("Canonical_Zip_Code"))
    )
    normalized_zip = normalize_zip_code(preferred_zip)
    locality = ", ".join(part for part in [municipality, state] if part)
    if normalized_zip:
        locality = f"{locality} {normalized_zip}".strip()

    preferred_subaddress = clean_field(row.get("Canonical_Subaddress"))
    unit = preferred_subaddress or normalize_unit(
        row.get("Canonical_Unit"),
        row.get("Canonical_UnitType"),
    )
    mailable_line = " ".join(part for part in [base_line, unit] if part)
    base_address = ", ".join(part for part in [base_line, locality] if part)
    mailable_address = ", ".join(
        part for part in [mailable_line, locality] if part
    )

    preferred_mailable = normalize_full_address(
        row.get("Canonical_Mailable_Address")
    )
    if preferred_mailable and is_usable_full_address(preferred_mailable):
        mailable_address = preferred_mailable

    return pd.Series(
        [base_address, mailable_address],
        index=["Base_Address", "Mailable_Address"],
    )


def normalize_address_key(value):
    """Normalize a complete address for cross-county duplicate review."""
    text = clean_field(value).upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def find_cross_county_duplicates(combined_gdf, state, metric_crs):
    """Flag close, same-address records supplied by different counties."""
    empty_columns = [
        "Mailable Address",
        "Primary Source County",
        "Primary Source Record ID",
        "Additional Source County",
        "Additional Source Record ID",
        "Distance (Meters)",
        "Review Status",
    ]
    combined_gdf = combined_gdf.copy()
    combined_gdf["Cross_County_Duplicate_Flag"] = ""

    if combined_gdf.empty or combined_gdf["Source_County"].nunique() < 2:
        return combined_gdf, pd.DataFrame(columns=empty_columns)

    duplicate_work = combined_gdf.copy()
    duplicate_work[["Base_Address", "Mailable_Address"]] = (
        duplicate_work.apply(
            lambda row: build_addresses(row, state),
            axis=1,
        )
    )
    duplicate_work["_Address_Key"] = duplicate_work[
        "Mailable_Address"
    ].map(normalize_address_key)
    duplicate_work = duplicate_work[
        duplicate_work["_Address_Key"].ne("")
    ].copy()
    duplicate_work = duplicate_work.to_crs(metric_crs)

    duplicate_rows = []
    flagged_source_ids = set()
    seen_pairs = set()
    candidate_keys = duplicate_work.groupby("_Address_Key")[
        "Source_County"
    ].nunique()
    candidate_keys = set(candidate_keys[candidate_keys > 1].index)

    for _, address_group in duplicate_work[
        duplicate_work["_Address_Key"].isin(candidate_keys)
    ].groupby("_Address_Key", sort=False):
        group_records = list(address_group.iterrows())
        for (_, first_row), (_, second_row) in combinations(group_records, 2):
            if first_row["Source_County"] == second_row["Source_County"]:
                continue

            pair_key = tuple(
                sorted(
                    [
                        str(first_row["Source_Record_ID"]),
                        str(second_row["Source_Record_ID"]),
                    ]
                )
            )
            if pair_key in seen_pairs:
                continue

            distance_meters = float(
                first_row.geometry.distance(second_row.geometry)
            )
            if distance_meters > CROSS_COUNTY_DUPLICATE_TOLERANCE_METERS:
                continue

            seen_pairs.add(pair_key)
            flagged_source_ids.update(pair_key)
            primary_row, additional_row = sorted(
                [first_row, second_row],
                key=lambda row: (
                    natural_keys(row["Source_County"]),
                    str(row["Source_Record_ID"]),
                ),
            )
            duplicate_rows.append(
                {
                    "Mailable Address": primary_row["Mailable_Address"],
                    "Primary Source County": primary_row["Source_County"],
                    "Primary Source Record ID": primary_row[
                        "Source_Record_ID"
                    ],
                    "Additional Source County": additional_row[
                        "Source_County"
                    ],
                    "Additional Source Record ID": additional_row[
                        "Source_Record_ID"
                    ],
                    "Distance (Meters)": round(distance_meters, 2),
                    "Review Status": (
                        "Possible cross-county duplicate. Both records were "
                        "retained for manual review."
                    ),
                }
            )

    if flagged_source_ids:
        combined_gdf.loc[
            combined_gdf["Source_Record_ID"].isin(flagged_source_ids),
            "Cross_County_Duplicate_Flag",
        ] = "Possible Cross-County Duplicate"

    duplicate_audit_df = pd.DataFrame(
        duplicate_rows,
        columns=empty_columns,
    )
    if not duplicate_audit_df.empty:
        duplicate_audit_df = duplicate_audit_df.sort_values(
            by=[
                "Mailable Address",
                "Primary Source County",
                "Additional Source County",
                "Primary Source Record ID",
            ],
            kind="stable",
        ).reset_index(drop=True)

    return combined_gdf, duplicate_audit_df


def append_quality_flag(base_flag, additional_flag):
    """Combine quality messages without duplicating text."""
    base_text = clean_field(base_flag)
    additional_text = clean_field(additional_flag)
    if not additional_text:
        return base_text
    if not base_text:
        return additional_text
    if additional_text in base_text.split(" | "):
        return base_text
    return f"{base_text} | {additional_text}"


def evaluate_data_quality(row):
    issues = []
    full_house_number = combine_house_number(row)
    street = clean_field(row.get("Canonical_Street"))
    municipality = clean_field(row.get("Canonical_Muni"))
    zip_code = normalize_zip_code(row.get("Canonical_Zip_Code"))
    base_address = clean_field(row.get("Base_Address"))
    mailable_address = clean_field(row.get("Mailable_Address"))
    fallback = normalize_full_address(row.get("Canonical_Full_Address"))

    parsed_is_usable = bool(full_house_number and street)
    fallback_is_usable = is_usable_full_address(fallback)

    if not street and not fallback_is_usable:
        issues.append("Missing Street")
    if not municipality:
        issues.append("Missing Municipality")
    if not zip_code:
        issues.append("Missing ZIP")
    if full_house_number and re.fullmatch(
        r"[+-]?0+(?:\.0+)?",
        full_house_number,
    ):
        issues.append("Zero House Number")

    zip_is_valid = not zip_code or bool(
        re.fullmatch(r"\d{5}(?:-\d{4})?", zip_code)
    )
    if (
        not (parsed_is_usable or fallback_is_usable)
        or not base_address
        or not mailable_address
        or ",," in base_address
        or ",," in mailable_address
        or not zip_is_valid
    ):
        issues.append("Malformed Address")

    if not parsed_is_usable and fallback_is_usable:
        issues.append("Full Address Fallback Used")
    elif fallback and not parsed_is_usable:
        issues.append("Unusable Full Address Fallback")

    result = " | ".join(issues)
    source_quality_flags = clean_field(row.get("Canonical_Quality_Flags"))
    if source_quality_flags:
        for source_flag in [
            flag.strip()
            for flag in source_quality_flags.split("|")
            if flag.strip()
        ]:
            result = append_quality_flag(result, source_flag)
    return result


def parse_house_number_components(full_house_number):
    text = normalize_house_number(full_house_number).upper()
    if not text:
        return "", "", ""

    text = re.sub(r"\s+", " ", text).strip()
    grid_match = re.fullmatch(
        r"([NSEW]\d+[NSEW])(\d+(?:\s+\d+/\d+|\.\d+)?)([A-Z]?)",
        text,
    )
    if grid_match:
        return grid_match.groups()

    directional_match = re.fullmatch(
        r"([NSEW])\s*(\d+(?:\s+\d+/\d+|\.\d+)?)([A-Z]?)",
        text,
    )
    if directional_match:
        return directional_match.groups()

    standard_match = re.fullmatch(
        r"(\d+(?:\s+\d+/\d+|\.\d+)?)([A-Z]?)",
        text,
    )
    if standard_match:
        house_main, suffix = standard_match.groups()
        return "", house_main, suffix

    numeric_matches = list(
        re.finditer(r"\d+(?:\s+\d+/\d+|\.\d+)?", text)
    )
    if not numeric_matches:
        return "", "", ""

    numeric_match = numeric_matches[-1]
    prefix = text[:numeric_match.start()].strip()
    suffix = text[numeric_match.end():].strip()
    prefix = prefix if re.fullmatch(r"[NSEW](?:\d+[NSEW])?", prefix) else ""
    suffix = suffix if re.fullmatch(r"[A-Z]", suffix) else ""
    return prefix, numeric_match.group(0), suffix


def parse_mailable_address(row, state):
    mailable_address = clean_field(row.get("Mailable_Address"))
    address_parts = [part.strip() for part in mailable_address.split(",")]

    street_line = address_parts[0] if address_parts else ""
    municipality = (
        address_parts[1]
        if len(address_parts) > 1
        else clean_field(row.get("Canonical_Muni"))
    )
    state_zip = address_parts[2] if len(address_parts) > 2 else ""

    state_value = state
    normalized_zip = normalize_zip_code(
        clean_field(row.get("Canonical_Full_ZIP"))
        or clean_field(row.get("Canonical_Zip_Code"))
    )
    zip_code = normalized_zip[:5] if normalized_zip else ""
    zip4_code = normalized_zip.split("-", 1)[1] if "-" in normalized_zip else ""
    state_zip_match = re.fullmatch(
        r"([A-Za-z]{2})(?:\s+(\d{5})(?:-(\d{4}))?)?",
        state_zip,
    )
    if state_zip_match:
        state_value = state_zip_match.group(1).upper()
        zip_code = state_zip_match.group(2) or zip_code
        zip4_code = state_zip_match.group(3) or zip4_code

    unit_type = ""
    unit_value = clean_field(row.get("Canonical_Unit"))
    normalized_unit = clean_field(row.get("Canonical_Subaddress")) or normalize_unit(
        unit_value,
        row.get("Canonical_UnitType"),
    )
    if normalized_unit:
        unit_match = re.match(
            r"^(APT(?:ARTMENT)?|UNIT|STE|SUITE|UPPER|LOWER|BSMT|"
            r"BASEMENT|REAR|FRONT|FLOOR|FL|BUILDING|BLDG|ROOM|RM|"
            r"TRLR|TRAILER|LOT|PH|PENTHOUSE|OFFICE)\b\s*(.*)$",
            normalized_unit,
            re.IGNORECASE,
        )
        if unit_match:
            unit_type = unit_match.group(1).upper()
            unit_value = unit_match.group(2).strip()
        else:
            unit_value = normalized_unit

    street_without_unit = street_line
    if normalized_unit:
        unit_pattern = re.compile(
            rf"\s+{re.escape(normalized_unit)}$",
            re.IGNORECASE,
        )
        street_without_unit = unit_pattern.sub("", street_line).strip()

    full_house_number = combine_house_number(row)
    if not full_house_number:
        house_match = re.match(r"^(\S+)\s+(.*)$", street_without_unit)
        if house_match:
            full_house_number = house_match.group(1)

    house_prefix, house_main, house_suffix = parse_house_number_components(
        full_house_number
    )

    street_prefix = clean_field(row.get("Canonical_Dir")).upper()
    street_name = clean_field(row.get("Canonical_Street"))
    street_type = clean_field(row.get("Canonical_StType")).upper()
    suffix_direction = clean_field(
        row.get("Canonical_SuffixDir")
    ).upper()
    full_street = clean_field(row.get("Canonical_Full_Street")) or " ".join(
        part
        for part in [
            street_prefix,
            street_name,
            street_type,
            suffix_direction,
        ]
        if part
    )

    if not street_name and street_without_unit:
        remaining_street = street_without_unit
        if full_house_number and remaining_street.upper().startswith(
            full_house_number.upper()
        ):
            remaining_street = remaining_street[
                len(full_house_number):
            ].strip()

        street_match = re.fullmatch(
            r"(?:(N|S|E|W|NE|NW|SE|SW)\s+)?(.+?)"
            r"(?:\s+(ST|STREET|AVE|AVENUE|RD|ROAD|BLVD|BOULEVARD|"
            r"DR|DRIVE|LN|LANE|CT|COURT|PL|PLACE|PKWY|PARKWAY|"
            r"HWY|HIGHWAY|WAY|TER|TERRACE|CIR|CIRCLE))?"
            r"(?:\s+(N|S|E|W|NE|NW|SE|SW))?",
            remaining_street,
            re.IGNORECASE,
        )
        if street_match:
            street_prefix = clean_field(street_match.group(1)).upper()
            street_name = clean_field(street_match.group(2))
            street_type = clean_field(street_match.group(3)).upper()
            suffix_direction = clean_field(street_match.group(4)).upper()
            full_street = " ".join(
                part
                for part in [
                    street_prefix,
                    street_name,
                    street_type,
                    suffix_direction,
                ]
                if part
            )

    return pd.Series(
        {
            "FullHouNumber": full_house_number,
            "FullStreet": full_street,
            "Municipality": municipality,
            "State": state_value,
            "ZipCode": zip_code,
            "ZIP4Code": zip4_code,
            "HouseNoPrefix": house_prefix,
            "HouseNoMain": house_main,
            "HouseSx": house_suffix,
            "StreetPrefixDir": street_prefix,
            "StreetName": street_name,
            "StreetType": street_type,
            "UnitType": unit_type,
            "Unit": unit_value,
        }
    )

# --- 3. EXCEL GENERATION ENGINE ---
def generate_excel_report(
    joined_gdf,
    unassigned_gdf,
    overlap_audit_df,
    cross_county_duplicate_df,
    kml_gdf,
    min_goal,
    max_goal,
    cong_name,
    analysis_config,
    apt_threshold,
    selected_excluded_statuses,
    selected_counties,
    county_source_files,
    bounding_record_counts,
    relevant_record_counts,
    assigned_record_counts,
    discarded_record_count,
    kml_filename,
    overlap_match_count,
    unassigned_address_count,
):
    output = io.BytesIO()
    run_timestamp = datetime.datetime.now()
    state = analysis_config["state"]
    metric_crs = analysis_config["metric_crs"]
    territory_order, territory_rank = build_territory_order(kml_gdf)
    territory_group_lookup = (
        kml_gdf[["Territory_Name", "Territory_Group"]]
        .drop_duplicates(subset=["Territory_Name"], keep="first")
        .set_index("Territory_Name")["Territory_Group"]
        .to_dict()
    )

    joined_gdf = joined_gdf.copy()
    joined_gdf["Canonical_Zip_Code"] = joined_gdf[
        "Canonical_Zip_Code"
    ].map(normalize_zip_code)
    joined_gdf[["Base_Address", "Mailable_Address"]] = joined_gdf.apply(
        lambda row: build_addresses(row, state),
        axis=1,
    )
    joined_gdf["Data_Quality_Flag"] = joined_gdf.apply(
        evaluate_data_quality,
        axis=1,
    )
    if "Cross_County_Duplicate_Flag" in joined_gdf.columns:
        joined_gdf["Data_Quality_Flag"] = joined_gdf.apply(
            lambda row: append_quality_flag(
                row.get("Data_Quality_Flag"),
                row.get("Cross_County_Duplicate_Flag"),
            ),
            axis=1,
        )
    flagged_record_count = int(
        joined_gdf["Data_Quality_Flag"].ne("").sum()
    )

    normalized_exclusions_by_county = {
        county_name: {
            clean_field(status).upper()
            for status in selected_excluded_statuses.get(county_name, [])
        }
        for county_name in selected_counties
    }
    joined_gdf["Canonical_Status_Normalized"] = joined_gdf[
        "Canonical_Status"
    ].map(clean_field).str.upper()
    exclusion_mask = pd.Series(False, index=joined_gdf.index)
    for county_name, county_statuses in normalized_exclusions_by_county.items():
        exclusion_mask |= (
            joined_gdf["Source_County"].eq(county_name)
            & joined_gdf["Canonical_Status_Normalized"].isin(
                county_statuses
            )
        )

    if "Canonical_Analyzer_Eligible" in joined_gdf.columns:
        analyzer_eligible = joined_gdf[
            "Canonical_Analyzer_Eligible"
        ].fillna(True).astype(bool)
        exclusion_mask |= ~analyzer_eligible
    if "Potential_Double_Count_Flag" in joined_gdf.columns:
        exclusion_mask |= joined_gdf[
            "Potential_Double_Count_Flag"
        ].fillna(False).astype(bool)

    excluded_gdf = joined_gdf[exclusion_mask].copy()
    valid_gdf = joined_gdf[~exclusion_mask].copy()

    valid_gdf["_Territory_Order"] = (
        valid_gdf["Territory_Name"].map(territory_rank).fillna(float("inf"))
    )
    excluded_gdf["_Territory_Order"] = (
        excluded_gdf["Territory_Name"].map(territory_rank).fillna(float("inf"))
    )

    unique_physical_address_count = int(
        valid_gdf["Source_Record_ID"].nunique()
    )
    territory_assignment_count = len(valid_gdf)
    excluded_address_count = int(
        excluded_gdf["Source_Record_ID"].nunique()
    )

    if overlap_audit_df is None or overlap_audit_df.empty:
        valid_overlap_audit_df = pd.DataFrame()
    else:
        valid_overlap_audit_df = overlap_audit_df.copy()
        valid_overlap_audit_df["Canonical_Status_Normalized"] = (
            valid_overlap_audit_df["Canonical_Status"]
            .map(clean_field)
            .str.upper()
        )
        overlap_exclusion_mask = pd.Series(
            False,
            index=valid_overlap_audit_df.index,
        )
        for county_name, county_statuses in (
            normalized_exclusions_by_county.items()
        ):
            overlap_exclusion_mask |= (
                valid_overlap_audit_df["Source_County"].eq(county_name)
                & valid_overlap_audit_df[
                    "Canonical_Status_Normalized"
                ].isin(county_statuses)
            )
        valid_overlap_audit_df = valid_overlap_audit_df[
            ~overlap_exclusion_mask
        ].copy()
        valid_overlap_audit_df["Canonical_Zip_Code"] = (
            valid_overlap_audit_df["Canonical_Zip_Code"].map(
                normalize_zip_code
            )
        )
        valid_overlap_audit_df[
            ["Base_Address", "Mailable_Address"]
        ] = valid_overlap_audit_df.apply(
            lambda row: build_addresses(row, state),
            axis=1,
        )
        valid_overlap_audit_df["Address"] = valid_overlap_audit_df[
            "Mailable_Address"
        ]

    if valid_overlap_audit_df.empty:
        cross_group_shared_address_count = 0
        shared_across_groups_by_territory = {}
        same_group_overlap_address_count = 0
        same_group_overlaps_by_territory = {}
    else:
        cross_group_audit = valid_overlap_audit_df[
            valid_overlap_audit_df["Overlap_Type"].eq(
                "Cross-Group Priority Assignment"
            )
        ].copy()
        cross_group_shared_address_count = int(
            cross_group_audit["Source_Record_ID"].nunique()
        )
        cross_assigned = cross_group_audit[
            ["Source_Record_ID", "Assigned_Territory"]
        ].rename(columns={"Assigned_Territory": "Territory_Name"})
        cross_additional = cross_group_audit[
            ["Source_Record_ID", "Additional_Territory"]
        ].rename(columns={"Additional_Territory": "Territory_Name"})
        cross_group_involvement = pd.concat(
            [cross_assigned, cross_additional],
            ignore_index=True,
        ).drop_duplicates()
        shared_across_groups_by_territory = (
            cross_group_involvement.groupby("Territory_Name")[
                "Source_Record_ID"
            ]
            .nunique()
            .to_dict()
        )

        same_group_audit = valid_overlap_audit_df[
            valid_overlap_audit_df["Overlap_Type"].eq(
                "Same-Group Review"
            )
        ].copy()
        same_group_overlap_address_count = int(
            same_group_audit["Source_Record_ID"].nunique()
        )
        same_assigned = same_group_audit[
            ["Source_Record_ID", "Assigned_Territory"]
        ].rename(columns={"Assigned_Territory": "Territory_Name"})
        same_additional = same_group_audit[
            ["Source_Record_ID", "Additional_Territory"]
        ].rename(columns={"Additional_Territory": "Territory_Name"})
        same_group_involvement = pd.concat(
            [same_assigned, same_additional],
            ignore_index=True,
        ).drop_duplicates()
        same_group_overlaps_by_territory = (
            same_group_involvement.groupby("Territory_Name")[
                "Source_Record_ID"
            ]
            .nunique()
            .to_dict()
        )

    all_territories_df = pd.DataFrame(
        {
            "Territory_Name": territory_order,
            "Territory_Group": [
                territory_group_lookup.get(name, "Residential")
                for name in territory_order
            ],
        }
    )
    valid_counts = (
        valid_gdf.groupby("Territory_Name")
        .size()
        .reset_index(name="Total_Addresses")
    )
    counts_df = all_territories_df.merge(
        valid_counts,
        on="Territory_Name",
        how="left",
    )
    counts_df["Total_Addresses"] = (
        counts_df["Total_Addresses"].fillna(0).astype(int)
    )
    counts_df["_Territory_Order"] = counts_df["Territory_Name"].map(
        territory_rank
    )

    def get_category(count):
        if count == 0:
            return "No Assigned Addresses"
        if count < min_goal:
            return "Undersized"
        if count <= max_goal:
            return "Ideal"
        return "Oversized"

    counts_df["Category"] = counts_df["Total_Addresses"].apply(
        get_category
    )

    valid_gdf[["NWS_Category", "NWS_Number"]] = valid_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
    valid_gdf["NWS_Category"] = valid_gdf["NWS_Category"].fillna("UNK")
    valid_gdf["NWS_Number"] = valid_gdf["NWS_Number"].fillna("0")

    if not excluded_gdf.empty:
        excluded_gdf[["NWS_Category", "NWS_Number"]] = excluded_gdf["Territory_Name"].astype(str).str.extract(r"^([A-Za-z]+)[-\s]+(.*)$")
        excluded_gdf["NWS_Category"] = excluded_gdf["NWS_Category"].fillna("UNK")
        excluded_gdf["NWS_Number"] = excluded_gdf["NWS_Number"].fillna("0")

    apartment_source = valid_gdf[
        [
            "Territory_Name",
            "Base_Address",
            "Canonical_Unit",
            "Canonical_Occupancy_Category",
            "Canonical_Occupancy_Confidence",
            "Canonical_Occupancy_Reason",
        ]
    ].copy()
    apartment_source["_Unit_Normalized"] = (
        apartment_source["Canonical_Unit"]
        .map(clean_field)
        .str.upper()
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    apartment_source["_Occupancy_Category"] = apartment_source[
        "Canonical_Occupancy_Category"
    ].map(clean_field)
    explicitly_nonresidential_categories = {
        "Commercial Suite or Office",
        "Hotel or Motel Room",
        "Campground Site",
        "Storage Unit",
    }
    apartment_source = apartment_source[
        ~apartment_source["_Occupancy_Category"].isin(
            explicitly_nonresidential_categories
        )
    ].copy()
    apartment_source = apartment_source[
        apartment_source["Base_Address"].map(clean_field).ne("")
    ].copy()
    apartment_source["_Has_Nonblank_Unit"] = apartment_source[
        "_Unit_Normalized"
    ].ne("")
    apartment_source["_Explicit_Residential_Occupancy"] = apartment_source[
        "_Occupancy_Category"
    ].isin(
        {
            "Residential Apartment or Condominium",
            "Residential Side or Duplex Unit",
            "Dormitory Room",
            "Mobile-home or Trailer Site",
        }
    )
    apartment_source["_Unknown_Occupancy"] = apartment_source[
        "_Occupancy_Category"
    ].isin({"", "Unknown Unit or Subaddress"})

    apt_groups = (
        apartment_source.groupby(
            ["Territory_Name", "Base_Address"],
            observed=True,
        )
        .agg(
            **{
                "Source Rows": ("_Unit_Normalized", "size"),
                "Nonblank Unit Rows": ("_Has_Nonblank_Unit", "sum"),
                "Unique Normalized Units": (
                    "_Unit_Normalized",
                    lambda values: values[values.ne("")].nunique(),
                ),
                "_Explicit Residential Evidence": (
                    "_Explicit_Residential_Occupancy",
                    "max",
                ),
                "_Unknown Occupancy Evidence": (
                    "_Unknown_Occupancy",
                    "max",
                ),
            }
        )
        .reset_index()
    )
    apt_groups["Blank Parent Rows"] = (
        apt_groups["Source Rows"] - apt_groups["Nonblank Unit Rows"]
    )
    apt_groups["Duplicate Units"] = (
        apt_groups["Nonblank Unit Rows"]
        - apt_groups["Unique Normalized Units"]
    ).clip(lower=0)
    apt_groups["Reported County Unit Count"] = ""
    apt_groups["Total Units"] = apt_groups["Unique Normalized Units"]

    def get_apartment_confidence(row):
        if row["Unique Normalized Units"] < apt_threshold:
            return "Below Threshold"
        if (
            bool(row["_Explicit Residential Evidence"])
            and row["Duplicate Units"] == 0
            and row["Blank Parent Rows"] <= 1
        ):
            return "High"
        if row["Duplicate Units"] <= 2:
            return "Medium"
        return "Low"

    apt_groups["Apartment Confidence"] = apt_groups.apply(
        get_apartment_confidence,
        axis=1,
    )
    def get_apartment_detection_reason(row):
        if bool(row["_Explicit Residential Evidence"]):
            evidence = "explicit residential occupancy evidence"
        elif bool(row["_Unknown Occupancy Evidence"]):
            evidence = "unit evidence with unresolved occupancy classification"
        else:
            evidence = "county unit evidence"
        return (
            f"{int(row['Unique Normalized Units'])} unique nonblank unit "
            f"identifier(s) met the threshold of {apt_threshold}; {evidence}."
        )

    apt_groups["Detection Reason"] = apt_groups.apply(
        get_apartment_detection_reason,
        axis=1,
    )
    apt_groups = apt_groups[
        apt_groups["Total Units"] >= apt_threshold
    ].copy()
    apt_groups["_Territory_Order"] = (
        apt_groups["Territory_Name"]
        .map(territory_rank)
        .fillna(float("inf"))
    )
    apt_groups = apt_groups.sort_values(
        by=["_Territory_Order", "Base_Address"],
        kind="stable",
    ).reset_index(drop=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        
        # --- TAB 1: DASHBOARD ---
        total_territories = len(counts_df)
        total_addresses = unique_physical_address_count
        largest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmax()] if total_territories > 0 else None
        smallest_terr = counts_df.loc[counts_df["Total_Addresses"].idxmin()] if total_territories > 0 else None
        ideal_pct = (len(counts_df[counts_df["Category"] == "Ideal"]) / total_territories * 100) if total_territories > 0 else 0

        largest_name = largest_terr["Territory_Name"] if largest_terr is not None else ""
        largest_count = largest_terr["Total_Addresses"] if largest_terr is not None else 0
        smallest_name = smallest_terr["Territory_Name"] if smallest_terr is not None else ""
        smallest_count = smallest_terr["Total_Addresses"] if smallest_terr is not None else 0

        dashboard_top = [
            [f"Territory Analysis for {cong_name}"],
            [f"Generated on {run_timestamp.strftime('%Y-%m-%d %H:%M')} by TerritoryToolbox (using the analysis tool)"],
            [""],
            ["Quick Facts:"],
            [f"Total Territories: {total_territories}"],
            [f"Total Valid Addresses: {total_addresses}"],
            [f"Excluded Addresses (See Tab 6): {excluded_address_count}"],
            [f"The largest territory has {largest_count} addresses in it ({largest_name})."],
            [f"The smallest territory has {smallest_count} addresses in it ({smallest_name})."],
            [""],
            [f"Ideal Address Range: ({min_goal}-{max_goal})"],
            [f"About {ideal_pct:.1f}% of territories fall within this range."],
        ]
        pd.DataFrame(dashboard_top).to_excel(
            writer,
            sheet_name="Dashboard",
            index=False,
            header=False,
        )

        fixed_bins = [
            ("Under 25", counts_df["Total_Addresses"] < 25),
            ("25-49", counts_df["Total_Addresses"].between(25, 49, inclusive="both")),
            ("50-74", counts_df["Total_Addresses"].between(50, 74, inclusive="both")),
            ("75-99", counts_df["Total_Addresses"].between(75, 99, inclusive="both")),
            ("100-125", counts_df["Total_Addresses"].between(100, 125, inclusive="both")),
            ("126-150", counts_df["Total_Addresses"].between(126, 150, inclusive="both")),
            ("151-175", counts_df["Total_Addresses"].between(151, 175, inclusive="both")),
            ("Over 175", counts_df["Total_Addresses"] > 175),
        ]

        distribution = []
        for range_label, range_mask in fixed_bins:
            range_rows = counts_df.loc[range_mask]
            range_count = len(range_rows)
            range_categories = range_rows["Category"].dropna().astype(str).unique().tolist()

            if len(range_categories) == 1:
                range_category = range_categories[0]
            elif len(range_categories) > 1:
                range_category = "Mixed"
            else:
                if range_label == "Under 25":
                    r_min, r_max = 0, 24
                elif range_label == "Over 175":
                    r_min, r_max = 176, float("inf")
                else:
                    r_min, r_max = [int(v) for v in range_label.split("-")]

                if r_max < min_goal:
                    range_category = "Undersized"
                elif r_min > max_goal:
                    range_category = "Oversized"
                elif r_min >= min_goal and r_max <= max_goal:
                    range_category = "Ideal"
                else:
                    range_category = "Mixed"
            distribution.append([range_category, range_label, range_count])

        distribution_start = 13
        pd.DataFrame(
            distribution,
            columns=["Category", "Range", "Count"],
        ).to_excel(
            writer,
            sheet_name="Dashboard",
            startrow=distribution_start - 1,
            index=False,
        )

        ws1 = writer.sheets["Dashboard"]
        for column_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws1.column_dimensions[column_letter].width = 18

        ws1["A1"].font = Font(size=20, bold=True, color="0D6B31")
        ws1["A2"].hyperlink = None
        ws1["A2"].font = Font(size=10, italic=True, color="0D6B31")
        ws1["A4"].font = Font(size=13, bold=True)

        bold_inline = InlineFont(b=True)
        ws1["A11"].value = CellRichText(
            TextBlock(bold_inline, "Ideal Address Range"),
            f": ({min_goal}-{max_goal})",
        )
        ws1["A11"].font = Font(size=13)
        ws1["A12"].value = CellRichText(
            "About ",
            TextBlock(bold_inline, f"{ideal_pct:.1f}%"),
            " of territories fall within this range.",
        )

        header_fill = PatternFill(
            start_color="C7CDDB",
            end_color="C7CDDB",
            fill_type="solid",
        )
        for col in range(1, 4):
            ws1.cell(row=distribution_start, column=col).fill = header_fill
            ws1.cell(row=distribution_start, column=col).font = Font(bold=True)

        distribution_end_row = distribution_start + len(distribution)
        for row_number in range(distribution_start + 1, distribution_end_row + 1):
            if ws1.cell(row=row_number, column=1).value == "Ideal":
                for col in range(1, 4):
                    ws1.cell(row=row_number, column=col).font = Font(bold=True)

        features_start = distribution_end_row + 2
        features_fill = PatternFill(
            start_color="0D6B31",
            end_color="0D6B31",
            fill_type="solid",
        )
        ws1.merge_cells(
            start_row=features_start,
            start_column=1,
            end_row=features_start,
            end_column=8,
        )
        features_header = ws1.cell(
            row=features_start,
            column=1,
            value="Features Of This Spreadsheet",
        )
        features_header.font = Font(bold=True, size=13, color="EFEFEF")
        features_header.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        for col in range(1, 9):
            ws1.cell(row=features_start, column=col).fill = features_fill

        feature_instructions = [
            (
                "DASHBOARD",
                "The ",
                " tab displays basic statistics about the territory that was "
                "analyzed. It also displays basic information on how the report "
                "is organized.",
            ),
            (
                "COUNTS",
                "The ",
                " tab organizes territories by size. This is done by 'counting' "
                "workable addresses, not geographical size (although that is a "
                "measured statistic). Through the counts tab, you gain insight "
                "into individual territories' density.",
            ),
            (
                "ADDRESS LIST",
                "The ",
                " tab displays every workable address in your territory. This is "
                "what most of the engine's analysis is based off of! Any "
                "questionable entries are flagged with a data warning.",
            ),
            (
                "APARTMENTS",
                "The ",
                " tab displays every multifamily at/above your apartment grouping "
                "threshold (defaulting @ 5 units = an apartment). Large apartment "
                "units can inflate territories sizes. These units can be turned "
                "into letter writing territory.",
            ),
            (
                "TERRITORY BALANCING",
                "The ",
                " tab provides reduction, consolidation, and border-shift "
                "recommendations for \"balancing\" your territory density. It's a "
                "great launching off spot for remapping territory borders.",
            ),
            (
                "EXCLUDED AUDIT",
                "The ",
                " tab displays addresses that are NOT counted towards your "
                "territory. These are usually addresses of highways, vacant lots, "
                "parks, etc. Addresses right outside your territory borders will "
                "be included here too. This is included for auditing purposes.",
            ),
            (
                "OVERLAP AUDIT",
                "The ",
                " tab displays addresses that matched multiple territories. "
                "Cross-group matches are assigned using the priority order "
                "Letter Writing, Residential, Business, then Other. Conflicting "
                "addresses within the same territory group are assigned once "
                "and flagged for manual review.",
            ),
            (
                "COUNTY DUPLICATE AUDIT",
                "The ",
                " tab flags same-address records supplied by different county "
                "datasets within five meters of one another. These records are "
                "retained and should be reviewed manually.",
            ),
        ]

        for offset, (tab_name, prefix, suffix) in enumerate(
            feature_instructions,
            start=1,
        ):
            row_number = features_start + offset
            ws1.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=8,
            )
            cell = ws1.cell(row=row_number, column=1)
            cell.value = CellRichText(
                prefix,
                TextBlock(bold_inline, tab_name),
                suffix,
            )
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

        final_feature_row = features_start + len(feature_instructions) + 1
        ws1.merge_cells(
            start_row=final_feature_row,
            start_column=1,
            end_row=final_feature_row,
            end_column=8,
        )
        final_feature_text = (
            "The addresses in this analysis, with a little reformatting, can be "
            "added to NWS or other supported programs (Visit "
            "https://territorytoolbox.com for details). It's suggested to export "
            "this file into a program you can easily edit, like excel or google "
            "sheets. That will allow you to expand cells to read easier, create "
            "custom filters to see specific data, and customize the sheet to make "
            "it more legible."
        )
        final_feature_cell = ws1.cell(
            row=final_feature_row,
            column=1,
            value=final_feature_text,
        )
        final_feature_cell.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )
        final_feature_cell.hyperlink = "https://territorytoolbox.com"
        final_feature_cell.font = Font(color="0563C1", underline="single")

        technical_start = final_feature_row + 2
        technical_fill = PatternFill(
            start_color="C7CDDB",
            end_color="C7CDDB",
            fill_type="solid",
        )
        ws1.merge_cells(
            start_row=technical_start,
            start_column=1,
            end_row=technical_start,
            end_column=8,
        )
        technical_header = ws1.cell(
            row=technical_start,
            column=1,
            value="Technical: Run Information",
        )
        technical_header.font = Font(bold=True, size=12)
        for col in range(1, 9):
            ws1.cell(row=technical_start, column=col).fill = technical_fill

        exclusion_summary = " | ".join(
            f"{county_name}: "
            + (
                ", ".join(selected_excluded_statuses.get(county_name, []))
                or "None selected"
            )
            for county_name in selected_counties
        )
        bounding_record_summary = " | ".join(
            f"{county_name}: {bounding_record_counts.get(county_name, 0):,}"
            for county_name in selected_counties
        )
        relevant_record_summary = " | ".join(
            f"{county_name}: {relevant_record_counts.get(county_name, 0):,}"
            for county_name in selected_counties
        )
        assigned_record_summary = " | ".join(
            f"{county_name}: {assigned_record_counts.get(county_name, 0):,}"
            for county_name in selected_counties
        )
        county_source_summary = " | ".join(
            f"{county_name}: {county_source_files[county_name]}"
            for county_name in selected_counties
        )
        possible_cross_county_duplicate_pairs = len(
            cross_county_duplicate_df
        )

        release_version = clean_field(
            analysis_config.get("release_version")
        )
        runtime_schema_version = clean_field(
            analysis_config.get("runtime_schema_version")
        )
        manifest_schema_version = clean_field(
            analysis_config.get("manifest_schema_version")
        )
        county_confidence_summary = clean_field(
            analysis_config.get("county_confidence_summary")
        )
        county_confidence_disclosure = clean_field(
            analysis_config.get("county_confidence_disclosure")
        )

        tech_info = [
            ("Run Timestamp", run_timestamp.strftime("%Y-%m-%d %H:%M")),
            ("Wisconsin NG911 Runtime Release", release_version),
            ("Runtime Schema Version", runtime_schema_version),
            ("Manifest Schema Version", manifest_schema_version),
            ("County Coverage Confidence", county_confidence_summary),
            ("Coverage Disclosure", county_confidence_disclosure or "Not applicable"),
            ("Ideal Address Range Setting", f"{min_goal}-{max_goal} addresses"),
            ("Apartment Grouping Threshold", f"{apt_threshold} units"),
            ("Counties Included", ", ".join(selected_counties)),
            ("County Datasets Loaded", f"{len(selected_counties):,}"),
            ("Records Inside KML Bounding Area", bounding_record_summary),
            (
                "Records Near or Inside Territory Polygons",
                relevant_record_summary,
            ),
            ("Assigned Records by County", assigned_record_summary),
            (
                "Records Discarded Outside Boundary Review Area",
                f"{discarded_record_count:,}",
            ),
            ("Excluded Audit Controls", exclusion_summary),
            (
                "Address Records Assigned to Map",
                f"{joined_gdf['Source_Record_ID'].nunique():,}",
            ),
            (
                "Unique Physical Addresses",
                f"{unique_physical_address_count:,}",
            ),
            (
                "Territory Assignments Created",
                f"{territory_assignment_count:,}",
            ),
            (
                "Addresses Shared Across Groups",
                f"{cross_group_shared_address_count:,}",
            ),
            (
                "Same-Group Overlaps Needing Review",
                f"{same_group_overlap_address_count:,}",
            ),
            ("Excluded Address Count", f"{excluded_address_count:,}"),
            ("Records Flagged with Warnings", f"{flagged_record_count:,}"),
            ("KML Filename", kml_filename),
            ("County Source Files", county_source_summary),
            (
                "County Spatial Read Method",
                clean_field(analysis_config.get("bbox_read_summary")),
            ),
            (
                "Possible Cross-County Duplicate Pairs",
                f"{possible_cross_county_duplicate_pairs:,}",
            ),
            (
                "Overlapping Address-to-Territory Matches",
                f"{overlap_match_count:,}",
            ),
            ("Unassigned Address Records", f"{unassigned_address_count:,}"),
        ]

        for offset, (label, value) in enumerate(tech_info, start=1):
            row_number = technical_start + offset
            ws1.merge_cells(
                start_row=row_number,
                start_column=1,
                end_row=row_number,
                end_column=2,
            )
            ws1.merge_cells(
                start_row=row_number,
                start_column=3,
                end_row=row_number,
                end_column=8,
            )
            label_cell = ws1.cell(row=row_number, column=1, value=label)
            value_cell = ws1.cell(row=row_number, column=3, value=value)
            label_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
            value_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

        assignment_note_row = technical_start + len(tech_info) + 1
        ws1.merge_cells(
            start_row=assignment_note_row,
            start_column=1,
            end_row=assignment_note_row,
            end_column=8,
        )
        assignment_note = ws1.cell(
            row=assignment_note_row,
            column=1,
            value=(
                f"The total territory assignment count is "
                f"{territory_assignment_count:,}. Each physical address is "
                "assigned to one final territory using this priority order: "
                "Letter Writing, Residential, Business, then Other. When an "
                "address matches multiple territories in the winning group, "
                "it is assigned once using the nearest territory reference "
                "point. All competing matches are documented in the Overlap "
                "Audit."
            ),
        )
        assignment_note.alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )
        ws1.row_dimensions[assignment_note_row].height = 60

        ws1.delete_cols(17, 10)

        def add_excel_table(worksheet, dataframe, table_name, show_stripes=False):
            if dataframe.empty: return
            max_row, max_col = dataframe.shape
            table_ref = f"A1:{openpyxl.utils.get_column_letter(max_col)}{max_row + 1}"
            tab = Table(displayName=table_name, ref=table_ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=show_stripes, showColumnStripes=False)
            worksheet.add_table(tab)

        # --- TAB 2: COUNTS ---
        territory_area_df = (
            kml_gdf[["Territory_Name", "geometry_terr"]]
            .dropna(subset=["Territory_Name", "geometry_terr"])
            .set_geometry("geometry_terr")
            .dissolve(by="Territory_Name")
            .to_crs(metric_crs)
        )
        territory_area_df["Size (Sq Acres)"] = (
            territory_area_df.geometry.area * 0.000247105
        ).round(2)
        territory_area_df = territory_area_df[["Size (Sq Acres)"]].reset_index()

        apartment_units_by_territory = (
            apt_groups.groupby("Territory_Name", observed=True)["Total Units"]
            .sum()
            .reset_index(name="# of Apartment Units")
        )
        counts_df_sorted = (
            counts_df.merge(
                territory_area_df,
                on="Territory_Name",
                how="left",
            )
            .merge(
                apartment_units_by_territory,
                on="Territory_Name",
                how="left",
            )
        )
        counts_df_sorted["Size (Sq Acres)"] = (
            counts_df_sorted["Size (Sq Acres)"].fillna(0).round(2)
        )
        counts_df_sorted["# of Apartment Units"] = (
            counts_df_sorted["# of Apartment Units"]
            .fillna(0)
            .astype(int)
        )
        counts_df_sorted["Addresses With Apartments Removed"] = (
            counts_df_sorted["Total_Addresses"]
            - counts_df_sorted["# of Apartment Units"]
        ).clip(lower=0)
        counts_df_sorted["Potential Status"] = counts_df_sorted[
            "Addresses With Apartments Removed"
        ].apply(get_category)

        suggested_actions = {
            ("Oversized", "Ideal"): (
                "Consider turning apartment units into letter writing"
            ),
            ("Ideal", "Undersized"): (
                "This is an apartment heavy territory. If you intend on "
                "transforming apartments into letter writing, consider a "
                "border adjustment to add more door-to-door territory."
            ),
            ("Oversized", "Undersized"): (
                "Consider turning some (but not all) apartment units into "
                "letter writing, or consider a border adjustment"
            ),
            ("Ideal", "Ideal"): "No action needed",
            ("Undersized", "Undersized"): (
                "Consider a border adjustment to add more door-to-door territory"
            ),
            ("Oversized", "Oversized"): (
                "Consider a border adjustment to subtract the amount of "
                "door-to-door needed to cover the territory"
            ),
            ("Undersized", "Ideal"): "This is impossible.",
            ("Ideal", "Oversized"): "This is impossible.",
            ("Undersized", "Oversized"): "This is impossible.",
        }
        def get_suggested_action(row):
            if row["Category"] == "No Assigned Addresses":
                return (
                    "Review this territory. No valid addresses were assigned. "
                    "It may overlap another territory, contain only excluded "
                    "records, or contain no county address points."
                )
            potential_status = row["Potential Status"]
            if potential_status == "No Assigned Addresses":
                potential_status = "Undersized"
            return suggested_actions.get(
                (row["Category"], potential_status),
                "Review territory manually",
            )

        counts_df_sorted["Suggested Action"] = counts_df_sorted.apply(
            get_suggested_action,
            axis=1,
        )
        counts_df_sorted["Shared Across Groups"] = (
            counts_df_sorted["Territory_Name"]
            .map(shared_across_groups_by_territory)
            .fillna(0)
            .astype(int)
        )
        counts_df_sorted["Same-Group Overlaps"] = (
            counts_df_sorted["Territory_Name"]
            .map(same_group_overlaps_by_territory)
            .fillna(0)
            .astype(int)
        )

        def get_overlap_review(row):
            same_group_count = int(row["Same-Group Overlaps"])
            shared_count = int(row["Shared Across Groups"])
            territory_count = int(row["Total_Addresses"])
            high_overlap = same_group_count >= 10 or (
                territory_count > 0
                and same_group_count / territory_count >= 0.10
            )

            if same_group_count and shared_count:
                if high_overlap:
                    return (
                        f"Review {same_group_count} same-group overlap(s); "
                        f"{shared_count} cross-group priority assignment(s)"
                    )
                return (
                    f"{same_group_count} same-group overlap(s) and "
                    f"{shared_count} cross-group assignment(s)"
                )
            if same_group_count:
                if high_overlap:
                    return "High-overlap territory — review recommended"
                return f"Review {same_group_count} same-group overlap(s)"
            if shared_count:
                return f"{shared_count} cross-group priority assignment(s)"
            return "No overlap detected"

        counts_df_sorted["Overlap Review"] = counts_df_sorted.apply(
            get_overlap_review,
            axis=1,
        )
        counts_df_sorted = (
            counts_df_sorted.sort_values(
                by="_Territory_Order",
                kind="stable",
            )
            .rename(
                columns={
                    "Territory_Name": "Territory Name",
                    "Total_Addresses": "Territory Address Count",
                    "Category": "Current Status",
                    "# of Apartment Units": "Apartment Units",
                    "Addresses With Apartments Removed": "Total Count w/o Apts",
                    "Potential Status": "Apartmentless Status",
                }
            )[
                [
                    "Territory Name",
                    "Size (Sq Acres)",
                    "Territory Address Count",
                    "Current Status",
                    "Apartment Units",
                    "Total Count w/o Apts",
                    "Apartmentless Status",
                    "Suggested Action",
                    "Shared Across Groups",
                    "Same-Group Overlaps",
                    "Overlap Review",
                ]
            ]
        )

        counts_df_sorted.to_excel(
            writer,
            sheet_name="Counts",
            index=False,
        )
        ws2 = writer.sheets["Counts"]
        ws2.freeze_panes = "B2"

        counts_widths = {
            "A": 14,
            "B": 14,
            "C": 18,
            "D": 18,
            "E": 14,
            "F": 18,
            "G": 18,
            "H": 92,
            "I": 18,
            "J": 18,
            "K": 38,
        }
        for column_letter, width in counts_widths.items():
            ws2.column_dimensions[column_letter].width = width

        header_fill_counts = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        green_status_fill = PatternFill(
            start_color="C4EFD0",
            end_color="C4EFD0",
            fill_type="solid",
        )
        green_data_fill = PatternFill(
            start_color="E1F2DB",
            end_color="E1F2DB",
            fill_type="solid",
        )
        red_current_status_fill = PatternFill(
            start_color="EA9D9C",
            end_color="EA9D9C",
            fill_type="solid",
        )
        red_potential_status_fill = PatternFill(
            start_color="DFC2D0",
            end_color="DFC2D0",
            fill_type="solid",
        )
        red_data_fill = PatternFill(
            start_color="E3D5DC",
            end_color="E3D5DC",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        bottom_border = Border(
            bottom=Side(style="thin", color="666666")
        )
        overlap_header_fill = PatternFill(
            start_color="C7CDDB",
            end_color="C7CDDB",
            fill_type="solid",
        )
        overlap_warning_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        overlap_left_side = Side(style="medium", color="666666")

        for column_number, cell in enumerate(ws2[1], start=1):
            if column_number >= 9:
                cell.fill = overlap_header_fill
                cell.font = Font(bold=True, color="000000", size=12)
            else:
                cell.fill = header_fill_counts
                cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = bottom_border
        ws2.cell(row=1, column=9).border = Border(
            left=overlap_left_side,
            bottom=Side(style="thin", color="666666"),
        )

        for row_number in range(2, len(counts_df_sorted) + 2):
            current_status = ws2.cell(row=row_number, column=4).value
            potential_status = ws2.cell(row=row_number, column=7).value
            suggested_action = ws2.cell(row=row_number, column=8).value

            alternate_fill = (
                white_fill if row_number % 2 == 0 else stripe_fill
            )
            ws2.cell(row=row_number, column=1).fill = alternate_fill
            ws2.cell(row=row_number, column=2).fill = alternate_fill
            ws2.cell(row=row_number, column=2).number_format = "0.00"

            current_is_ideal = current_status == "Ideal"
            potential_is_ideal = potential_status == "Ideal"

            current_status_fill = (
                green_status_fill
                if current_is_ideal
                else red_current_status_fill
            )
            current_data_fill = (
                green_data_fill
                if current_is_ideal
                else red_data_fill
            )
            potential_status_fill = (
                green_status_fill
                if potential_is_ideal
                else red_potential_status_fill
            )

            for column_number in [3, 5, 6]:
                ws2.cell(
                    row=row_number,
                    column=column_number,
                ).fill = current_data_fill

            ws2.cell(
                row=row_number,
                column=4,
            ).fill = current_status_fill
            ws2.cell(
                row=row_number,
                column=7,
            ).fill = potential_status_fill
            suggested_action_fill = (
                green_data_fill if potential_is_ideal else red_data_fill
            )
            ws2.cell(
                row=row_number,
                column=8,
            ).fill = suggested_action_fill

            for column_number in range(9, 12):
                ws2.cell(
                    row=row_number,
                    column=column_number,
                ).fill = alternate_fill

            same_group_overlap_value = ws2.cell(
                row=row_number,
                column=10,
            ).value
            if same_group_overlap_value not in {None, "", 0}:
                ws2.cell(
                    row=row_number,
                    column=10,
                ).fill = overlap_warning_fill
                ws2.cell(
                    row=row_number,
                    column=11,
                ).fill = overlap_warning_fill

            for column_number in range(1, 12):
                cell = ws2.cell(
                    row=row_number,
                    column=column_number,
                )
                cell.font = Font(
                    bold=column_number == 3,
                    italic=column_number in {5, 6, 7},
                    color="000000",
                )
                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_number in {1, 8, 11}
                        else "center"
                    ),
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = bottom_border

            ws2.cell(row=row_number, column=9).border = Border(
                left=overlap_left_side,
                bottom=Side(style="thin", color="666666"),
            )

        # --- TAB 3: ADDRESS LIST ---
        valid_wgs84 = valid_gdf.to_crs("EPSG:4326")
        valid_gdf["Latitude"] = valid_wgs84.geometry.y.astype(float)
        valid_gdf["Longitude"] = valid_wgs84.geometry.x.astype(float)
        valid_gdf[
            [
                "HouseNum_Sort",
                "HouseNum_Suffix_Rank",
                "HouseNum_Text_Sort",
            ]
        ] = valid_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
        valid_gdf["Unit_Sort"] = (
            valid_gdf["Canonical_Unit"].map(clean_field).str.upper()
        )

        address_list_df = valid_gdf.sort_values(
            by=[
                "_Territory_Order",
                "Canonical_Street",
                "HouseNum_Sort",
                "HouseNum_Suffix_Rank",
                "HouseNum_Text_Sort",
                "Unit_Sort",
            ],
            kind="stable",
        ).copy()

        parsed_address_df = address_list_df.apply(
            lambda row: parse_mailable_address(row, state),
            axis=1,
        )
        address_list_df = pd.concat(
            [address_list_df, parsed_address_df],
            axis=1,
        )

        export_df = address_list_df[
            [
                "Territory_Name",
                "Mailable_Address",
                "FullHouNumber",
                "FullStreet",
                "Municipality",
                "State",
                "ZipCode",
                "ZIP4Code",
                "HouseNoPrefix",
                "HouseNoMain",
                "HouseSx",
                "StreetPrefixDir",
                "StreetName",
                "StreetType",
                "UnitType",
                "Unit",
                "Latitude",
                "Longitude",
                "Source_County",
                "Source_Record_ID",
                "Data_Quality_Flag",
            ]
        ].rename(
            columns={
                "Territory_Name": "Territory Name",
                "Mailable_Address": "Mailable Address",
                "Source_County": "Source County",
                "Source_Record_ID": "Source record ID",
                "Data_Quality_Flag": "Data Quality Flag",
            }
        )

        export_df.to_excel(
            writer,
            sheet_name="Address List",
            index=False,
        )
        ws3 = writer.sheets["Address List"]
        ws3.freeze_panes = "C2"

        address_table = Table(
            displayName="AddressListTable",
            ref=(
                f"A1:{openpyxl.utils.get_column_letter(len(export_df.columns))}"
                f"{len(export_df) + 1}"
            ),
        )
        ws3.add_table(address_table)

        hidden_address_columns = [
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
            "R",
            "S",
            "T",
        ]
        for column_letter in hidden_address_columns:
            ws3.column_dimensions[column_letter].hidden = True

        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 57
        ws3.column_dimensions["Q"].width = 25
        ws3.column_dimensions["R"].width = 25
        ws3.column_dimensions["S"].width = 18
        ws3.column_dimensions["T"].width = 36
        ws3.column_dimensions["U"].width = 38

        header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        quality_warning_fill = PatternFill(
            start_color="EA9F9D",
            end_color="EA9F9D",
            fill_type="solid",
        )
        address_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        source_record_column = (
            export_df.columns.get_loc("Source record ID") + 1
        )
        latitude_column = export_df.columns.get_loc("Latitude") + 1
        longitude_column = export_df.columns.get_loc("Longitude") + 1
        quality_flag_column = (
            export_df.columns.get_loc("Data Quality Flag") + 1
        )

        for row_number in range(2, len(export_df) + 2):
            row_fill = (
                white_fill if row_number % 2 == 0 else stripe_fill
            )
            for column_number in range(1, len(export_df.columns) + 1):
                ws3.cell(
                    row=row_number,
                    column=column_number,
                ).fill = row_fill

            ws3.cell(
                row=row_number,
                column=1,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )
            ws3.cell(
                row=row_number,
                column=2,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            source_id_cell = ws3.cell(
                row=row_number,
                column=source_record_column,
            )
            source_id_cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=False,
            )

            for coordinate_column in [latitude_column, longitude_column]:
                coordinate_cell = ws3.cell(
                    row=row_number,
                    column=coordinate_column,
                )
                if coordinate_cell.value not in {None, ""}:
                    coordinate_cell.value = float(coordinate_cell.value)
                    coordinate_cell.number_format = "0.################"

            quality_flag = ws3.cell(
                row=row_number,
                column=quality_flag_column,
            ).value
            if clean_field(quality_flag):
                for column_number in range(1, len(export_df.columns) + 1):
                    ws3.cell(
                        row=row_number,
                        column=column_number,
                    ).fill = quality_warning_fill

            for column_number in [1, 2, source_record_column]:
                ws3.cell(
                    row=row_number,
                    column=column_number,
                ).border = address_border

        # --- TAB 4: APARTMENTS ---
        if not counts_df.empty:
            category_mapping = counts_df.set_index("Territory_Name")[
                "Category"
            ].to_dict()
            address_count_mapping = counts_df.set_index("Territory_Name")[
                "Total_Addresses"
            ].to_dict()
            apt_groups["Current Territory Status"] = apt_groups[
                "Territory_Name"
            ].map(category_mapping)
            apt_groups["Total Addresses in Territory"] = apt_groups[
                "Territory_Name"
            ].map(address_count_mapping).fillna(0).astype(int)
        else:
            apt_groups["Current Territory Status"] = "Unknown"
            apt_groups["Total Addresses in Territory"] = 0

        apt_groups["_Potential Address Count"] = (
            apt_groups["Total Addresses in Territory"]
            - apt_groups["Total Units"]
        ).clip(lower=0)
        apt_groups["_Potential Status"] = apt_groups[
            "_Potential Address Count"
        ].apply(get_category)

        def get_apartment_action_code(row):
            units = int(row["Total Units"])
            current_status = row["Current Territory Status"]
            potential_status = row["_Potential Status"]
            if units >= 10:
                return "TEN_PLUS"
            if current_status == "Undersized":
                return "CURRENT_UNDERSIZED"
            if potential_status in {
                "Undersized",
                "No Assigned Addresses",
            }:
                return "POTENTIAL_UNDERSIZED"
            if current_status == "Ideal" and potential_status == "Ideal":
                return "IDEAL_TO_IDEAL"
            if current_status == "Oversized" and potential_status == "Oversized":
                return "OVERSIZED_TO_OVERSIZED"
            if current_status == "Oversized" and potential_status == "Ideal":
                return "OVERSIZED_TO_IDEAL"
            return "MANUAL"

        apartment_action_text = {
            "TEN_PLUS": (
                "Ideal for letter writing. This building has 10 or more units, "
                "which is difficult to cover if not easily accessible."
            ),
            "CURRENT_UNDERSIZED": (
                "Keep as door-to-door. The territory is already undersized, "
                "removing address would further shrink it"
            ),
            "POTENTIAL_UNDERSIZED": (
                "Keep as door-to-door if accessible from street level. "
                "Removing these units would shrink the territory to undersized."
            ),
            "IDEAL_TO_IDEAL": (
                "Indifferent. The territory remains in the target range with "
                "or without its units. If a border is adjusted, reconsider."
            ),
            "OVERSIZED_TO_OVERSIZED": (
                "Consider letter writing. Even without these units, the "
                "territory remains oversized. Further adjustments would be needed."
            ),
            "OVERSIZED_TO_IDEAL": (
                "Ideal candidate for letter writing. Removing this building "
                "brings the territory into the target range."
            ),
            "MANUAL": "Review building manually",
        }
        apt_groups["_Action_Code"] = apt_groups.apply(
            get_apartment_action_code,
            axis=1,
        )
        apt_groups["Suggested Action"] = apt_groups["_Action_Code"].map(
            apartment_action_text
        )
        apt_groups["Cell on Address List"] = ""

        if not apt_groups.empty:
            apt_export = apt_groups.rename(
                columns={
                    "Territory_Name": "Territory Name",
                    "Base_Address": "Base Address",
                    "Total Units": "Units",
                }
            )[
                [
                    "Base Address",
                    "Units",
                    "Territory Name",
                    "Total Addresses in Territory",
                    "Current Territory Status",
                    "Suggested Action",
                    "Cell on Address List",
                    "Source Rows",
                    "Nonblank Unit Rows",
                    "Unique Normalized Units",
                    "Blank Parent Rows",
                    "Duplicate Units",
                    "Reported County Unit Count",
                    "Apartment Confidence",
                    "Detection Reason",
                ]
            ]
        else:
            apt_export = pd.DataFrame(
                columns=[
                    "Base Address",
                    "Units",
                    "Territory Name",
                    "Total Addresses in Territory",
                    "Current Territory Status",
                    "Suggested Action",
                    "Cell on Address List",
                    "Source Rows",
                    "Nonblank Unit Rows",
                    "Unique Normalized Units",
                    "Blank Parent Rows",
                    "Duplicate Units",
                    "Reported County Unit Count",
                    "Apartment Confidence",
                    "Detection Reason",
                ]
            )

        apt_export.to_excel(writer, sheet_name="Apartments", index=False)
        ws4 = writer.sheets["Apartments"]
        ws4.freeze_panes = "C2"

        apartment_table = Table(
            displayName="ApartmentsTable",
            ref=(
                f"A1:{openpyxl.utils.get_column_letter(len(apt_export.columns))}"
                f"{len(apt_export) + 1}"
            ),
        )
        ws4.add_table(apartment_table)

        apartment_widths = {
            "A": 43,
            "B": 11,
            "C": 14,
            "D": 18,
            "E": 18,
            "F": 107,
            "G": 14,
            "H": 18,
            "I": 18,
            "J": 18,
            "K": 18,
            "L": 18,
            "M": 18,
            "N": 18,
            "O": 57,
        }
        for column_letter, width in apartment_widths.items():
            ws4.column_dimensions[column_letter].width = width
        for column_letter in ["H", "I", "J", "K", "L", "M", "N", "O"]:
            ws4.column_dimensions[column_letter].hidden = True
        ws4.delete_cols(14, 12)
        apartment_table.ref = f"A1:M{len(apt_export) + 1}"

        header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        apartment_bottom_border = Border(
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        address_row_lookup = {}
        for dataframe_position, (_, address_row) in enumerate(
            address_list_df.iterrows(),
            start=2,
        ):
            base_address_value = clean_field(address_row.get("Base_Address"))
            if base_address_value and base_address_value not in address_row_lookup:
                address_row_lookup[base_address_value] = dataframe_position

        apartment_bold_phrases = {
            "TEN_PLUS": "Ideal for letter writing.",
            "CURRENT_UNDERSIZED": "Keep as door-to-door.",
            "POTENTIAL_UNDERSIZED": (
                "Keep as door-to-door if accessible from street level."
            ),
            "IDEAL_TO_IDEAL": "Indifferent.",
            "OVERSIZED_TO_OVERSIZED": "Consider letter writing.",
            "OVERSIZED_TO_IDEAL": "Ideal candidate for letter writing.",
            "MANUAL": "Review building manually",
        }

        for row_number in range(2, len(apt_export) + 2):
            row_fill = white_fill if row_number % 2 == 0 else stripe_fill
            for column_number in range(1, len(apt_export.columns) + 1):
                cell = ws4.cell(row=row_number, column=column_number)
                cell.fill = row_fill
                cell.border = apartment_bottom_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_number in {1, 3, 6, 13} else "center"
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            base_address = clean_field(ws4.cell(row=row_number, column=1).value)
            units_cell = ws4.cell(row=row_number, column=2)
            units_cell.hyperlink = None
            units_cell.font = Font(color="000000", underline=None)
            if units_cell.value not in {None, ""}:
                units_cell.value = int(units_cell.value)

            address_list_row = address_row_lookup.get(base_address)
            address_cell_reference = ws4.cell(row=row_number, column=7)
            if address_list_row is not None:
                address_cell_reference.value = f"A{address_list_row}"
            address_cell_reference.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=False,
            )

            action_code = apt_groups.iloc[row_number - 2]["_Action_Code"]
            full_text = apartment_action_text[action_code]
            bold_phrase = apartment_bold_phrases[action_code]
            remainder = full_text[len(bold_phrase):]
            ws4.cell(row=row_number, column=6).value = CellRichText(
                [TextBlock(InlineFont(b=True), bold_phrase), remainder]
            )

        # --- TAB 5: TERRITORY BALANCING ---
        balancing_columns = [
            "Territory",
            "Action Type",
            "Target Territory",
            "Priority",
            "Addresses Affected",
            "Projected Statuses",
            "Why",
        ]

        def distance_from_ideal(count, goal_minimum, goal_maximum):
            if count < goal_minimum:
                return goal_minimum - count
            if count > goal_maximum:
                return count - goal_maximum
            return 0

        def status_with_count(count, category_function):
            integer_count = int(count)
            return f"{integer_count} ({category_function(integer_count)})"

        def materially_improved(original_distance, projected_distance):
            if original_distance <= 0:
                return False

            improvement = original_distance - projected_distance
            proportional_improvement = improvement / original_distance

            return improvement >= 10 or proportional_improvement >= 0.25

        def evaluate_spatial_shift(
            source_count,
            target_count,
            goal_minimum,
            goal_maximum,
            category_function,
        ):
            equalizing_shift = max(
                int((source_count - target_count) // 2),
                0,
            )

            if equalizing_shift < 1:
                return None

            minimum_for_source_to_be_ideal = max(
                source_count - goal_maximum,
                0,
            )
            minimum_for_target_to_be_ideal = max(
                goal_minimum - target_count,
                0,
            )
            minimum_to_make_both_ideal = max(
                minimum_for_source_to_be_ideal,
                minimum_for_target_to_be_ideal,
            )

            maximum_that_keeps_source_ideal = max(
                source_count - goal_minimum,
                0,
            )
            maximum_that_keeps_target_ideal = max(
                goal_maximum - target_count,
                0,
            )
            maximum_for_both_ideal = min(
                maximum_that_keeps_source_ideal,
                maximum_that_keeps_target_ideal,
                equalizing_shift,
            )

            if (
                minimum_to_make_both_ideal >= 1
                and minimum_to_make_both_ideal <= maximum_for_both_ideal
            ):
                shift = int(minimum_to_make_both_ideal)
                projected_source = source_count - shift
                projected_target = target_count + shift

                return {
                    "shift": shift,
                    "priority": "High",
                    "projected_source": projected_source,
                    "projected_target": projected_target,
                    "why": (
                        f"Shifting {shift} addresses brings both territories "
                        "into the ideal range."
                    ),
                }

            original_source_distance = distance_from_ideal(
                source_count,
                goal_minimum,
                goal_maximum,
            )
            original_target_distance = distance_from_ideal(
                target_count,
                goal_minimum,
                goal_maximum,
            )

            evaluated_shifts = []

            for shift in range(1, equalizing_shift + 1):
                projected_source = source_count - shift
                projected_target = target_count + shift

                projected_source_status = category_function(
                    projected_source
                )
                projected_target_status = category_function(
                    projected_target
                )

                projected_source_distance = distance_from_ideal(
                    projected_source,
                    goal_minimum,
                    goal_maximum,
                )
                projected_target_distance = distance_from_ideal(
                    projected_target,
                    goal_minimum,
                    goal_maximum,
                )

                source_is_ideal = projected_source_status == "Ideal"
                target_is_ideal = projected_target_status == "Ideal"
                ideal_count = int(source_is_ideal) + int(target_is_ideal)

                source_materially_improved = materially_improved(
                    original_source_distance,
                    projected_source_distance,
                )
                target_materially_improved = materially_improved(
                    original_target_distance,
                    projected_target_distance,
                )

                one_is_ideal = ideal_count == 1
                other_materially_improved = (
                    source_is_ideal and target_materially_improved
                ) or (
                    target_is_ideal and source_materially_improved
                )

                total_distance_improvement = (
                    original_source_distance
                    + original_target_distance
                    - projected_source_distance
                    - projected_target_distance
                )

                if one_is_ideal and other_materially_improved:
                    priority = "Medium"
                    priority_rank = 2
                else:
                    priority = "Low"
                    priority_rank = 1

                evaluated_shifts.append(
                    {
                        "shift": shift,
                        "priority": priority,
                        "priority_rank": priority_rank,
                        "ideal_count": ideal_count,
                        "projected_source": projected_source,
                        "projected_target": projected_target,
                        "projected_source_distance": (
                            projected_source_distance
                        ),
                        "projected_target_distance": (
                            projected_target_distance
                        ),
                        "total_distance_improvement": (
                            total_distance_improvement
                        ),
                    }
                )

            if not evaluated_shifts:
                return None

            evaluated_shifts.sort(
                key=lambda result: (
                    -result["priority_rank"],
                    -result["ideal_count"],
                    -result["total_distance_improvement"],
                    (
                        result["projected_source_distance"]
                        + result["projected_target_distance"]
                    ),
                    result["shift"],
                )
            )

            best_result = evaluated_shifts[0]

            if best_result["priority"] == "Medium":
                best_result["why"] = (
                    f"Shifting {best_result['shift']} addresses brings one "
                    "territory into the ideal range and materially improves "
                    "the other."
                )
            else:
                best_result["why"] = (
                    f"Shifting {best_result['shift']} addresses improves the "
                    "imbalance but does not fully resolve both territories."
                )

            return best_result

        apartment_units_by_territory = (
            apt_groups.groupby(
                "Territory_Name",
                observed=True,
            )["Total Units"]
            .sum()
            .to_dict()
        )

        territory_metrics = counts_df[
            [
                "Territory_Name",
                "Total_Addresses",
                "Category",
            ]
        ].copy()

        territory_metrics["Apartment_Units"] = (
            territory_metrics["Territory_Name"]
            .map(apartment_units_by_territory)
            .fillna(0)
            .astype(int)
        )

        territory_metrics["Potential_Count"] = (
            territory_metrics["Total_Addresses"]
            - territory_metrics["Apartment_Units"]
        ).clip(lower=0)

        territory_metrics["Potential_Status"] = territory_metrics[
            "Potential_Count"
        ].apply(get_category)

        territory_metrics["Shift_Baseline_Count"] = territory_metrics[
            "Total_Addresses"
        ]
        territory_metrics["Resolved"] = False

        territory_metrics = territory_metrics.set_index(
            "Territory_Name"
        )

        balancing_rows = []

        # Phase 1: Reductions.
        for territory_name, territory_row in territory_metrics.iterrows():
            raw_count = int(territory_row["Total_Addresses"])
            current_status = territory_row["Category"]
            apartment_units = int(territory_row["Apartment_Units"])
            potential_count = int(territory_row["Potential_Count"])
            potential_status = territory_row["Potential_Status"]

            if current_status != "Oversized" or apartment_units <= 0:
                continue

            if potential_status == "Ideal":
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Reduction",
                        "Target Territory": "Internal",
                        "Priority": "High",
                        "Addresses Affected": apartment_units,
                        "Projected Statuses": (
                            f"{raw_count} (Oversized with Apartments) -> "
                            f"{potential_count} (Ideal W/O Apartments)"
                        ),
                        "Why": (
                            "Converting all apartments to letter writing "
                            "brings this territory into the ideal range "
                            "without border adjustments."
                        ),
                    }
                )
                territory_metrics.at[
                    territory_name,
                    "Resolved",
                ] = True

            elif potential_status == "Oversized":
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Reduction",
                        "Target Territory": "Internal",
                        "Priority": "Medium",
                        "Addresses Affected": apartment_units,
                        "Projected Statuses": (
                            f"{raw_count} (Oversized) -> "
                            f"{potential_count} (Oversized)"
                        ),
                        "Why": (
                            "Converting all apartments reduces bloat, but "
                            "the territory remains oversized. A border shift "
                            "is still required."
                        ),
                    }
                )
                territory_metrics.at[
                    territory_name,
                    "Shift_Baseline_Count",
                ] = potential_count

            elif potential_status in {
                "Undersized",
                "No Assigned Addresses",
            }:
                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Review Warning",
                        "Target Territory": "Internal",
                        "Priority": "Low",
                        "Addresses Affected": "Review",
                        "Projected Statuses": (
                            f"{raw_count} (Oversized) -> "
                            f"{potential_count} (Undersized)"
                        ),
                        "Why": (
                            "Warning: Converting all apartments drops this "
                            "territory below the minimum goal. To avoid a "
                            "partial conversion (e.g., converting 2 buildings "
                            "but leaving 1), consider a border shift instead."
                        ),
                    }
                )

        terr_geoms = (
            kml_gdf[
                [
                    "Territory_Name",
                    "geometry_terr",
                ]
            ]
            .dropna(
                subset=[
                    "Territory_Name",
                    "geometry_terr",
                ]
            )
            .set_geometry("geometry_terr")
            .dissolve(by="Territory_Name")
        )

        terr_geoms["geometry_terr"] = (
            terr_geoms.geometry.make_valid()
        )

        terr_geoms = terr_geoms[
            terr_geoms.geometry.notna()
            & ~terr_geoms.geometry.is_empty
        ].copy()

        terr_geoms_metric = terr_geoms.to_crs(metric_crs)
        territory_sindex = terr_geoms_metric.sindex

        # Phase 2: Consolidations.
        unresolved_undersized = set(
            territory_metrics.index[
                territory_metrics["Category"].eq("Undersized")
                & ~territory_metrics["Resolved"]
            ]
        )

        consolidation_pairs = set()

        for territory_name in sorted(
            unresolved_undersized,
            key=lambda name: territory_rank.get(name, float("inf")),
        ):
            if (
                territory_name not in terr_geoms_metric.index
                or territory_name not in unresolved_undersized
            ):
                continue

            territory_geom = terr_geoms_metric.at[
                territory_name,
                "geometry_terr",
            ]

            candidate_positions = territory_sindex.query(
                territory_geom.buffer(45.0),
                predicate="intersects",
            )

            for candidate_position in candidate_positions:
                neighbor_name = terr_geoms_metric.index[
                    candidate_position
                ]

                if (
                    neighbor_name == territory_name
                    or neighbor_name not in unresolved_undersized
                ):
                    continue

                pair_key = tuple(
                    sorted(
                        (
                            str(territory_name),
                            str(neighbor_name),
                        )
                    )
                )

                if pair_key in consolidation_pairs:
                    continue

                neighbor_geom = terr_geoms_metric.iloc[
                    candidate_position
                ].geometry_terr

                if territory_geom.distance(neighbor_geom) > 45.0:
                    continue

                combined_count = int(
                    territory_metrics.at[
                        territory_name,
                        "Total_Addresses",
                    ]
                    + territory_metrics.at[
                        neighbor_name,
                        "Total_Addresses",
                    ]
                )

                if get_category(combined_count) != "Ideal":
                    continue

                consolidation_pairs.add(pair_key)

                balancing_rows.append(
                    {
                        "Territory": territory_name,
                        "Action Type": "Consolidation",
                        "Target Territory": neighbor_name,
                        "Priority": "High",
                        "Addresses Affected": "Merge",
                        "Projected Statuses": (
                            f"{territory_name}: "
                            f"{int(territory_metrics.at[territory_name, 'Total_Addresses'])} + "
                            f"{neighbor_name}: "
                            f"{int(territory_metrics.at[neighbor_name, 'Total_Addresses'])} = "
                            f"{combined_count} (Ideal when combined)"
                        ),
                        "Why": (
                            "Merging these adjacent undersized territories "
                            "creates a single ideal territory and reduces "
                            "map bloat."
                        ),
                    }
                )

                territory_metrics.at[
                    territory_name,
                    "Resolved",
                ] = True
                territory_metrics.at[
                    neighbor_name,
                    "Resolved",
                ] = True

                unresolved_undersized.discard(territory_name)
                unresolved_undersized.discard(neighbor_name)
                break

        # Phase 3: Border Shifts.
        unresolved_territories = territory_metrics.index[
            ~territory_metrics["Resolved"]
        ]

        shift_sources = [
            territory_name
            for territory_name in unresolved_territories
            if get_category(
                int(
                    territory_metrics.at[
                        territory_name,
                        "Shift_Baseline_Count",
                    ]
                )
            )
            == "Oversized"
        ]

        seen_shift_pairs = set()

        for source_name in shift_sources:
            if source_name not in terr_geoms_metric.index:
                continue

            source_count = int(
                territory_metrics.at[
                    source_name,
                    "Shift_Baseline_Count",
                ]
            )

            source_geom = terr_geoms_metric.at[
                source_name,
                "geometry_terr",
            ]

            candidate_positions = territory_sindex.query(
                source_geom.buffer(45.0),
                predicate="intersects",
            )

            for candidate_position in candidate_positions:
                target_name = terr_geoms_metric.index[
                    candidate_position
                ]

                if (
                    target_name == source_name
                    or target_name not in unresolved_territories
                ):
                    continue

                target_count = int(
                    territory_metrics.at[
                        target_name,
                        "Shift_Baseline_Count",
                    ]
                )

                if get_category(target_count) != "Undersized":
                    continue

                pair_key = tuple(
                    sorted(
                        (
                            str(source_name),
                            str(target_name),
                        )
                    )
                )

                if pair_key in seen_shift_pairs:
                    continue

                target_geom = terr_geoms_metric.iloc[
                    candidate_position
                ].geometry_terr

                if source_geom.distance(target_geom) > 45.0:
                    continue

                seen_shift_pairs.add(pair_key)

                shift_result = evaluate_spatial_shift(
                    source_count,
                    target_count,
                    min_goal,
                    max_goal,
                    get_category,
                )

                if shift_result is None:
                    continue

                projected_source = int(
                    shift_result["projected_source"]
                )
                projected_target = int(
                    shift_result["projected_target"]
                )

                balancing_rows.append(
                    {
                        "Territory": source_name,
                        "Action Type": "Border Shift",
                        "Target Territory": target_name,
                        "Priority": shift_result["priority"],
                        "Addresses Affected": int(
                            shift_result["shift"]
                        ),
                        "Projected Statuses": (
                            f"{source_name}: "
                            f"{status_with_count(projected_source, get_category)}"
                            " | "
                            f"{target_name}: "
                            f"{status_with_count(projected_target, get_category)}"
                        ),
                        "Why": (
                            "(Candidate territories are within the configured "
                            "45m boundary tolerance). "
                            f"{shift_result['why']}"
                        ),
                    }
                )

        territory_balancing_df = pd.DataFrame(
            balancing_rows,
            columns=balancing_columns,
        )

        count_lookup = (
            counts_df.set_index("Territory_Name")["Total_Addresses"].to_dict()
            if not counts_df.empty
            else {}
        )

        if not territory_balancing_df.empty:
            territory_balancing_df["Target Territory"] = (
                territory_balancing_df["Target Territory"]
                .replace({"Internal": "Internal (Apartments)"})
            )
            territory_balancing_df["OT Count"] = (
                territory_balancing_df["Territory"]
                .map(count_lookup)
                .fillna(0)
                .astype(int)
            )
            territory_balancing_df["TT Count"] = (
                territory_balancing_df["Target Territory"].map(
                    lambda target: (
                        "N/A"
                        if target == "Internal (Apartments)"
                        else int(count_lookup.get(target, 0))
                    )
                )
            )
            territory_balancing_df["_Originating_Sort"] = (
                territory_balancing_df["Territory"]
                .astype(str)
                .map(territory_rank)
                .fillna(float("inf"))
            )
            territory_balancing_df = (
                territory_balancing_df.sort_values(
                    by="_Originating_Sort",
                    kind="stable",
                )
                .drop(columns=["_Originating_Sort"])
                .reset_index(drop=True)
            )
        else:
            territory_balancing_df["OT Count"] = pd.Series(dtype="int64")
            territory_balancing_df["TT Count"] = pd.Series(dtype="object")

        territory_balancing_df = (
            territory_balancing_df.rename(
                columns={
                    "Territory": "Originating Territory",
                    "Target Territory": "Targeted Territory",
                    "Action Type": "Balancing Method",
                    "Addresses Affected": "Est. Addresses Affected",
                    "Why": "Why?/Comments",
                }
            )[
                [
                    "Originating Territory",
                    "OT Count",
                    "Targeted Territory",
                    "TT Count",
                    "Balancing Method",
                    "Priority",
                    "Est. Addresses Affected",
                    "Projected Statuses",
                    "Why?/Comments",
                ]
            ]
        )

        territory_balancing_df.to_excel(
            writer,
            sheet_name="Territory Balancing",
            index=False,
        )

        ws5 = writer.sheets["Territory Balancing"]
        ws5.freeze_panes = "E2"

        balancing_widths = {
            "A": 21,
            "B": 7,
            "C": 21,
            "D": 7,
            "E": 18,
            "F": 14,
            "G": 16,
            "H": 64,
            "I": 100,
        }

        for column_letter, width in balancing_widths.items():
            ws5.column_dimensions[column_letter].width = width

        balancing_header_fill = PatternFill(
            start_color="046A34",
            end_color="046A34",
            fill_type="solid",
        )
        balancing_stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        balancing_white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        high_priority_fill = PatternFill(
            start_color="EA9D9C",
            end_color="EA9D9C",
            fill_type="solid",
        )
        medium_priority_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        balancing_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws5[1]:
            cell.fill = balancing_header_fill
            cell.font = Font(
                bold=True,
                color="EAECEB",
                size=12,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = balancing_border

        review_warning_fill = PatternFill(
            start_color="C8A8A8",
            end_color="C8A8A8",
            fill_type="solid",
        )

        for row_number in range(
            2,
            len(territory_balancing_df) + 2,
        ):
            ws5.row_dimensions[row_number].height = 26.25

            left_fill = (
                balancing_white_fill
                if row_number % 2 == 0
                else balancing_stripe_fill
            )
            right_fill = (
                balancing_stripe_fill
                if row_number % 2 == 0
                else balancing_white_fill
            )

            for column_number in range(1, 10):
                cell = ws5.cell(
                    row=row_number,
                    column=column_number,
                )
                cell.fill = left_fill if column_number in {1, 2} else right_fill
                cell.border = balancing_border
                cell.alignment = Alignment(
                    horizontal=(
                        "right"
                        if column_number in {1, 3}
                        else ("left" if column_number in {8, 9} else "center")
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            black_side = Side(style="thin", color="000000")
            ws5.cell(row=row_number, column=1).border = Border(
                left=black_side,
                top=black_side,
                bottom=black_side,
            )
            ws5.cell(row=row_number, column=2).border = Border(
                right=black_side,
                top=black_side,
                bottom=black_side,
            )
            ws5.cell(row=row_number, column=3).border = Border(
                left=black_side,
                top=black_side,
                bottom=black_side,
            )
            ws5.cell(row=row_number, column=4).border = Border(
                right=black_side,
                top=black_side,
                bottom=black_side,
            )

            balancing_method_cell = ws5.cell(row=row_number, column=5)
            priority_cell = ws5.cell(row=row_number, column=6)
            affected_cell = ws5.cell(row=row_number, column=7)

            if balancing_method_cell.value == "Review Warning":
                balancing_method_cell.fill = review_warning_fill
                balancing_method_cell.font = Font(italic=True)
                affected_cell.fill = review_warning_fill
                affected_cell.font = Font(italic=True)

            if priority_cell.value == "High":
                priority_cell.fill = high_priority_fill
            elif priority_cell.value == "Medium":
                priority_cell.fill = medium_priority_fill

        # --- TAB 6: EXCLUDED AUDIT ---
        audit_frames = []

        if not excluded_gdf.empty:
            excluded_audit = excluded_gdf.copy()
            def build_exclusion_explanation(row):
                if bool(row.get("Potential_Double_Count_Flag", False)):
                    return (
                        "Excluded from territory counts to avoid parent-building "
                        "and child-unit double counting. The record is retained "
                        "for audit review."
                    )
                if not bool(row.get("Canonical_Analyzer_Eligible", True)):
                    category = clean_field(
                        row.get("Canonical_Exclusion_Category")
                    ).replace("_", " ")
                    handling = clean_field(
                        row.get("Canonical_Analyzer_Handling")
                    ).replace("_", " ")
                    reason = category or handling or clean_field(
                        row.get("Canonical_Status")
                    )
                    return (
                        "Excluded by the statewide runtime classification: "
                        f"{reason}."
                    )
                return (
                    f"Excluded due to {clean_field(row.get('Source_County'))} "
                    "category: "
                    f"{clean_field(row.get('Canonical_Status')).upper()}"
                )

            excluded_audit["Exclusion Explanation"] = excluded_audit.apply(
                build_exclusion_explanation,
                axis=1,
            )
            audit_frames.append(excluded_audit)

        if unassigned_gdf is not None and not unassigned_gdf.empty:
            unassigned_audit = unassigned_gdf.copy()
            unassigned_audit["Territory_Name"] = "Unassigned"
            unassigned_audit["Exclusion Explanation"] = (
                "Unassigned: Address falls outside the drawn territory boundary "
                f"but is within the {BOUNDARY_AUDIT_BUFFER_METERS:g}-meter "
                "boundary audit area."
            )
            audit_frames.append(unassigned_audit)

        if audit_frames:
            audit_gdf = pd.concat(audit_frames, ignore_index=True, sort=False)
            audit_gdf["Canonical_Zip_Code"] = audit_gdf[
                "Canonical_Zip_Code"
            ].map(normalize_zip_code)
            audit_gdf[["Base_Address", "Mailable_Address"]] = audit_gdf.apply(
                lambda row: build_addresses(row, state),
                axis=1,
            )
            audit_gdf["Data_Quality_Flag"] = audit_gdf.apply(
                evaluate_data_quality,
                axis=1,
            )
            if "Cross_County_Duplicate_Flag" in audit_gdf.columns:
                audit_gdf["Data_Quality_Flag"] = audit_gdf.apply(
                    lambda row: append_quality_flag(
                        row.get("Data_Quality_Flag"),
                        row.get("Cross_County_Duplicate_Flag"),
                    ),
                    axis=1,
                )
            audit_wgs84 = gpd.GeoDataFrame(
                audit_gdf,
                geometry="geometry",
                crs=joined_gdf.crs,
            ).to_crs("EPSG:4326")
            audit_gdf["Latitude"] = audit_wgs84.geometry.y.astype(float)
            audit_gdf["Longitude"] = audit_wgs84.geometry.x.astype(float)
            audit_gdf[
                [
                    "HouseNum_Sort",
                    "HouseNum_Suffix_Rank",
                    "HouseNum_Text_Sort",
                ]
            ] = audit_gdf["Canonical_HouseNo"].apply(house_number_sort_parts)
            audit_gdf["Unit_Sort"] = (
                audit_gdf["Canonical_Unit"].map(clean_field).str.upper()
            )
            audit_gdf["_Territory_Natural_Sort"] = (
                audit_gdf["Territory_Name"]
                .astype(str)
                .map(territory_rank)
                .fillna(float("inf"))
            )
            audit_gdf = (
                audit_gdf.sort_values(
                    by=[
                        "_Territory_Natural_Sort",
                        "Canonical_Street",
                        "HouseNum_Sort",
                        "HouseNum_Suffix_Rank",
                        "HouseNum_Text_Sort",
                        "Unit_Sort",
                    ],
                    kind="stable",
                )
                .drop(columns=["_Territory_Natural_Sort"])
                .copy()
            )

            parsed_audit_df = audit_gdf.apply(
                lambda row: parse_mailable_address(row, state),
                axis=1,
            )
            audit_gdf = pd.concat([audit_gdf, parsed_audit_df], axis=1)

            export_ex_df = audit_gdf[
                [
                    "Territory_Name",
                    "Mailable_Address",
                    "Source_County",
                    "Canonical_Status",
                    "Exclusion Explanation",
                    "FullHouNumber",
                    "FullStreet",
                    "Municipality",
                    "State",
                    "ZipCode",
                    "ZIP4Code",
                    "HouseNoPrefix",
                    "HouseNoMain",
                    "HouseSx",
                    "StreetPrefixDir",
                    "StreetName",
                    "StreetType",
                    "UnitType",
                    "Unit",
                    "Latitude",
                    "Longitude",
                    "Source_Record_ID",
                    "Data_Quality_Flag",
                ]
            ].rename(
                columns={
                    "Territory_Name": "Territory Name",
                    "Mailable_Address": "Mailable Address",
                    "Source_County": "Source County",
                    "Canonical_Status": "County Status",
                    "Source_Record_ID": "Source record ID",
                    "Data_Quality_Flag": "Data Quality Flag",
                }
            )
        else:
            export_ex_df = pd.DataFrame(
                columns=[
                    "Territory Name",
                    "Mailable Address",
                    "Source County",
                    "County Status",
                    "Exclusion Explanation",
                    "FullHouNumber",
                    "FullStreet",
                    "Municipality",
                    "State",
                    "ZipCode",
                    "ZIP4Code",
                    "HouseNoPrefix",
                    "HouseNoMain",
                    "HouseSx",
                    "StreetPrefixDir",
                    "StreetName",
                    "StreetType",
                    "UnitType",
                    "Unit",
                    "Latitude",
                    "Longitude",
                    "Source record ID",
                    "Data Quality Flag",
                ]
            )

        export_ex_df.to_excel(writer, sheet_name="Excluded Audit", index=False)
        ws6 = writer.sheets["Excluded Audit"]
        ws6.freeze_panes = "D2"

        for column_letter in [
            "F", "G", "H", "I", "J", "K", "L", "M", "N", "O",
            "P", "Q", "R", "S", "T", "U", "V",
        ]:
            ws6.column_dimensions[column_letter].hidden = True

        ws6.column_dimensions["A"].width = 14
        ws6.column_dimensions["B"].width = 57
        ws6.column_dimensions["C"].width = 18
        ws6.column_dimensions["D"].width = 22
        ws6.column_dimensions["E"].width = 64
        ws6.column_dimensions["T"].width = 25
        ws6.column_dimensions["U"].width = 25
        ws6.column_dimensions["V"].width = 36
        ws6.column_dimensions["W"].width = 38

        excluded_header_fill = PatternFill(
            start_color="434343",
            end_color="434343",
            fill_type="solid",
        )
        excluded_stripe_fill = PatternFill(
            start_color="F3F3F3",
            end_color="F3F3F3",
            fill_type="solid",
        )
        excluded_white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid",
        )
        excluded_border = Border(
            left=Side(style="thin", color="999999"),
            right=Side(style="thin", color="999999"),
            top=Side(style="thin", color="999999"),
            bottom=Side(style="thin", color="999999"),
        )

        for cell in ws6[1]:
            cell.fill = excluded_header_fill
            cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = excluded_border

        source_record_column_ex = export_ex_df.columns.get_loc(
            "Source record ID"
        ) + 1
        latitude_column_ex = export_ex_df.columns.get_loc("Latitude") + 1
        longitude_column_ex = export_ex_df.columns.get_loc("Longitude") + 1

        for row_number in range(2, len(export_ex_df) + 2):
            row_fill = (
                excluded_white_fill
                if row_number % 2 == 0
                else excluded_stripe_fill
            )
            for column_number in range(1, len(export_ex_df.columns) + 1):
                cell = ws6.cell(row=row_number, column=column_number)
                cell.fill = row_fill
                cell.border = excluded_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left"
                        if column_number in {1, 2, 3, 4, 5, source_record_column_ex}
                        else "center"
                    ),
                    vertical="center",
                    wrap_text=column_number in {1, 2, 3, 4, 5},
                )

            for coordinate_column in [latitude_column_ex, longitude_column_ex]:
                coordinate_cell = ws6.cell(
                    row=row_number,
                    column=coordinate_column,
                )
                if coordinate_cell.value not in {None, ""}:
                    coordinate_cell.value = float(coordinate_cell.value)
                    coordinate_cell.number_format = "0.################"

        if export_ex_df.empty:
            ws6["A2"] = "This page is intentionally blank. There is nothing to audit."
            ws6["A2"].font = Font(italic=True, color="666666")

        # --- TAB 7: OVERLAP AUDIT ---
        overlap_columns = [
            "Address",
            "Assigned Territory",
            "Additional Territory",
            "Assigned Group",
            "Additional Group",
            "Source County",
            "Overlap Type",
            "Resolution",
            "Source Record ID",
        ]
        if valid_overlap_audit_df.empty:
            overlap_export_df = pd.DataFrame(columns=overlap_columns)
        else:
            overlap_export_df = valid_overlap_audit_df.copy()
            overlap_export_df["_Assigned_Order"] = (
                overlap_export_df["Assigned_Territory"]
                .map(territory_rank)
                .fillna(float("inf"))
            )
            overlap_export_df["_Additional_Order"] = (
                overlap_export_df["Additional_Territory"]
                .map(territory_rank)
                .fillna(float("inf"))
            )
            overlap_export_df = (
                overlap_export_df.sort_values(
                    by=[
                        "_Assigned_Order",
                        "_Additional_Order",
                        "Address",
                        "Source_Record_ID",
                        "Overlap_Type",
                    ],
                    kind="stable",
                )
                .rename(
                    columns={
                        "Assigned_Territory": "Assigned Territory",
                        "Additional_Territory": "Additional Territory",
                        "Assigned_Group": "Assigned Group",
                        "Additional_Group": "Additional Group",
                        "Source_County": "Source County",
                        "Overlap_Type": "Overlap Type",
                        "Source_Record_ID": "Source Record ID",
                    }
                )[overlap_columns]
                .reset_index(drop=True)
            )

        overlap_export_df.to_excel(
            writer,
            sheet_name="Overlap Audit",
            index=False,
        )
        ws7 = writer.sheets["Overlap Audit"]
        ws7.freeze_panes = "A2"

        overlap_widths = {
            "A": 57,
            "B": 22,
            "C": 22,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 30,
            "H": 90,
            "I": 36,
        }
        for column_letter, width in overlap_widths.items():
            ws7.column_dimensions[column_letter].width = width

        for cell in ws7[1]:
            cell.fill = excluded_header_fill
            cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = excluded_border

        same_group_audit_fill = PatternFill(
            start_color="FFF2CC",
            end_color="FFF2CC",
            fill_type="solid",
        )
        for row_number in range(2, len(overlap_export_df) + 2):
            row_fill = (
                excluded_white_fill
                if row_number % 2 == 0
                else excluded_stripe_fill
            )
            overlap_type = ws7.cell(row=row_number, column=7).value
            if overlap_type == "Same-Group Review":
                row_fill = same_group_audit_fill

            for column_number in range(1, 10):
                cell = ws7.cell(row=row_number, column=column_number)
                cell.fill = row_fill
                cell.border = excluded_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_number in {1, 2, 3, 7, 8, 9}
                        else "center"
                    ),
                    vertical="center",
                    wrap_text=column_number in {1, 2, 3, 7, 8},
                )

        filter_end_row = max(len(overlap_export_df) + 1, 1)
        ws7.auto_filter.ref = f"A1:I{filter_end_row}"
        if overlap_export_df.empty:
            ws7["A2"] = "This page is intentionally blank. There is nothing to audit."
            ws7["A2"].font = Font(italic=True, color="666666")

        # --- TAB 8: COUNTY DUPLICATE AUDIT ---
        county_duplicate_columns = [
            "Mailable Address",
            "Primary Source County",
            "Primary Source Record ID",
            "Additional Source County",
            "Additional Source Record ID",
            "Distance (Meters)",
            "Review Status",
        ]
        if cross_county_duplicate_df is None:
            county_duplicate_export = pd.DataFrame(
                columns=county_duplicate_columns
            )
        else:
            county_duplicate_export = cross_county_duplicate_df.reindex(
                columns=county_duplicate_columns
            ).copy()

        county_duplicate_export.to_excel(
            writer,
            sheet_name="County Duplicate Audit",
            index=False,
        )
        ws8 = writer.sheets["County Duplicate Audit"]
        ws8.freeze_panes = "A2"
        county_duplicate_widths = {
            "A": 57,
            "B": 20,
            "C": 38,
            "D": 20,
            "E": 38,
            "F": 18,
            "G": 72,
        }
        for column_letter, width in county_duplicate_widths.items():
            ws8.column_dimensions[column_letter].width = width

        for cell in ws8[1]:
            cell.fill = excluded_header_fill
            cell.font = Font(bold=True, color="EAECEB", size=12)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = excluded_border

        for row_number in range(2, len(county_duplicate_export) + 2):
            row_fill = (
                excluded_white_fill
                if row_number % 2 == 0
                else excluded_stripe_fill
            )
            for column_number in range(1, 8):
                cell = ws8.cell(row=row_number, column=column_number)
                cell.fill = row_fill
                cell.border = excluded_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_number in {1, 2, 3, 4, 5, 7}
                        else "center"
                    ),
                    vertical="center",
                    wrap_text=column_number in {1, 2, 3, 4, 5, 7},
                )
            ws8.cell(row=row_number, column=6).number_format = "0.00"

        county_duplicate_filter_end = max(
            len(county_duplicate_export) + 1,
            1,
        )
        ws8.auto_filter.ref = f"A1:G{county_duplicate_filter_end}"
        if county_duplicate_export.empty:
            ws8["A2"] = "This page is intentionally blank. There is nothing to audit."
            ws8["A2"].font = Font(italic=True, color="666666")

        # --- EXCEL UX POLISH ---
        writer.sheets["Territory Balancing"].freeze_panes = "E2"
        writer.sheets["Excluded Audit"].freeze_panes = "A2"
        writer.sheets["Overlap Audit"].freeze_panes = "A2"
        writer.sheets["County Duplicate Audit"].freeze_panes = "A2"

        writer.sheets["Dashboard"].sheet_properties.tabColor = "1E90FF"
        writer.sheets["Counts"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Address List"].sheet_properties.tabColor = "32CD32"
        writer.sheets["Apartments"].sheet_properties.tabColor = "FF8C00"
        writer.sheets["Territory Balancing"].sheet_properties.tabColor = "FF0000"
        writer.sheets["Excluded Audit"].sheet_properties.tabColor = "808080"
        writer.sheets["Overlap Audit"].sheet_properties.tabColor = "808080"
        writer.sheets["County Duplicate Audit"].sheet_properties.tabColor = "808080"

    output.seek(0)
    return output

# --- 4. EXECUTION FLOW ---
if "last_uploaded_kml" not in st.session_state:
    st.session_state["last_uploaded_kml"] = None
if "last_group_signature" not in st.session_state:
    st.session_state["last_group_signature"] = None
if "last_county_signature" not in st.session_state:
    st.session_state["last_county_signature"] = None
if "last_exclusion_signature" not in st.session_state:
    st.session_state["last_exclusion_signature"] = None
if "last_settings_signature" not in st.session_state:
    st.session_state["last_settings_signature"] = None

settings_signature = (
    congregation_name,
    goal_range,
    apartment_threshold,
    active_release["release_version"],
    active_release["runtime_schema_version"],
    active_release["manifest_schema_version"],
)

inputs_changed = (
    uploaded_kml != st.session_state["last_uploaded_kml"]
    or group_signature != st.session_state["last_group_signature"]
    or county_signature != st.session_state["last_county_signature"]
    or exclusion_signature
    != st.session_state["last_exclusion_signature"]
    or settings_signature != st.session_state["last_settings_signature"]
)
if inputs_changed:
    if "excel_data" in st.session_state:
        del st.session_state["excel_data"]
    st.session_state["last_uploaded_kml"] = uploaded_kml
    st.session_state["last_group_signature"] = group_signature
    st.session_state["last_county_signature"] = county_signature
    st.session_state["last_exclusion_signature"] = exclusion_signature
    st.session_state["last_settings_signature"] = settings_signature

if uploaded_kml and "excel_data" not in st.session_state:
    if st.button("Generate Territory Analysis"):
        status_placeholder = st.empty()
        show_loading_status(
            status_placeholder,
            "Loading the active county data release…",
        )

        try:
            analysis_state, analysis_crs = validate_selected_counties(
                selected_counties,
                manifest_county_lookup,
            )
            analysis_config = {
                "state": analysis_state,
                "metric_crs": analysis_crs,
                "release_version": active_release["release_version"],
                "runtime_schema_version": active_release[
                    "runtime_schema_version"
                ],
                "manifest_schema_version": active_release[
                    "manifest_schema_version"
                ],
                "county_confidence_summary": confidence_summary,
                "county_confidence_disclosure": confidence_disclosure,
                "county_confidence_by_name": county_confidence_by_name,
            }

            kml_gdf = gpd.read_file(
                io.BytesIO(uploaded_kml.getvalue()),
                driver="KML",
            )
            if kml_gdf.crs is None:
                kml_gdf = kml_gdf.set_crs(
                    "EPSG:4326",
                    allow_override=True,
                )
            kml_gdf = kml_gdf.copy()
            kml_gdf["geometry"] = kml_gdf.geometry.make_valid()
            kml_gdf = kml_gdf[
                kml_gdf.geometry.notna() & ~kml_gdf.geometry.is_empty
            ].copy()
            if kml_gdf.empty:
                raise ValueError(
                    "The uploaded KML contains no usable territory geometry."
                )
            kml_gdf = assign_territory_names(kml_gdf)
            kml_gdf = apply_territory_groups(
                kml_gdf,
                territory_group_overrides,
            )

            kml_bounds = tuple(float(value) for value in kml_gdf.total_bounds)
            kml_crs = str(kml_gdf.crs)
            county_frames = []
            county_record_counts = {}
            county_source_files = {}

            bbox_pushdown_by_county = {}
            for county_name in selected_counties:
                show_loading_status(status_placeholder)
                try:
                    county_gdf, source_description, bbox_pushdown_used = (
                        prepare_county_data(
                            county_name=county_name,
                            kml_bounds=kml_bounds,
                            kml_crs=kml_crs,
                            analysis_crs=analysis_crs,
                            release=active_release,
                            manifest_lookup=manifest_county_lookup,
                        )
                    )
                except Exception as county_error:
                    raise RuntimeError(
                        f"{county_name} County load failure: "
                        f"{type(county_error).__name__}: {county_error}"
                    ) from county_error
                county_record_counts[county_name] = len(county_gdf)
                county_source_files[county_name] = source_description
                bbox_pushdown_by_county[county_name] = bbox_pushdown_used
                county_frames.append(county_gdf)

            if not county_frames:
                raise ValueError("No county datasets were selected or loaded.")

            parcel_gdf = gpd.GeoDataFrame(
                pd.concat(
                    county_frames,
                    ignore_index=True,
                    sort=False,
                ),
                geometry="geometry",
                crs=analysis_crs,
            )
            if parcel_gdf.empty:
                raise ValueError(
                    "No county address records intersect the uploaded KML "
                    "extent."
                )

            kml_gdf = kml_gdf.to_crs(analysis_crs)
            territory_union = gpd.GeoSeries(
                [kml_gdf.geometry.union_all()],
                crs=analysis_crs,
            ).make_valid().iloc[0]
            if territory_union.is_empty:
                raise ValueError(
                    "The uploaded KML contains no usable combined territory geometry."
                )
            territory_envelope = territory_union.envelope
            parcel_gdf = parcel_gdf[
                parcel_gdf.geometry.intersects(territory_envelope)
            ].copy()
            bounding_record_counts = {
                county_name: int(
                    parcel_gdf["Source_County"].eq(county_name).sum()
                )
                for county_name in selected_counties
            }

            # Compute representative points once, then retain only records inside
            # the exact territory union or its small boundary-review buffer.
            parcel_gdf["_join_point"] = (
                parcel_gdf.geometry.representative_point()
            )
            territory_review_area = territory_union.buffer(
                BOUNDARY_AUDIT_BUFFER_METERS
            )
            relevant_mask = parcel_gdf["_join_point"].covered_by(
                territory_review_area
            )
            relevant_gdf = parcel_gdf[relevant_mask].copy()
            discarded_record_count = int((~relevant_mask).sum())
            relevant_record_counts = {
                county_name: int(
                    relevant_gdf["Source_County"].eq(county_name).sum()
                )
                for county_name in selected_counties
            }

            inside_mask = relevant_gdf["_join_point"].covered_by(
                territory_union
            )
            assignment_candidates_gdf = relevant_gdf[inside_mask].copy()
            unassigned_gdf = relevant_gdf[~inside_mask].copy()
            unassigned_address_count = len(unassigned_gdf)

            relevant_gdf, cross_county_duplicate_df = (
                find_cross_county_duplicates(
                    relevant_gdf,
                    state=analysis_state,
                    metric_crs=analysis_crs,
                )
            )
            duplicate_flag_lookup = relevant_gdf.set_index(
                "Source_Record_ID"
            )["Cross_County_Duplicate_Flag"]
            assignment_candidates_gdf["Cross_County_Duplicate_Flag"] = (
                assignment_candidates_gdf["Source_Record_ID"]
                .map(duplicate_flag_lookup)
                .fillna("")
            )
            unassigned_gdf["Cross_County_Duplicate_Flag"] = (
                unassigned_gdf["Source_Record_ID"]
                .map(duplicate_flag_lookup)
                .fillna("")
            )

            show_loading_status(status_placeholder)
            parcel_join_gdf = assignment_candidates_gdf.set_geometry(
                "_join_point"
            )
            kml_gdf = kml_gdf.rename(
                columns={"geometry": "geometry_terr"}
            ).set_geometry("geometry_terr")

            joined_gdf = gpd.sjoin(
                parcel_join_gdf,
                kml_gdf[
                    [
                        "Territory_Name",
                        "Territory_Group",
                        "geometry_terr",
                    ]
                ],
                how="inner",
                predicate="covered_by",
            )
            joined_gdf = joined_gdf.dropna(
                subset=["Territory_Name"]
            ).copy()
            assigned_record_counts = {
                county_name: int(
                    joined_gdf.loc[
                        joined_gdf["Source_County"].eq(county_name),
                        "Source_Record_ID",
                    ].nunique()
                )
                for county_name in selected_counties
            }

            show_loading_status(status_placeholder)
            _, territory_rank = build_territory_order(kml_gdf)
            joined_gdf, overlap_audit_df, overlap_match_count = (
                resolve_overlapping_assignments(
                    joined_gdf,
                    kml_gdf,
                    analysis_crs,
                    territory_rank,
                )
            )

            show_loading_status(status_placeholder)

            bbox_read_summary_parts = []
            for county_name in selected_counties:
                strategy = get_county_source_strategy(
                    county_name,
                    manifest_county_lookup,
                )
                if strategy == "county_override":
                    bbox_read_summary_parts.append(
                        f"{county_name}: county-source bbox filter"
                    )
                elif bbox_pushdown_by_county.get(county_name):
                    bbox_read_summary_parts.append(
                        f"{county_name}: GeoParquet bbox pushdown"
                    )
                else:
                    bbox_read_summary_parts.append(
                        f"{county_name}: safe full-read fallback"
                    )
            analysis_config["bbox_read_summary"] = " | ".join(
                bbox_read_summary_parts
            )

            excel_file = generate_excel_report(
                joined_gdf,
                unassigned_gdf,
                overlap_audit_df,
                cross_county_duplicate_df,
                kml_gdf,
                MIN_GOAL,
                MAX_GOAL,
                congregation_name.replace(" ", ""),
                analysis_config,
                apt_threshold=apartment_threshold,
                selected_excluded_statuses=selected_excluded_statuses,
                selected_counties=selected_counties,
                county_source_files=county_source_files,
                bounding_record_counts=bounding_record_counts,
                relevant_record_counts=relevant_record_counts,
                assigned_record_counts=assigned_record_counts,
                discarded_record_count=discarded_record_count,
                kml_filename=uploaded_kml.name,
                overlap_match_count=overlap_match_count,
                unassigned_address_count=unassigned_address_count,
            )

            safe_congregation_name = re.sub(
                r"[^A-Za-z0-9_-]+",
                "",
                congregation_name.replace(" ", ""),
            ) or "Congregation"
            filename = (
                f"{safe_congregation_name}-TerritoryAnalysis-"
                f"{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
            )
            st.session_state["excel_data"] = excel_file.getvalue()
            st.session_state["excel_filename"] = filename

            status_placeholder.success(
                "Analysis Complete! Download the generated file below"
            )

        except Exception as error:
            status_placeholder.empty()
            st.error(f"An error occurred during processing: {error}")

if "excel_data" in st.session_state:
    st.download_button(
        label="Download Your Analysis File!",
        data=st.session_state["excel_data"],
        file_name=st.session_state["excel_filename"],
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

st.markdown(
    "<div style='margin-top: 2rem;'><a href='https://territorytoolbox.com' "
    "target='_blank' rel='noopener noreferrer'>Back to TerritoryToolbox</a></div>",
    unsafe_allow_html=True,
)
